import io
import json
import base64
import re
from datetime import datetime, timedelta, timezone
from flask import send_file
from PIL import Image as PilImage
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.lib.utils import ImageReader


DEFAULT_TZ_OFFSET = 7  # WIB


def _format_dt(val, offset=DEFAULT_TZ_OFFSET):
    """Format ISO datetime string to human-readable local time."""
    if not val:
        return "-"
    try:
        if isinstance(val, str):
            dt = datetime.fromisoformat(val.replace("Z", "+00:00"))
        else:
            dt = val
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        dt = dt.astimezone(timezone(timedelta(hours=offset)))
        return dt.strftime("%d %b %Y %H:%M")
    except (ValueError, TypeError):
        return str(val)[:19]


def _nisn_val(answers, fallback):
    """Extract readable NISN or fallback to student UUID."""
    if isinstance(answers, str):
        return fallback
    nisn = (answers or {}).get("_nisn", "")
    if nisn and nisn.replace("?", "") and not nisn.startswith("?" * 8):
        return nisn
    return fallback


def _parse_answer(ans_data, q_idx, q_type, answer_key):
    if ans_data is None:
        return "", ""
    if isinstance(ans_data, str):
        return ans_data, ""
    if isinstance(ans_data, dict):
        ans = ans_data.get("answer", "")
        pages = ans_data.get("pages", {}) or {}
        text_parts = []
        has_canvas = False
        for p_data in pages.values():
            if isinstance(p_data, dict):
                if p_data.get("canvas"):
                    has_canvas = True
                for tb in (p_data.get("textBoxes") or []):
                    if isinstance(tb, dict) and tb.get("text", "").strip():
                        text_parts.append(tb["text"].strip())
        text_content = " | ".join(text_parts) if text_parts else ""
        return ans, text_content
    return str(ans_data), ""


def _get_answer_key(exam, q_idx):
    """Get correct answer for a question."""
    if not exam:
        return ""
    ak = exam.get("answer_key") or {}
    if isinstance(ak, str):
        try:
            ak = json.loads(ak)
        except (json.JSONDecodeError, TypeError):
            ak = {}
    return ak.get(str(q_idx), "")


def export_to_xlsx(submissions: list, exam: dict = None) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hasil"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    wrap_align = Alignment(horizontal="center", vertical="top", wrap_text=True)

    q_types = exam.get("question_types") or {}
    if isinstance(q_types, str):
        try:
            q_types = json.loads(q_types)
        except (json.JSONDecodeError, TypeError):
            q_types = {}
    total_q = exam.get("total_questions", 0)

    # Headers
    headers = ["No", "Nama Siswa", "NISN / ID", "Skor MCQ", "Penalti", "Nilai Final", "Status", "Waktu"]
    for qi in range(total_q):
        qtype = q_types.get(str(qi), "mcq")
        label = f"Soal {qi+1}"
        if qtype != "mcq":
            label += " (Esai)"
        headers.append(label)
        if qtype != "mcq":
            headers.append(f"Soal {qi+1} Tulisan")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for i, s in enumerate(submissions, 1):
        answers = s.get("answers") or {}
        if isinstance(answers, str):
            try:
                answers = json.loads(answers)
            except (json.JSONDecodeError, TypeError):
                answers = {}

        row_data = [
            i,
            s.get("student_name", s.get("student_id", "")[:12]),
            _nisn_val(answers, s.get("student_id", "")),
            s.get("score", 0),
            s.get("penalty", 0),
            s.get("final_score", s.get("score", 0)),
            s.get("status", "submitted"),
            _format_dt(s.get("submitted_at", "")),
        ]

        for qi in range(total_q):
            qtype = q_types.get(str(qi), "mcq")
            ans_data = answers.get(str(qi))
            correct = _get_answer_key(exam, qi)
            ans, text_content = _parse_answer(ans_data, qi, qtype, exam.get("answer_key"))

            if qtype == "mcq":
                display = f"{ans}" if ans else "-"
                if correct and ans == correct:
                    display += " ✓"
                row_data.append(display)
            else:
                row_data.append(ans if ans else "✓")
                row_data.append(text_content[:200] if text_content else "")

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.border = thin_border
            cell.alignment = wrap_align

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 32
    for col in range(4, 9):
        ws.column_dimensions[chr(64 + col)].width = 14
    for qi in range(total_q):
        col_idx = 9 + qi
        if col_idx <= 26:
            ws.column_dimensions[chr(64 + col_idx)].width = 12
        else:
            ws.column_dimensions[chr(64 + (col_idx // 26)) + chr(65 + (col_idx % 26))].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_to_pdf(submissions: list, exam_title: str = "Hasil Ujian", exam: dict = None) -> io.BytesIO:
    buf = io.BytesIO()
    c = pdf_canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    q_types = exam.get("question_types") or {} if exam else {}
    if isinstance(q_types, str):
        try:
            q_types = json.loads(q_types)
        except (json.JSONDecodeError, TypeError):
            q_types = {}
    total_q = exam.get("total_questions", 0) if exam else 0


    for s in submissions:
        answers = s.get("answers") or {}
        if isinstance(answers, str):
            try:
                answers = json.loads(answers)
            except (json.JSONDecodeError, TypeError):
                answers = {}

        student_name = s.get("student_name", s.get("student_id", "")[:12])
        student_nisn = _nisn_val(s.get("answers", ""), "")
        score = s.get("score", 0)
        penalty = s.get("penalty", 0)
        final_score = s.get("final_score", score)

        # Header
        c.setFont("Helvetica-Bold", 16)
        c.drawString(40, height - 40, exam_title)
        c.setFont("Helvetica", 10)
        c.drawString(40, height - 58, f"Siswa: {student_name}")
        if student_nisn:
            c.drawString(40, height - 72, f"NISN: {student_nisn}")
            y_start = 72
        else:
            y_start = 58
        c.drawString(300, height - y_start, f"Final: {final_score}")
        c.drawString(450, height - y_start, f"Penalti: {penalty}")
        c.line(40, height - y_start - 6, width - 40, height - y_start - 6)

        y = height - y_start - 20
        c.setFont("Helvetica-Bold", 9)

        for qi in range(total_q):
            if y < 60:
                c.showPage()
                y = height - 40
                c.setFont("Helvetica-Bold", 9)

            qtype = q_types.get(str(qi), "mcq")
            ans_data = answers.get(str(qi))
            correct = _get_answer_key(exam, qi)
            ans, text_content = _parse_answer(ans_data, qi, qtype, exam.get("answer_key"))

            q_label = f"Soal {qi+1} ({qtype})"
            c.setFont("Helvetica-Bold", 10)
            c.drawString(40, y, q_label)
            y -= 14

            c.setFont("Helvetica", 9)
            if qtype == "mcq":
                status = "✓" if correct and ans == correct else "✗"
                c.drawString(50, y, f"Jawaban: {ans if ans else '-'} | Kunci: {correct} {status}")
                y -= 14
            else:
                if ans:
                    c.drawString(50, y, f"Skor guru: {ans}")
                    y -= 14
                if text_content:
                    c.setFont("Helvetica-Oblique", 8)
                    lines = _wrap_text(text_content, 80)
                    for line in lines[:8]:
                        c.drawString(50, y, line)
                        y -= 10
                    c.setFont("Helvetica", 9)
                    y -= 4

            # Draw canvas image if exists
            if isinstance(ans_data, dict):
                pages = ans_data.get("pages") or {}
                for p_idx, p_data in pages.items():
                    if isinstance(p_data, dict) and p_data.get("canvas"):
                        canvas_data = p_data["canvas"]
                        try:
                            _draw_canvas_on_pdf(c, canvas_data, 50, y - 10, width - 100)
                            y -= 110
                        except Exception:
                            c.drawString(50, y, "[Canvas tidak dapat dirender]")
                            y -= 14

            y -= 6

        c.showPage()

    c.save()
    buf.seek(0)
    return buf


def _draw_canvas_on_pdf(c, data_url, x, y, max_width):
    """Decode a base64 PNG data URL and draw it on the PDF canvas."""
    match = re.match(r'^data:image/\w+;base64,(.+)$', data_url)
    if not match:
        return
    img_data = base64.b64decode(match.group(1))
    img_buf = io.BytesIO(img_data)
    pil_img = PilImage.open(img_buf)
    img_reader = ImageReader(pil_img)

    img_w, img_h = pil_img.size
    aspect = img_h / max(img_w, 1)
    draw_w = min(img_w, max_width)
    draw_h = draw_w * aspect
    if draw_h > 200:
        draw_h = 200
        draw_w = draw_h / aspect

    c.drawImage(img_reader, x, y - draw_h, width=draw_w, height=draw_h, preserveAspectRatio=True)


def _wrap_text(text, max_chars):
    """Simple word wrap for PDF text."""
    words = text.split()
    lines = []
    current = ""
    for w in words:
        if len(current + " " + w) <= max_chars:
            current = (current + " " + w).strip()
        else:
            if current:
                lines.append(current)
            current = w
    if current:
        lines.append(current)
    return lines if lines else [text]


def generate_bubble_sheet_pdf(
    title: str,
    total_questions: int = 20,
    subject: str = "",
    student_name: str = "",
    options: int = 5,
) -> io.BytesIO:
    """Generate bubble sheet PDF like ZipGrade."""
    buf = io.BytesIO()
    c = pdf_canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    margin = 30 * mm

    c.setFont("Helvetica-Bold", 18)
    c.drawString(margin, height - 30 * mm, title)
    c.setFont("Helvetica", 11)
    c.drawString(margin, height - 40 * mm, f"Mata Pelajaran: {subject}")
    if student_name:
        c.drawString(margin, height - 48 * mm, f"Nama: {student_name}")
    c.drawString(margin, height - 56 * mm, f"Jumlah Soal: {total_questions}")
    c.line(margin, height - 60 * mm, width - margin, height - 60 * mm)

    label_y = height - 68 * mm
    c.setFont("Helvetica-Bold", 9)
    c.drawString(margin + 20 * mm, label_y, "Jawaban")
    for j, opt in enumerate(["A", "B", "C", "D", "E"][:options]):
        cx = margin + 38 * mm + j * 14 * mm
        c.drawString(cx, label_y, opt)

    bubble_y = label_y - 6 * mm
    c.setFont("Helvetica", 8)
    for i in range(total_questions):
        y_pos = bubble_y - i * 10 * mm
        c.drawString(margin, y_pos + 2 * mm, str(i + 1))
        for j in range(options):
            cx = margin + 38 * mm + j * 14 * mm
            c.circle(cx + 3 * mm, y_pos + 2 * mm, 2.5 * mm, fill=0, stroke=1)
        if (i + 1) % 25 == 0 or i == total_questions - 1:
            c.showPage()
            if i < total_questions - 1:
                c.setFont("Helvetica-Bold", 18)
                c.drawString(margin, height - 30 * mm, title)

    c.save()
    buf.seek(0)
    return buf
