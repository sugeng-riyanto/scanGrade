import io
import os
import json
import zipfile
import uuid
import time
from datetime import datetime, timezone
from flask import Blueprint, request, jsonify, g, render_template, redirect, send_file, current_app
from app.utils.auth import login_required, get_supabase
from app.services.anti_cheat_service import validate_violation_log
from app.utils.logger import get_logger
from app.errors import ValidationError, NotFoundError, GradingError, AIProcessingError
from app.utils.rate_limiter import limiter

def _rate_limit(n):
    return limiter.limit(n) if limiter else (lambda f: f)

api_bp = Blueprint("api", __name__)

# ── Upload security constants ──────────────────────────
ALLOWED_EXTENSIONS = {".jpg", ".jpeg", ".png"}
MAX_IMAGE_SIZE = 20 * 1024 * 1024  # 20MB
UPLOAD_SCAN_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "static", "uploads", "scans")

def _redis_lock(key, timeout=10):
    """Redis-based cross-worker lock (SETNX pattern)."""
    try:
        from redis import Redis
        r = Redis.from_url(current_app.config.get("REDIS_URL", "redis://localhost:6379/0"))
        lock_key = f"scan_grade:lock:{key}"
        # SETNX: only set if key doesn't exist, with expiry
        if r.setnx(lock_key, "1"):
            r.expire(lock_key, timeout)
            return r
        return None
    except Exception:
        return None

def _release_lock(redis_conn, key):
    """Release a Redis lock."""
    try:
        if redis_conn:
            redis_conn.delete(f"scan_grade:lock:{key}")
    except Exception:
        pass

_sync_last = {}


def _check_rate_limit(user_id, exam_id, min_interval=5):
    key = f"{user_id}:{exam_id}"
    now = time.time()
    last = _sync_last.get(key, 0)
    if now - last < min_interval:
        return False
    _sync_last[key] = now
    return True


@api_bp.route("/violation/log", methods=["POST"])
@login_required
def log_violation():
    data = request.get_json()
    logs = data if isinstance(data, list) else [data]

    supabase = get_supabase()
    results = []
    for log in logs:
        valid = validate_violation_log(
            g.user_id,
            log.get("exam_id", ""),
            log.get("timestamp", 0),
        )
        if valid["valid"]:
            exam_id = log.get("exam_id", "")
            try:
                supabase.table("violation_logs").insert({
                    "exam_id": exam_id,
                    "user_id": g.user_id,
                    "violation_type": log.get("violation_type", "unknown"),
                    "metadata": log.get("metadata", {}),
                }).execute()
            except Exception:
                current_app.logger.warning("Violation log insert failed for exam %s", exam_id)
                results.append({"logged": False, "reason": "db_error"})
                continue
            total_count = supabase.table("violation_logs").select("id", count="exact").eq("user_id", g.user_id).eq("exam_id", exam_id).execute().count or 0
            exam = supabase.table("exams").select("anti_cheat_enabled, penalty_per_violation, max_violations, auto_submit_on_max").eq("id", exam_id).single().execute().data or {}
            from app.services.anti_cheat_service import calculate_graduated_penalty
            penalty_info = calculate_graduated_penalty(total_count, exam)
            # Save penalty to the student's submission
            try:
                sub = supabase.table("submissions").select("id,penalty").eq("exam_id", exam_id).eq("student_id", g.user_id).order("created_at", desc=True).limit(1).execute()
                if sub.data:
                    current_penalty = float(sub.data[0].get("penalty") or 0)
                    new_penalty = current_penalty + penalty_info.get("current_penalty_this_violation", 0)
                    supabase.table("submissions").update({"penalty": new_penalty}).eq("id", sub.data[0]["id"]).execute()
            except Exception:
                pass
            results.append({"logged": True, "violation_count": total_count, **penalty_info})
        else:
            results.append({"logged": False, "reason": valid.get("reason")})

    return jsonify({"violations": results})


@api_bp.route("/student/force-submit", methods=["POST"])
@login_required
def force_submit():
    """Force-submit an exam when anti-cheat max violations reached.
    This is a fallback when the Alpine component cannot be triggered.
    """
    data = request.get_json()
    exam_id = (data or {}).get("exam_id", "")
    if not exam_id:
        return jsonify({"error": "exam_id required"}), 400

    supabase = get_supabase()
    # Find the latest draft submission and submit it
    try:
        sub = supabase.table("submissions") \
            .select("id,answers") \
            .eq("exam_id", exam_id) \
            .eq("student_id", g.user_id) \
            .eq("status", "draft") \
            .order("created_at", desc=True) \
            .limit(1) \
            .execute()
        if sub.data:
            answers = sub.data[0].get("answers") or {}
            # Re-fetch submission for answer key
            exam = supabase.table("exams").select("answer_key,question_types,question_weights,total_questions,penalty_per_violation").eq("id", exam_id).single().execute().data or {}
            # Parse JSON fields that may be strings
            for _fld in ("answer_key", "question_types", "question_weights"):
                _v = exam.get(_fld)
                if isinstance(_v, str):
                    try: exam[_fld] = json.loads(_v)
                    except (json.JSONDecodeError, TypeError): exam[_fld] = {}
            key = exam.get("answer_key") or {}
            qtypes = exam.get("question_types") or {}
            weights = exam.get("question_weights") or {}
            total_q = exam.get("total_questions", 0)
            penalty = float(exam.get("penalty_per_violation", 5))
            # Calculate MCQ score
            earned = 0.0
            mcq_count = 0
            for i in range(total_q):
                qi = str(i)
                qt = qtypes.get(qi, "mcq")
                kv = key.get(qi)
                w = float(weights.get(qi, 0))
                if qt == "mcq" and kv and w > 0:
                    mcq_count += 1
                    ans = answers.get(qi)
                    if isinstance(ans, dict): ans = ans.get("answer", "")
                    if kv == "bonus":
                        if ans and str(ans).strip(): earned += w
                    elif isinstance(kv, list):
                        if ans in kv: earned += w
                    elif ans == kv: earned += w
            score = round(min(earned, 100), 2)
            # Get actual penalty from violation logs
            viol_count = supabase.table("violation_logs").select("id", count="exact").eq("user_id", g.user_id).eq("exam_id", exam_id).execute().count or 0
            from app.services.anti_cheat_service import calculate_graduated_penalty
            pinfo = calculate_graduated_penalty(viol_count, exam)
            total_penalty = pinfo.get("penalty", 0)
            final = max(0.0, round(score - total_penalty, 2))
            supabase.table("submissions") \
                .update({"status": "submitted", "answers": answers, "score": score, "final_score": final, "penalty": total_penalty}) \
                .eq("id", sub.data[0]["id"]) \
                .execute()
            return jsonify({"success": True, "score": score, "final_score": final, "penalty": total_penalty})
        return jsonify({"error": "No draft submission found"}), 404
    except Exception as e:
        current_app.logger.error("force_submit error: %s", e)
        return jsonify({"error": str(e)[:100]}), 500


@api_bp.route("/violation/count", methods=["GET"])
@login_required
def violation_count():
    exam_id = request.args.get("exam_id")
    supabase = get_supabase()
    res = supabase.table("violation_logs") \
        .select("count", count="exact") \
        .eq("user_id", g.user_id) \
        .eq("exam_id", exam_id) \
        .execute()
    return jsonify({"count": res.count})


@api_bp.route("/scan/process", methods=["POST"])
@_rate_limit("20 per minute")
@login_required
def scan_process():
    """Process a scanned bubble sheet image and return detected answers.

    Security: validates extension, MIME type, image integrity, strips EXIF.
    """
    if "image" not in request.files:
        return jsonify({"error": "Tidak ada gambar yang dikirim"}), 400

    image_file = request.files["image"]
    raw_bytes = image_file.read()

    # 0. Handle PDF upload — convert first page to image
    ext = os.path.splitext(image_file.filename or "")[1].lower()
    if ext == ".pdf":
        try:
            from pdf2image import convert_from_bytes
            pages = convert_from_bytes(raw_bytes, first_page=1, last_page=1)
            if not pages:
                return jsonify({"error": "Gagal membaca PDF"}), 422
            out = io.BytesIO()
            pages[0].save(out, format="JPEG", quality=90)
            raw_bytes = out.getvalue()
        except ImportError:
            return jsonify({"error": "PDF support tidak tersedia (install pdf2image)"}), 422
        except Exception as e:
            return jsonify({"error": f"Gagal memproses PDF: {str(e)[:100]}"}), 422

    # 1. Validate extension
    if ext not in ALLOWED_EXTENSIONS and ext != ".pdf":
        return jsonify({"error": "Format file tidak didukung. Gunakan .jpg, .jpeg, .png, atau .pdf"}), 422

    # 2. Validate MIME type via python-magic
    mime_type = None
    try:
        import magic
        header = raw_bytes[:2048]
        mime_type = magic.from_buffer(header, mime=True)
        if ext != ".pdf" and mime_type not in ("image/jpeg", "image/png"):
            return jsonify({"error": f"Tipe file tidak valid ({mime_type}). Hanya jpg/png yang diizinkan."}), 422
    except ImportError:
        current_app.logger.warning("python-magic not installed; skipping MIME validation")

    # 3. Validate size
    file_size = len(raw_bytes)
    if file_size > MAX_IMAGE_SIZE:
        return jsonify({"error": f"Gambar terlalu besar ({file_size/1024/1024:.1f}MB). Maksimal 20MB."}), 413
    if file_size == 0:
        return jsonify({"error": "File kosong"}), 422

    # 4. Verify image integrity + strip EXIF via Pillow
    try:
        from PIL import Image
        buf = io.BytesIO(raw_bytes)
        img_pil = Image.open(buf)
        img_pil.verify()
        buf.seek(0)
        img_pil = Image.open(buf)
        clean_buf = io.BytesIO()
        save_format = "PNG" if mime_type == "image/png" else "JPEG"
        if "exif" in img_pil.info:
            img_pil.info.pop("exif")
        if save_format == "JPEG" and img_pil.mode != "RGB":
            img_pil = img_pil.convert("RGB")
        img_pil.save(clean_buf, format=save_format)
        image_data = clean_buf.getvalue()
        current_app.logger.info("EXIF stripped from scan image, cleaned size=%d", len(image_data))
    except Exception as e:
        return jsonify({"error": "Gambar tidak valid atau corrupt", "detail": str(e)[:100]}), 422

    exam_id = request.form.get("exam_id", "")
    total_questions = int(request.form.get("total_questions", 50))

    _cleanup_scan_tmp()

    # Save original image to temp
    scan_file_id = str(uuid.uuid4())[:8]
    scan_dir = os.path.join(UPLOAD_SCAN_DIR, "tmp")
    os.makedirs(scan_dir, exist_ok=True)
    scan_ext = ".png" if mime_type == "image/png" else ".jpg"
    scan_temp_path = os.path.join(scan_dir, scan_file_id + scan_ext)
    try:
        with open(scan_temp_path, "wb") as f:
            f.write(image_data)
    except Exception as e:
        current_app.logger.warning("Failed to save scan temp: %s", e)

    # Enqueue async Celery task
    try:
        from app.services.omr_tasks import process_omr_scan
        task = process_omr_scan.delay(
            image_path=scan_temp_path,
            total_questions=total_questions,
            exam_id=exam_id,
        )
        current_app.logger.info("OMR task enqueued: %s", task.id)
        return jsonify({
            "async": True,
            "task_id": task.id,
            "scan_file_id": scan_file_id,
            "status": "processing",
        })
    except Exception as e:
        current_app.logger.error("Failed to enqueue OMR task: %s", e)
        # Fallback to synchronous processing
        from app.services.omr_service import process_scan, draw_debug_image, preprocess_scan
        try:
            result = process_scan(image_data, total_questions=total_questions, preprocess=True)
        except Exception as e2:
            return jsonify({"error": f"Gagal memproses: {str(e2)[:200]}"}), 422

        if "error" not in result and exam_id:
            supabase = get_supabase()
            exam = supabase.table("exams").select("*").eq("id", exam_id).single().execute().data
            if exam and exam.get("answer_key"):
                key = exam["answer_key"]
                if isinstance(key, str):
                    key = json.loads(key)
                detected = result.get("answers", {})
                correct = 0
                for k, v in key.items():
                    if k in detected and detected[k] == v and v not in ("essay", "essay_text", "essay_canvas"):
                        correct += 1
                mcq_count = sum(1 for v in key.values() if v not in ("essay", "essay_text", "essay_canvas"))
                result["score"] = round((correct / max(mcq_count, 1)) * 100, 2)
                result["correct"] = correct

        if "error" not in result:
            from app.services.omr_service import load_image, find_registration_marks
            img = load_image(image_data)
            if img is not None:
                corners = find_registration_marks(img)
                debug_jpg = draw_debug_image(img, corners, result.get("answers"))
                import base64
                result["debug_image"] = base64.b64encode(debug_jpg).decode()
                result["scan_file_id"] = scan_file_id

        result["async"] = False
        return jsonify(result)




def _scan_essay_vision(image_bytes, api_key="", lang="en"):
    try:
        from google import genai
        import PIL.Image
        import io
        client = genai.Client(api_key=api_key)
        prompt = "Extract ALL handwritten text from this exam answer sheet. Preserve the original language. Output only the text content."
        if lang and not lang.startswith("en"):
            prompt += " The handwriting is in " + lang + "."
        buf = io.BytesIO(image_bytes)
        img = PIL.Image.open(buf)
        response = client.models.generate_content(
            model="gemini-2.0-flash",
            contents=[prompt, img]
        )
        return response.text.strip()
    except ImportError:
        raise ImportError("google-genai tidak terinstall. Install: pip install google-genai")
    except ValueError:
        raise
    except Exception as e:
        err = str(e)
        if "API_KEY" in err or "not found" in err.lower() or "unauthorized" in err.lower():
            raise ValueError("API Key Gemini tidak valid. Setup di Pengaturan AI.")
        if "quota" in err.lower() or "429" in err or "RATE_LIMIT" in err.upper():
            raise ValueError("Kuota Gemini habis. Tunggu sebentar.")
        raise


@api_bp.route("/scan/essay", methods=["POST"])
@login_required
def scan_essay():
    if "image" not in request.files:
        return jsonify({"error": "Tidak ada gambar"}), 400
    image_file = request.files["image"]
    raw = image_file.read()
    if len(raw) > 20 * 1024 * 1024:
        return jsonify({"error": "Gambar terlalu besar. Maksimal 20MB"}), 413
    from app.services.ai_service import _get_active_key
    key_data = _get_active_key(g.user_id)
    api_key = ""
    if key_data and key_data.get("provider") == "gemini":
        api_key = key_data.get("api_key", "")
    elif key_data:
        supabase = get_supabase()
        gemini_keys = supabase.table("teacher_ai_keys").select("*").eq("teacher_id", g.user_id).eq("provider", "gemini").limit(1).execute().data
        if gemini_keys:
            api_key = gemini_keys[0].get("api_key", "")
    if not api_key:
        return jsonify({"error": "API Key Gemini tidak ditemukan. Setup Gemini di Pengaturan AI."}), 400
    try:
        text = _scan_essay_vision(raw, api_key=api_key)
        return jsonify({"success": True, "text": text})
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": "Gagal OCR: " + str(e)[:200]}), 500


@api_bp.route("/grade/vision-canvas", methods=["POST"])
@login_required
def vision_canvas_ocr():
    data = request.get_json() or {}
    image_data = data.get("image", "")
    if not image_data or not image_data.startswith("data:image/"):
        return jsonify({"error": "Data URL gambar tidak valid"}), 400
    try:
        import base64
        header, encoded = image_data.split(",", 1)
        img_bytes = base64.b64decode(encoded)
    except Exception as e:
        return jsonify({"error": "Gagal decode: " + str(e)[:100]}), 400
    from app.services.ai_service import _get_active_key
    key_data = _get_active_key(g.user_id)
    api_key = ""
    if key_data and key_data.get("provider") == "gemini":
        api_key = key_data.get("api_key", "")
    elif key_data:
        supabase = get_supabase()
        gemini_keys = supabase.table("teacher_ai_keys").select("*").eq("teacher_id", g.user_id).eq("provider", "gemini").limit(1).execute().data
        if gemini_keys:
            api_key = gemini_keys[0].get("api_key", "")
    if not api_key:
        return jsonify({"error": "API Key Gemini tidak ditemukan. Setup Gemini di Pengaturan AI."}), 400
    try:
        from google import genai
        import PIL.Image
        import io
        client = genai.Client(api_key=api_key)
        prompt = "Extract ALL handwritten text, formulas, equations, and diagram labels from this student exam drawing. Output only the text."
        buf = io.BytesIO(img_bytes)
        img = PIL.Image.open(buf)
        response = client.models.generate_content(model="gemini-2.0-flash", contents=[prompt, img])
        text = response.text.strip()
        return jsonify({"success": True, "text": text})
    except ImportError:
        return jsonify({"error": "google-genai tidak terinstall. Install: pip install google-genai"}), 400
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": "Gagal OCR: " + str(e)[:200]}), 500


@api_bp.route("/scan/task/<task_id>", methods=["GET"])
@login_required
def scan_task_status(task_id):
    """Poll Celery task status and get result when done."""
    from app.celery_app import celery_app
    task = celery_app.AsyncResult(task_id)
    response = {"task_id": task_id, "status": task.state}

    if task.state == "PENDING":
        response["status"] = "pending"
    elif task.state == "PROCESSING":
        response["status"] = "processing"
    elif task.state == "PROGRESS":
        response["status"] = "processing"
        response["progress"] = task.info
    elif task.state == "SUCCESS":
        response["status"] = "done"
        response["result"] = task.result
    elif task.state == "FAILURE":
        response["status"] = "error"
        response["error"] = str(task.info)
        # Clean up task result
        task.forget()

    return jsonify(response)


@api_bp.route("/scan/bulk", methods=["POST"])
@_rate_limit("10 per minute")
@login_required
def scan_bulk():
    """Process a ZIP file containing multiple LJK scan images."""
    if "archive" not in request.files:
        return jsonify({"error": "Tidak ada file ZIP yang dikirim"}), 400

    archive_file = request.files["archive"]
    ext = os.path.splitext(archive_file.filename or "")[1].lower()
    if ext not in (".zip",):
        return jsonify({"error": "Hanya file ZIP yang didukung"}), 400

    exam_id = request.form.get("exam_id", "")
    total_questions = int(request.form.get("total_questions", 50))

    _cleanup_scan_tmp()

    # Save ZIP to temp
    zip_id = str(uuid.uuid4())[:8]
    zip_dir = os.path.join(UPLOAD_SCAN_DIR, "tmp")
    os.makedirs(zip_dir, exist_ok=True)
    zip_path = os.path.join(zip_dir, f"bulk_{zip_id}.zip")
    archive_file.save(zip_path)

    # Enqueue async Celery task
    try:
        from app.services.omr_tasks import process_bulk_scan
        task = process_bulk_scan.delay(
            zip_path=zip_path,
            total_questions=total_questions,
            exam_id=exam_id,
        )
        current_app.logger.info("Bulk OMR task enqueued: %s (%d images)", task.id, 0)
        return jsonify({
            "async": True,
            "task_id": task.id,
            "status": "processing",
        })
    except Exception as e:
        current_app.logger.error("Failed to enqueue bulk task: %s", e)
        # Fallback: remove temp ZIP and return error
        try:
            os.remove(zip_path)
        except OSError:
            pass
        return jsonify({"error": "Gagal mengantrekan pemrosesan. Coba lagi."}), 500


@api_bp.route("/scan/bulk-save", methods=["POST"])
@login_required
@_rate_limit("10 per minute")
def scan_bulk_save():
    """Save multiple scan results as submissions at once."""
    data = request.get_json()
    if not data or "submissions" not in data:
        return jsonify({"error": "No submissions data"}), 400

    exam_id = data.get("exam_id")
    if not exam_id:
        return jsonify({"error": "exam_id required"}), 400

    supabase = get_supabase()
    saved = []
    failed = []

    for sub in data["submissions"]:
        student_id = sub.get("student_id")
        answers = sub.get("answers", {})
        nisn = sub.get("nisn", "")
        confidence = sub.get("confidence", {})
        needs_review = sub.get("needs_review", [])
        scan_file_id = sub.get("scan_file_id", "")

        if not student_id or not answers:
            failed.append({"student_id": student_id, "error": "Missing student_id or answers"})
            continue

        # Grade
        exam = supabase.table("exams").select("answer_key").eq("id", exam_id).single().execute().data
        key = exam.get("answer_key", {}) if exam else {}
        if isinstance(key, str):
            key = json.loads(key)
        correct = 0
        for k, v in key.items():
            if k in answers and answers[k] == v and v not in ("essay", "essay_text", "essay_canvas"):
                correct += 1
        mcq_count = sum(1 for v in key.values() if v not in ("essay", "essay_text", "essay_canvas"))
        score = round((correct / max(mcq_count, 1)) * 100, 2) if mcq_count > 0 else 0

        enriched = {}
        for k, v in answers.items():
            entry = {"answer": v}
            if k in confidence:
                entry["confidence"] = confidence[k]
            if k in needs_review:
                entry["review"] = True
            enriched[k] = entry
        if nisn:
            enriched["_nisn"] = nisn

        # Move scan image from tmp to permanent storage
        if scan_file_id:
            for ext in (".png", ".jpg"):
                src = os.path.join(UPLOAD_SCAN_DIR, "tmp", scan_file_id + ext)
                if os.path.exists(src):
                    dst_dir = os.path.join(UPLOAD_SCAN_DIR, exam_id)
                    os.makedirs(dst_dir, exist_ok=True)
                    dst = os.path.join(dst_dir, student_id + ext)
                    try:
                        os.rename(src, dst)
                        enriched["_scan_image"] = f"/static/uploads/scans/{exam_id}/{student_id}{ext}"
                    except Exception as e:
                        current_app.logger.warning("Failed to move scan image: %s", e)
                    break

        try:
            existing = supabase.table("submissions").select("id").eq("exam_id", exam_id).eq("student_id", student_id).execute().data
            update_data = {"answers": enriched, "score": score, "max_score": mcq_count, "status": "graded"}
            if existing:
                supabase.table("submissions").update(update_data).eq("id", existing[0]["id"]).execute()
            else:
                update_data.update({"exam_id": exam_id, "student_id": student_id})
                supabase.table("submissions").insert(update_data).execute()
            saved.append({"student_id": student_id, "score": score, "correct": correct, "nisn": nisn})
        except Exception as e:
            failed.append({"student_id": student_id, "error": str(e)[:100]})

    return jsonify({
        "success": True,
        "saved": len(saved),
        "failed": len(failed),
        "details": {"saved": saved, "failed": failed},
    })


def _cleanup_scan_tmp(age_hours=1):
    """Remove stale temp scan files older than age_hours."""
    from app.services.cleanup_service import clean_temp_files
    clean_temp_files(max_age=age_hours * 3600)


@api_bp.route("/student/auto-save", methods=["POST"])
@login_required
def student_auto_save():
    """Auto-save student's in-progress exam draft."""
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data"}), 400
    return jsonify({"saved": True, "at": int(time.time())})


@api_bp.route("/student/sync-draft", methods=["POST"])
@login_required
def student_sync_draft():
    """Sync student draft — lightweight MCQ/text every 20s, canvas every 60s."""
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"saved": True, "at": int(time.time())})
    exam_id = data.get("exam_id")
    answers = data.get("answers", {})
    is_light = data.get("light", True)
    if not exam_id:
        return jsonify({"saved": True, "at": int(time.time())})
    if not _check_rate_limit(g.user_id, exam_id, min_interval=3 if is_light else 10):
        return jsonify({"saved": True, "at": int(time.time()), "throttled": True})
    lock_key = f"sync:{g.user_id}:{exam_id}"
    rlock = _redis_lock(lock_key)
    if not rlock:
        return jsonify({"saved": True, "at": int(time.time()), "busy": True})
    try:
        from app.utils.auth import get_supabase
        supabase = get_supabase()
        existing = supabase.table("submissions").select("id,status,answers").eq("exam_id", exam_id).eq("student_id", g.user_id).execute().data
        if existing and existing[0].get("status") == "draft":
            if is_light and existing[0].get("answers"):
                merged = existing[0]["answers"]
                if isinstance(merged, dict):
                    merged.update(answers)
                    answers = merged
            supabase.table("submissions").update({"answers": answers}).eq("id", existing[0]["id"]).execute()
        elif not existing:
            client_started = data.get("started_at")
            started_at_ts = client_started if client_started else int(time.time())
            # Use client's started_at if earlier than now (more accurate to when user opened exam)
            started_at_dt = datetime.fromtimestamp(started_at_ts / 1000 if started_at_ts > 1e10 else started_at_ts, tz=timezone.utc).isoformat()
            supabase.table("submissions").insert({
                "exam_id": exam_id,
                "student_id": g.user_id,
                "answers": answers,
                "score": 0,
                "max_score": 100,
                "status": "draft",
                "started_at": started_at_dt,
            }).execute()
        elif existing and existing[0].get("started_at") is None:
            # Backfill started_at if missing (e.g. old drafts before this fix)
            client_started = data.get("started_at")
            if client_started:
                started_at_dt = datetime.fromtimestamp(client_started / 1000 if client_started > 1e10 else client_started, tz=timezone.utc).isoformat()
                supabase.table("submissions").update({"started_at": started_at_dt}).eq("id", existing[0]["id"]).execute()
    except Exception:
        pass
    finally:
        _release_lock(rlock, lock_key)
    return jsonify({"saved": True, "at": int(time.time())})


@api_bp.route("/grade/auto-save/<submission_id>", methods=["POST"])
@login_required
def grade_auto_save(submission_id):
    """Auto-save teacher's in-progress grading."""
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data"}), 400
    supabase = get_supabase()
    # Save teacher_feedback draft into submission
    feedback = data.get("teacher_feedback", {})
    supabase.table("submissions") \
        .update({"teacher_feedback": feedback}) \
        .eq("id", submission_id) \
        .execute()
    return jsonify({"saved": True, "at": int(time.time())})


@api_bp.route("/grade/batch", methods=["POST"])
@login_required
def grade_batch():
    """Batch grade all submissions for an exam. Optimized for <2s grading speed."""
    import json, time as _time
    from app.utils.auth import get_supabase, teacher_or_admin_required
    data = request.get_json()
    exam_id = data.get("exam_id") if data else None
    if not exam_id:
        return jsonify({"error": "exam_id required"}), 400
    if g.user_role not in ("guru", "super_admin", "admin_sekolah"):
        return jsonify({"error": "Forbidden"}), 403
    supabase = get_supabase()
    t0 = _time.time()
    exam = supabase.table("exams").select("*").eq("id", exam_id).single().execute().data
    if not exam:
        return jsonify({"error": "Exam not found"}), 404
    # Parse JSON fields that may be strings
    for _fld in ("answer_key", "question_types", "question_weights", "question_pages"):
        _v = exam.get(_fld)
        if isinstance(_v, str):
            try: exam[_fld] = json.loads(_v)
            except (json.JSONDecodeError, TypeError): exam[_fld] = {}
    answer_key = exam.get("answer_key") or {}
    question_types = exam.get("question_types") or {}
    question_weights = exam.get("question_weights") or {}
    total_q = exam.get("total_questions", 0)
    if not question_weights and total_q > 0:
        mcq_count = sum(1 for i in range(total_q) if question_types.get(str(i), "mcq") == "mcq")
        if mcq_count > 0:
            each = round(100 / mcq_count, 2)
            for i in range(total_q):
                if question_types.get(str(i), "mcq") == "mcq":
                    question_weights[str(i)] = each
    subs = supabase.table("submissions").select("id,answers,penalty,teacher_feedback").eq("exam_id", exam_id).in_("status", ["submitted", "graded", "published"]).execute().data or []
    graded = 0
    for sub in subs:
        # Parse JSON fields in each submission
        for _sf in ("answers", "teacher_feedback"):
            _sv = sub.get(_sf)
            if isinstance(_sv, str):
                try: sub[_sf] = json.loads(_sv)
                except (json.JSONDecodeError, TypeError): sub[_sf] = {}
        answers = sub.get("answers") or {}
        earned = 0.0
        for i in range(total_q):
            qtype = question_types.get(str(i), "mcq")
            key_val = answer_key.get(str(i))
            w = float(question_weights.get(str(i), 0))
            if w <= 0:
                continue
            if qtype == "mcq":
                ans = answers.get(str(i))
                if key_val == "bonus":
                    if ans and str(ans).strip():
                        earned += w
                elif isinstance(key_val, list):
                    if ans in key_val:
                        earned += w
                elif ans == key_val:
                    earned += w
        fb = sub.get("teacher_feedback") or {}
        fb_scores = fb.get("scores", {}) or {}
        for qi, sv in fb_scores.items():
            if sv is not None and sv != "":
                ew = float(question_weights.get(str(qi), 0))
                if ew > 0:
                    earned += float(sv) / 100.0 * ew
        final = round(min(earned, 100), 2)
        penalty = float(sub.get("penalty") or 0)
        final = max(0, round(final - penalty, 2))
        mcq_correct = 0
        mcq_count_q = sum(1 for i in range(total_q) if question_types.get(str(i), "mcq") == "mcq")
        for i in range(total_q):
            qtype = question_types.get(str(i), "mcq")
            key_val = answer_key.get(str(i))
            if qtype == "mcq" and key_val:
                ans = answers.get(str(i))
                if key_val == "bonus":
                    if ans and str(ans).strip():
                        mcq_correct += 1
                elif isinstance(key_val, list):
                    if ans in key_val:
                        mcq_correct += 1
                elif ans == key_val:
                    mcq_correct += 1
        mcq_score = round((mcq_correct / max(mcq_count_q, 1)) * 100, 2) if mcq_count_q > 0 else 0
        supabase.table("submissions").update({
            "score": mcq_score,
            "final_score": final,
        }).eq("id", sub["id"]).execute()
        graded += 1
    elapsed = round((_time.time() - t0) * 1000)
    return jsonify({
        "success": True,
        "graded": graded,
        "elapsed_ms": elapsed,
        "per_submission_ms": round(elapsed / max(graded, 1), 1),
    })


@api_bp.route("/scan/save", methods=["POST"])
@login_required
@_rate_limit("30 per minute")
def scan_save():
    """Save scanned answers as a submission."""
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400

    exam_id = data.get("exam_id")
    student_id = data.get("student_id")
    answers = data.get("answers")
    nisn = data.get("nisn") or ""
    confidence = data.get("confidence") or {}
    needs_review = data.get("needs_review") or []
    scan_file_id = data.get("scan_file_id") or ""

    if not all([exam_id, student_id, answers]):
        return jsonify({"error": "exam_id, student_id, and answers are required"}), 400

    supabase = get_supabase()

    # Verify exam exists and user has access (teacher or admin_sekolah)
    exam = supabase.table("exams").select("*").eq("id", exam_id).single().execute().data
    if not exam:
        return jsonify({"error": "Exam not found"}), 404
    user_role = g.get("user_role")
    user_school = g.get("user_school_id")
    if user_role == "guru" and str(exam.get("teacher_id")) != g.user_id:
        return jsonify({"error": "Forbidden"}), 403
    if user_role == "admin_sekolah" and str(exam.get("school_id")) != str(user_school):
        return jsonify({"error": "Forbidden"}), 403
    if user_role not in ("guru", "admin_sekolah"):
        return jsonify({"error": "Forbidden"}), 403

    # Grade MCQ answers
    key = exam.get("answer_key", {})
    if isinstance(key, str):
        key = json.loads(key)
    detected = answers
    correct = 0
    for k, v in key.items():
        if k in detected and detected[k] == v and v not in ("essay", "essay_text", "essay_canvas"):
            correct += 1
    mcq_count = sum(1 for v in key.values() if v not in ("essay", "essay_text", "essay_canvas"))
    score = round((correct / max(mcq_count, 1)) * 100, 2) if mcq_count > 0 else 0

    # Build enriched answers dict with confidence metadata
    enriched_answers = {}
    for k, v in answers.items():
        entry = {"answer": v}
        if k in confidence:
            entry["confidence"] = confidence[k]
        if k in needs_review:
            entry["review"] = True
        enriched_answers[k] = entry
    if nisn:
        enriched_answers["_nisn"] = nisn
    if needs_review:
        enriched_answers["_needs_review"] = needs_review

    # Save original scan image permanently
    if scan_file_id:
        for ext in (".png", ".jpg"):
            src = os.path.join(UPLOAD_SCAN_DIR, "tmp", scan_file_id + ext)
            if os.path.exists(src):
                dst_dir = os.path.join(UPLOAD_SCAN_DIR, exam_id)
                os.makedirs(dst_dir, exist_ok=True)
                dst = os.path.join(dst_dir, student_id + ext)
                try:
                    os.rename(src, dst)
                    enriched_answers["_scan_image"] = f"/static/uploads/scans/{exam_id}/{student_id}{ext}"
                except Exception as e:
                    current_app.logger.warning("Failed to move scan image: %s", e)
                break

    # Check for existing submission and update or create
    existing = supabase.table("submissions") \
        .select("*") \
        .eq("exam_id", exam_id) \
        .eq("student_id", student_id) \
        .execute().data

    update_data = {
        "answers": enriched_answers,
        "score": score,
        "max_score": mcq_count,
        "status": "graded",
    }

    if existing:
        sub = supabase.table("submissions") \
            .update(update_data) \
            .eq("id", existing[0]["id"]) \
            .execute().data
    else:
        update_data.update({"exam_id": exam_id, "student_id": student_id})
        sub = supabase.table("submissions") \
            .insert(update_data) \
            .execute().data

    return jsonify({
        "success": True,
        "score": score,
        "correct": correct,
        "total": mcq_count,
        "needs_review": len(needs_review),
        "submission": sub[0] if sub else None,
    })


@api_bp.route("/scan/image/<exam_id>/<student_id>", methods=["GET"])
@login_required
def scan_image(exam_id, student_id):
    """Serve the annotated scan image with answer overlays."""
    for ext in (".png", ".jpg"):
        path = os.path.join(UPLOAD_SCAN_DIR, exam_id, student_id + ext)
        if os.path.exists(path):
            from app.services.omr_service import load_image, find_registration_marks
            img = load_image(open(path, "rb").read())
            if img is not None:
                # Get answers from submission
                supabase = get_supabase()
                sub = supabase.table("submissions").select("answers").eq("exam_id", exam_id).eq("student_id", student_id).limit(1).execute().data
                answers = {}
                if sub and sub[0].get("answers"):
                    raw = sub[0]["answers"]
                    if isinstance(raw, str):
                        raw = json.loads(raw)
                    for k, v in raw.items():
                        if isinstance(v, dict) and "answer" in v:
                            answers[k] = v["answer"]
                        elif not k.startswith("_"):
                            answers[k] = v
                # Draw annotated image
                from app.services.omr_service import draw_debug_image
                corners = find_registration_marks(img)
                annotated = draw_debug_image(img, corners, answers)
                return send_file(
                    io.BytesIO(annotated),
                    mimetype="image/jpeg",
                    as_attachment=False,
                )
            return send_file(path, mimetype=f"image/{ext[1:]}")
    return jsonify({"error": "Scan image not found"}), 404


@api_bp.route("/transaction/status", methods=["GET"])
@login_required
def transaction_status():
    order_id = request.args.get("order_id", "")
    if not order_id:
        return jsonify({"status": "error", "message": "order_id required"}), 400
    supabase = get_supabase()
    try:
        res = supabase.table("payment_transactions").select("status, gross_amount, activation_code, payment_type, payment_details").eq("order_id", order_id).limit(1).execute()
        if not res.data:
            return jsonify({"status": "not_found"})

        tx = res.data[0]
        # If pending, also check Midtrans for latest status
        if tx.get("status") == "pending":
            try:
                from app.services.midtrans_service import _load_midtrans_config
                cfg = _load_midtrans_config()
                if cfg.get("server_key"):
                    import midtransclient
                    snap = midtransclient.Snap(
                        is_production=cfg.get("is_production", False),
                        server_key=cfg["server_key"],
                        client_key=cfg.get("client_key", ""),
                    )
                    status_resp = snap.transaction.status(order_id)
                    trans_status = status_resp.get("transaction_status", "")
                    fraud_status = status_resp.get("fraud_status", "")
                    payment_type = status_resp.get("payment_type", "")

                    # Store payment details (VA, etc.)
                    details = {}
                    va = status_resp.get("va_numbers")
                    if va:
                        details["va_numbers"] = va
                    permata_va = status_resp.get("permata_va_number")
                    if permata_va:
                        details["permata_va"] = permata_va
                    payment_code = status_resp.get("payment_code")
                    if payment_code:
                        details["payment_code"] = payment_code

                    new_status = tx["status"]
                    if trans_status in ("settlement", "capture") and fraud_status != "deny":
                        new_status = "success"
                    elif trans_status == "expire":
                        new_status = "expired"
                    elif trans_status in ("deny", "cancel", "failure"):
                        new_status = "failure"

                    update = {"payment_type": payment_type, "payment_details": details}
                    if new_status != tx["status"]:
                        update["status"] = new_status
                        if new_status == "success":
                            update["activation_code"] = None  # will be set by _activate_subscription

                    supabase.table("payment_transactions").update(update).eq("id", tx["id"]).execute()

                    if new_status == "success" and tx.get("status") != "success":
                        from app.services.midtrans_service import _activate_subscription
                        _activate_subscription(tx["school_id"], tx.get("plan_id"), order_id, supabase)

                    tx["status"] = new_status
                    tx["payment_type"] = payment_type
                    tx["payment_details"] = details
            except Exception as e:
                current_app.logger.error(f"Midtrans status check error: {e}")

        return jsonify(tx)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)[:60]}), 500


@api_bp.route("/activation/redeem", methods=["POST"])
@login_required
def redeem_activation_code():
    data = request.get_json()
    code = (data or {}).get("code", "").strip().upper()
    if not code:
        return jsonify({"success": False, "message": "Kode aktivasi tidak boleh kosong"}), 400

    school_id = g.get("user_school_id")
    if not school_id:
        return jsonify({"success": False, "message": "Sekolah tidak terdaftar"}), 400

    supabase = get_supabase()

    # Find transaction with this activation code
    tx = supabase.table("payment_transactions") \
        .select("*") \
        .eq("activation_code", code) \
        .limit(1) \
        .execute()

    if not tx.data:
        # Also check school_subscriptions
        sub = supabase.table("school_subscriptions") \
            .select("*") \
            .eq("activation_code", code) \
            .order("created_at", desc=True) \
            .limit(1) \
            .execute()
        if sub.data:
            return jsonify({"success": False, "message": "Kode aktivasi ini sudah digunakan"}), 400
        return jsonify({"success": False, "message": "Kode aktivasi tidak ditemukan. Periksa kembali kode Anda."}), 404

    tx_data = tx.data[0]
    if tx_data["school_id"] != school_id:
        return jsonify({"success": False, "message": "Kode aktivasi bukan untuk sekolah Anda"}), 403

    # Check if already used
    existing = supabase.table("school_subscriptions") \
        .select("id") \
        .eq("activation_code", code) \
        .limit(1) \
        .execute()
    if existing.data:
        return jsonify({"success": False, "message": "Kode aktivasi sudah digunakan sebelumnya"}), 400

    # Activate
    from app.services.midtrans_service import _activate_subscription
    _activate_subscription(school_id, tx_data.get("plan_id"), tx_data["order_id"], supabase)

    return jsonify({"success": True, "message": "Langganan berhasil diaktifkan!"})


@api_bp.route("/ai/test-key", methods=["POST"])
@login_required
def ai_test_key():
    data = request.get_json()
    key_id = (data or {}).get("key_id", "")
    if not key_id:
        raise ValidationError("key_id", "Parameter key_id diperlukan")
    from app.services.ai_service import test_api_key
    result = test_api_key(g.user_id, key_id)
    if result.get("error"):
        app.logger.warning("AI key test failed: %s", result["error"], extra={"user_id": g.user_id, "key_id": key_id})
    return jsonify(result)


@api_bp.route("/ai/test-key-raw", methods=["POST"])
@login_required
def ai_test_key_raw():
    """Test a raw API key (before saving to DB)."""
    data = request.get_json() or {}
    api_key = data.get("api_key", "")
    provider = data.get("provider", "gemini")
    if not api_key:
        return jsonify({"error": "API Key tidak boleh kosong"}), 400
    from app.services.ai_service import _test_key_internal
    key = {"api_key": api_key, "provider": provider, "label": "Test"}
    result = _test_key_internal(key)
    return jsonify(result)


@api_bp.route("/grade/ai-suggest", methods=["POST"])
@login_required
def ai_suggest():
    data = request.get_json()
    if not data:
        raise ValidationError("request body", "Data tidak boleh kosong")
    question = data.get("question", "")
    answer = data.get("answer", "")
    max_score = data.get("max_score", 100)
    rubric = data.get("rubric", "")
    if not answer:
        raise ValidationError("answer", "Jawaban siswa kosong")
    from app.services.ai_service import suggest_grade
    result = suggest_grade(g.user_id, question, answer, max_score, rubric)
    if result.get("error"):
        raise AIProcessingError(result.get("provider", "unknown"), result["error"])
    return jsonify(result)


# ═══════════════════════════════════════════════════════════
# STUDENT BULK IMPORT (via API)
# ═══════════════════════════════════════════════════════════

@api_bp.route("/students/import", methods=["POST"])
@login_required
def api_import_students():
    """Bulk import students via CSV (API version, uses pandas).

    Accepts multipart/form-data with a 'csv_file' field.
    CSV must have columns: nama, nisn (required); kelas, email, password (optional).

    Returns: { success, failed, total, errors, message }
    """
    import pandas as pd

    school_id = g.get("user_school_id")
    if not school_id:
        return jsonify({"error": "Akses ditolak: sekolah tidak terdaftar"}), 403

    if "csv_file" not in request.files:
        return jsonify({"error": "File CSV diperlukan"}), 400

    file = request.files["csv_file"]
    if file.filename == "":
        return jsonify({"error": "Pilih file terlebih dahulu"}), 400

    if not file.filename.lower().endswith(".csv"):
        return jsonify({"error": "File harus berformat .csv"}), 422

    class_id = request.form.get("class_id") or None

    try:
        df = pd.read_csv(file, dtype=str).fillna("")
    except Exception as e:
        return jsonify({"error": f"Gagal membaca CSV: {str(e)[:100]}"}), 422

    # Validate required columns
    required = {"nama", "nisn"}
    missing_cols = required - set(df.columns.str.lower())
    if missing_cols:
        return jsonify({
            "error": f"Kolom wajib tidak ditemukan: {', '.join(sorted(missing_cols))}",
        }), 422

    # Normalize column names to lowercase
    df.columns = df.columns.str.lower()

    supabase = get_supabase()
    results = {"success": 0, "failed": 0, "total": len(df), "errors": []}

    # Process in chunks for memory efficiency
    chunk_size = 100
    for start in range(0, len(df), chunk_size):
        chunk = df.iloc[start:start + chunk_size]
        for idx, row in chunk.iterrows():
            row_num = idx + 2  # 1-indexed + header row
            try:
                nama = str(row.get("nama", "")).strip()
                nisn = str(row.get("nisn", "")).strip()
                if not nama or not nisn:
                    results["failed"] += 1
                    continue

                email = str(row.get("email", "")).strip() or None
                kelas = str(row.get("kelas", "")).strip()
                password = str(row.get("password", "")).strip() or "siswa123"

                # Check duplicate NISN
                existing = supabase.table("students") \
                    .select("id").eq("nisn", nisn) \
                    .eq("school_id", school_id).maybe_single().execute()
                if existing.data:
                    results["failed"] += 1
                    results["errors"].append({"row": row_num, "nisn": nisn, "message": "NISN sudah terdaftar"})
                    continue

                # Resolve class from name
                resolved_class_id = class_id
                if not resolved_class_id and kelas:
                    c = supabase.table("classes") \
                        .select("id").eq("school_id", school_id) \
                        .eq("name", kelas).maybe_single().execute()
                    if c.data:
                        resolved_class_id = c.data["id"]

                # Create auth user
                user_email = email or f"{nisn}@siswa.scan-grade.app"
                created = supabase.auth.admin.create_user({
                    "email": user_email, "password": password,
                    "user_metadata": {"role": "murid", "full_name": nama},
                    "email_confirm": True,
                })
                uid = created.user.id

                supabase.table("profiles").upsert({
                    "id": uid, "full_name": nama, "role": "murid",
                    "nisn": nisn, "school_id": school_id, "status": "active",
                }).execute()
                supabase.table("students").upsert({
                    "id": uid, "school_id": school_id, "nisn": nisn,
                    "class_id": resolved_class_id, "status": "active",
                }).execute()
                results["success"] += 1

            except Exception as e:
                results["failed"] += 1
                results["errors"].append({
                    "row": row_num,
                    "nisn": str(row.get("nisn", "")),
                    "message": str(e)[:100],
                })

    current_app.logger.info(
        "API import: %d success, %d failed of %d",
        results["success"], results["failed"], results["total"],
        extra={"school_id": str(school_id)},
    )
    return jsonify({
        "success": True,
        "results": results,
        "message": f"Berhasil: {results['success']}, Gagal: {results['failed']}",
    })


# ═══════════════════════════════════════════════════════════
# EXAM REPORT & EXPORT
# ═══════════════════════════════════════════════════════════

@api_bp.route("/exams/<exam_id>/report", methods=["GET"])
@login_required
def exam_report(exam_id):
    """Generate exam report with statistics. Supports ?format=excel for XLSX download.

    Returns JSON stats by default. With ?format=excel returns a formatted XLSX file.
    """
    from app.decorators.security import require_school_access

    # Manually apply school access check
    deco = require_school_access("exams", "exam_id")
    result = deco(lambda: None)(exam_id=exam_id)
    # If the decorator returns a non-None tuple, it's an error response
    if isinstance(result, tuple):
        return result

    supabase = get_supabase()
    exam = supabase.table("exams").select("*").eq("id", exam_id).single().execute().data
    if not exam:
        return jsonify({"error": "Ujian tidak ditemukan"}), 404

    subs = supabase.table("submissions") \
        .select("*, profiles!inner(full_name, nisn)") \
        .eq("exam_id", exam_id) \
        .in_("status", ["submitted", "graded", "published"]) \
        .execute().data or []

    # Build student list
    students = []
    scores = []
    for s in subs:
        profile = s.get("profiles") or {}
        score = s.get("final_score") or s.get("score") or 0
        students.append({
            "nama": profile.get("full_name", ""),
            "nisn": profile.get("nisn", ""),
            "nilai": float(score),
            "status": s.get("status", ""),
            "penalty": float(s.get("penalty") or 0),
        })
        scores.append(float(score))

    if not scores:
        stats = {"mean": 0, "median": 0, "highest": 0, "lowest": 0, "count": 0}
    else:
        sorted_s = sorted(scores)
        n = len(sorted_s)
        stats = {
            "mean": round(sum(scores) / n, 2),
            "median": round(sorted_s[n // 2] if n % 2 else (sorted_s[n // 2 - 1] + sorted_s[n // 2]) / 2, 2),
            "highest": max(scores),
            "lowest": min(scores),
            "count": n,
        }

    # Check format parameter
    fmt = request.args.get("format", "").lower()

    if fmt == "excel":
        return _generate_report_excel(exam, students, stats)

    # Default: JSON
    passing_score = exam.get("passing_score") or 0
    for s in students:
        s["keterangan"] = "Lulus" if s["nilai"] >= passing_score else "Tidak Lulus"

    return jsonify({
        "exam_title": exam.get("title", ""),
        "exam_id": exam_id,
        "stats": stats,
        "students": students,
        "passing_score": passing_score,
    })


def _generate_report_excel(exam, students, stats):
    """Generate formatted Excel report with Indonesian column names."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Nilai"

    # ── Styles ──
    title_font = Font(bold=True, size=14, color="1E293B")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
    stat_label_font = Font(bold=True, size=11, color="1E293B")
    stat_value_font = Font(size=11, color="334155")
    border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    # ── Title ──
    ws.merge_cells("A1:E1")
    ws["A1"] = f"Laporan Nilai — {exam.get('title', 'Ujian')}"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    # ── Statistics section ──
    stat_start = 3
    row = stat_start
    for label, key in [("Rata-rata", "mean"), ("Median", "median"), ("Tertinggi", "highest"), ("Terendah", "lowest"), ("Jumlah Siswa", "count")]:
        ws.cell(row=row, column=1, value=label).font = stat_label_font
        ws.cell(row=row, column=2, value=stats.get(key, 0)).font = stat_value_font
        row += 1

    # ── Student table header ──
    header_row = row + 1
    headers = ["No", "Nama Siswa", "NISN", "Nilai", "Keterangan"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # ── Student data ──
    passing_score = exam.get("passing_score") or 0
    for i, s in enumerate(students, 1):
        row_idx = header_row + i
        ws.cell(row=row_idx, column=1, value=i).border = border
        ws.cell(row=row_idx, column=2, value=s["nama"]).border = border
        ws.cell(row=row_idx, column=3, value=s.get("nisn", "")).border = border
        ws.cell(row=row_idx, column=4, value=s["nilai"]).border = border
        ket = "Lulus" if s["nilai"] >= passing_score else "Tidak Lulus"
        ws.cell(row=row_idx, column=5, value=ket).border = border

    # ── Auto-width ──
    for col in range(1, 6):
        max_len = len(str(headers[col - 1]))
        for row_idx in range(header_row, header_row + len(students) + 1):
            val = ws.cell(row=row_idx, column=col).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    title_slug = "".join(c for c in exam.get("title", "laporan") if c.isalnum() or c in " _").strip()[:20]
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"laporan_{title_slug}.xlsx",
    )


# ── AI Essay Grading API ──

@api_bp.route("/ai/grade-essay", methods=["POST"])
@login_required
def ai_grade_essay():
    """Grade a single essay answer."""
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data"}), 400

    submission_id = data.get("submission_id")
    question_index = data.get("question_index", 0)
    question_text = data.get("question_text", "")
    student_answer = data.get("student_answer", "")
    max_score = int(data.get("max_score", 100))
    rubric = data.get("rubric", "")
    diagram_context = data.get("diagram_context", "")
    lang = data.get("lang", "en")

    if not student_answer:
        return jsonify({"error": "Tidak ada jawaban siswa"}), 400

    from app.services.ai_grading import grade_essay as ai_grade
    result = ai_grade(
        teacher_id=g.user_id,
        submission_id=submission_id,
        question_index=question_index,
        question_text=question_text,
        student_answer=student_answer,
        max_score=max_score,
        rubric=rubric,
        diagram_context=diagram_context,
        lang=lang,
    )
    if "error" in result:
        return jsonify(result), 422
    return jsonify(result)


@api_bp.route("/ai/grade-bulk", methods=["POST"])
@login_required
def ai_grade_bulk():
    """Grade all pending essay questions for an exam or submission list."""
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data"}), 400

    exam_id = data.get("exam_id")
    submission_ids = data.get("submission_ids")

    if not exam_id:
        return jsonify({"error": "exam_id required"}), 400

    from app.services.ai_grading import grade_bulk_essays
    result = grade_bulk_essays(
        teacher_id=g.user_id,
        exam_id=exam_id,
        submission_ids=submission_ids,
    )
    return jsonify(result)

@api_bp.route("/broadcast/send", methods=["POST"])
@login_required
def api_send_broadcast():
    """Send broadcast notification."""
    data = request.get_json() or {}
    title = str(data.get("title", "")).strip()
    message = str(data.get("message", "")).strip()
    target_role = data.get("target_role", "")
    target_school_id = data.get("school_id", None)

    if not message:
        return jsonify({"error": "Pesan wajib diisi"}), 400
    if not title:
        title = (message[:50] + '...') if len(message) > 50 else message

    supabase = get_supabase()
    role = g.get("user_role")
    school_id = g.get("user_school_id")

    if role == "guru":
        if target_role not in ("murid", "guru"):
            return jsonify({"error": "Guru hanya bisa kirim ke guru/murid"}), 403
        target_school_id = school_id
    elif role == "admin_sekolah":
        if target_role not in ("guru", "murid"):
            return jsonify({"error": "Admin hanya bisa kirim ke guru/murid"}), 403
        target_school_id = school_id
    elif role == "murid":
        if target_role not in ("guru", "admin_sekolah"):
            return jsonify({"error": "Murid hanya bisa kirim ke guru/admin"}), 403
        target_school_id = school_id
    elif role != "super_admin":
        return jsonify({"error": "Unauthorized"}), 403

    # Build recipient list first (before notification insert)
    recipient_ids = data.get("recipient_ids")
    recipients = []
    if recipient_ids and isinstance(recipient_ids, list) and len(recipient_ids) > 0:
        if role == "murid":
            try:
                valid = supabase.table("profiles").select("id, role").in_("id", recipient_ids).execute().data or []
                recipients = [{"id": r["id"]} for r in valid if r.get("role") in ("guru", "admin_sekolah", "super_admin")]
            except Exception:
                recipients = []
        elif target_school_id:
            try:
                valid = supabase.table("profiles").select("id").in_("id", recipient_ids).eq("school_id", target_school_id).execute().data or []
                recipients = [{"id": r["id"]} for r in valid]
            except Exception:
                recipients = [{"id": uid} for uid in recipient_ids]
        else:
            recipients = [{"id": uid} for uid in recipient_ids]
    else:
        # Broadcast to all matching role + school
        try:
            query = supabase.table("profiles").select("id")
            if target_role and target_role != "all":
                query = query.eq("role", target_role)
            if target_school_id:
                query = query.eq("school_id", target_school_id)
            recipients = query.execute().data or []
        except Exception:
            try:
                query = supabase.table("profiles").select("id")
                if target_role and target_role != "all":
                    query = query.eq("role", target_role)
                recipients = query.execute().data or []
            except Exception:
                pass

    # For one-on-one messages, create/update conversation FIRST
    conversation_id = None
    if recipient_ids and len(recipient_ids) == 1 and target_role and target_role != "all":
        other_id = recipient_ids[0]
        try:
            conv1 = supabase.table("conversations").select("id").eq("participant_1", g.user_id).eq("participant_2", other_id).eq("title", title).order("created_at", desc=True).limit(1).execute().data or []
            conv2 = supabase.table("conversations").select("id").eq("participant_1", other_id).eq("participant_2", g.user_id).eq("title", title).order("created_at", desc=True).limit(1).execute().data or []
            conv = conv1 + conv2
            if conv:
                conversation_id = conv[0]["id"]
                if conv[0].get("status") != "open":
                    supabase.table("conversations").update({
                        "status": "open", "completed_at": None, "completed_by": None
                    }).eq("id", conversation_id).execute()
                supabase.table("conversations").update({
                    "last_message_at": datetime.now(timezone.utc).isoformat()
                }).eq("id", conversation_id).execute()
            else:
                conv_res = supabase.table("conversations").insert({
                    "participant_1": g.user_id, "participant_2": other_id,
                    "title": title, "status": "open",
                }).execute()
                conversation_id = conv_res.data[0]["id"]
        except Exception as e:
            return jsonify({"error": f"Gagal membuat percakapan: {str(e)[:100]}"}), 500

    # Insert notification WITH conversation_id already set
    notif = {
        "sender_id": g.user_id, "sender_role": role,
        "title": title, "message": message,
        "target_role": target_role if target_role != "all" else None,
        "target_school_id": target_school_id,
    }
    if conversation_id:
        notif["conversation_id"] = conversation_id
    try:
        res = supabase.table("notifications").insert(notif).execute()
        notif_id = res.data[0]["id"]
    except Exception as e:
        return jsonify({"error": f"Gagal: {str(e)[:100]}"}), 500

    # Create individual recipient entries for read tracking
    if recipients:
        rec_data = [{"notification_id": notif_id, "recipient_id": r["id"]} for r in recipients]
        for i in range(0, len(rec_data), 100):
            try:
                supabase.table("notification_recipients").upsert(rec_data[i:i+100], ignore_duplicates=True).execute()
            except Exception:
                pass

    resp = {"success": True, "recipients": len(recipients)}
    if conversation_id:
        resp["conversation_id"] = conversation_id
    return jsonify(resp)


@api_bp.route("/broadcast/reply", methods=["POST"])
@login_required
def api_reply_broadcast():
    """Reply to a conversation thread."""
    data = request.get_json() or {}
    conversation_id = data.get("conversation_id")
    message = str(data.get("message", "")).strip()
    if not conversation_id or not message:
        return jsonify({"error": "Data tidak lengkap"}), 400

    supabase = get_supabase()
    try:
        # Verify user is a participant
        conv = supabase.table("conversations").select("id, participant_1, participant_2, status, title").eq("id", conversation_id).single().execute().data
        if not conv:
            return jsonify({"error": "Percakapan tidak ditemukan"}), 404
        uid = g.user_id
        if conv["status"] != "open":
            update_data = {"status": "open", "completed_at": None, "completed_by": None, "last_message_at": datetime.now(timezone.utc).isoformat()}
            supabase.table("conversations").update(update_data).eq("id", conversation_id).execute()
        if uid != conv["participant_1"] and uid != conv["participant_2"]:
            return jsonify({"error": "Anda bukan peserta percakapan ini"}), 403

        other_id = conv["participant_1"] if uid == conv["participant_2"] else conv["participant_2"]
        role = g.get("user_role")

        # Use original conversation title, not reply title
        orig_title = conv.get("title", "")
        notif = {
            "sender_id": uid, "sender_role": role,
            "title": orig_title, "message": message,
            "target_role": None, "target_school_id": None,
            "conversation_id": conversation_id,
        }
        res = supabase.table("notifications").insert(notif).execute()
        notif_id = res.data[0]["id"]

        # Add recipient = the other participant
        supabase.table("notification_recipients").insert({
            "notification_id": notif_id, "recipient_id": other_id
        }).execute()

        # Update conversation last_message_at only (keep original title)
        supabase.table("conversations").update({
            "last_message_at": datetime.now(timezone.utc).isoformat()
        }).eq("id", conversation_id).execute()

        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": f"Gagal: {str(e)[:100]}"}), 500


@api_bp.route("/broadcast/complete", methods=["POST"])
@login_required
def api_complete_conversation():
    """Mark a conversation as complete."""
    data = request.get_json() or {}
    conversation_id = data.get("conversation_id")
    if not conversation_id:
        return jsonify({"error": "conversation_id diperlukan"}), 400
    supabase = get_supabase()
    try:
        conv = supabase.table("conversations").select("id, participant_1, participant_2, status").eq("id", conversation_id).single().execute().data
        if not conv:
            return jsonify({"error": "Percakapan tidak ditemukan"}), 404
        uid = g.user_id
        if uid != conv["participant_1"] and uid != conv["participant_2"]:
            return jsonify({"error": "Anda bukan peserta percakapan ini"}), 403
        if conv["status"] != "open":
            return jsonify({"error": "Percakapan sudah selesai"}), 400
        supabase.table("conversations").update({
            "status": "completed", "completed_at": datetime.now(timezone.utc).isoformat(), "completed_by": uid
        }).eq("id", conversation_id).execute()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)[:100]}), 500


@api_bp.route("/broadcast/edit-message", methods=["POST"])
@login_required
def api_edit_broadcast_message():
    """Edit a message within 2 minutes of sending."""
    data = request.get_json() or {}
    notification_id = data.get("notification_id")
    new_message = str(data.get("message", "")).strip()
    if not notification_id or not new_message:
        return jsonify({"error": "Data tidak lengkap"}), 400
    supabase = get_supabase()
    try:
        notif = supabase.table("notifications").select("id, sender_id, message, created_at, conversation_id").eq("id", notification_id).single().execute().data
        if not notif:
            return jsonify({"error": "Pesan tidak ditemukan"}), 404
        if notif["sender_id"] != g.user_id:
            return jsonify({"error": "Anda bukan pengirim pesan ini"}), 403
        from datetime import datetime, timezone, timedelta
        created = datetime.fromisoformat(notif["created_at"].replace("Z", "+00:00"))
        if datetime.now(timezone.utc) - created > timedelta(minutes=2):
            return jsonify({"error": "Batas edit 2 menit telah berlalu"}), 400
        supabase.table("notifications").update({"message": new_message}).eq("id", notification_id).execute()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)[:100]}), 500


@api_bp.route("/broadcast/unsend", methods=["POST"])
@login_required
def api_unsend_message():
    """Unsend a message for everyone (delete for all, within 2 minutes)."""
    data = request.get_json() or {}
    notification_id = data.get("notification_id")
    if not notification_id:
        return jsonify({"error": "notification_id diperlukan"}), 400
    supabase = get_supabase()
    try:
        notif = supabase.table("notifications").select("id, sender_id, created_at, conversation_id").eq("id", notification_id).single().execute().data
        if not notif:
            return jsonify({"error": "Pesan tidak ditemukan"}), 404
        if notif["sender_id"] != g.user_id:
            return jsonify({"error": "Anda bukan pengirim pesan ini"}), 403
        if not notif.get("conversation_id"):
            return jsonify({"error": "Hanya bisa unsend di percakapan"}), 400
        from datetime import datetime, timezone, timedelta
        created = datetime.fromisoformat(notif["created_at"].replace("Z", "+00:00"))
        if datetime.now(timezone.utc) - created > timedelta(minutes=2):
            return jsonify({"error": "Batas unsend 2 menit telah berlalu"}), 400
        supabase.table("notifications").update({
            "message": "Pesan telah dihapus"
        }).eq("id", notification_id).execute()
        try:
            supabase.table("notifications").update({
                "is_deleted": True, "deleted_at": datetime.now(timezone.utc).isoformat(), "deleted_by": g.user_id
            }).eq("id", notification_id).execute()
        except Exception:
            pass
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)[:100]}), 500


@api_bp.route("/broadcast/delete-for-me", methods=["POST"])
@login_required
def api_delete_message_for_me():
    """Hide a message for the current user only."""
    data = request.get_json() or {}
    notification_id = data.get("notification_id")
    if not notification_id:
        return jsonify({"error": "notification_id diperlukan"}), 400
    supabase = get_supabase()
    try:
        notif = supabase.table("notifications").select("id, conversation_id").eq("id", notification_id).single().execute().data
        if not notif:
            return jsonify({"error": "Pesan tidak ditemukan"}), 404
        uid = g.user_id
        try:
            existing = supabase.table("message_hides").select("id").eq("notification_id", notification_id).eq("user_id", uid).maybe_single().execute().data
            if not existing:
                supabase.table("message_hides").insert({"notification_id": notification_id, "user_id": uid}).execute()
        except Exception:
            pass
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)[:100]}), 500


@api_bp.route("/broadcast/conversation/delete", methods=["POST"])
@login_required
def api_delete_conversation():
    """Delete entire conversation for current user."""
    data = request.get_json() or {}
    conversation_id = data.get("conversation_id")
    if not conversation_id:
        return jsonify({"error": "conversation_id diperlukan"}), 400
    supabase = get_supabase()
    uid = g.user_id
    try:
        conv = supabase.table("conversations").select("id, participant_1, participant_2").eq("id", conversation_id).single().execute().data
        if not conv:
            return jsonify({"error": "Percakapan tidak ditemukan"}), 404
        if uid != conv["participant_1"] and uid != conv["participant_2"]:
            return jsonify({"error": "Anda bukan peserta percakapan ini"}), 403
        now = datetime.now(timezone.utc).isoformat()
        del_key = "deleted_at_p1" if uid == conv["participant_1"] else "deleted_at_p2"
        try:
            supabase.table("conversations").update({del_key: now}).eq("id", conversation_id).execute()
        except Exception:
            supabase.table("conversations").update({"status": "archived"}).eq("id", conversation_id).execute()
            return jsonify({"success": True})
        conv2 = supabase.table("conversations").select("deleted_at_p1, deleted_at_p2").eq("id", conversation_id).single().execute().data
        if conv2.get("deleted_at_p1") and conv2.get("deleted_at_p2"):
            supabase.table("conversations").delete().eq("id", conversation_id).execute()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)[:100]}), 500


@api_bp.route("/broadcast/conversation/archive", methods=["POST"])
@login_required
def api_archive_conversation():
    """Archive a conversation (hide from main list)."""
    data = request.get_json() or {}
    conversation_id = data.get("conversation_id")
    if not conversation_id:
        return jsonify({"error": "conversation_id diperlukan"}), 400
    supabase = get_supabase()
    try:
        conv = supabase.table("conversations").select("id, participant_1, participant_2, status").eq("id", conversation_id).single().execute().data
        if not conv:
            return jsonify({"error": "Percakapan tidak ditemukan"}), 404
        uid = g.user_id
        if uid != conv["participant_1"] and uid != conv["participant_2"]:
            return jsonify({"error": "Anda bukan peserta percakapan ini"}), 403
        if conv["status"] != "open":
            return jsonify({"error": "Hanya percakapan aktif yang bisa diarsipkan"}), 400
        supabase.table("conversations").update({"status": "archived"}).eq("id", conversation_id).execute()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)[:100]}), 500


@api_bp.route("/broadcast/conversation/unarchive", methods=["POST"])
@login_required
def api_unarchive_conversation():
    """Restore an archived conversation."""
    data = request.get_json() or {}
    conversation_id = data.get("conversation_id")
    if not conversation_id:
        return jsonify({"error": "conversation_id diperlukan"}), 400
    supabase = get_supabase()
    try:
        conv = supabase.table("conversations").select("id, participant_1, participant_2, status").eq("id", conversation_id).single().execute().data
        if not conv:
            return jsonify({"error": "Percakapan tidak ditemukan"}), 404
        uid = g.user_id
        if uid != conv["participant_1"] and uid != conv["participant_2"]:
            return jsonify({"error": "Anda bukan peserta percakapan ini"}), 403
        if conv["status"] != "archived":
            return jsonify({"error": "Percakapan tidak dalam status arsip"}), 400
        supabase.table("conversations").update({"status": "open"}).eq("id", conversation_id).execute()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)[:100]}), 500


@api_bp.route("/broadcast/conversations", methods=["GET"])
@login_required
def api_conversations():
    """Get conversation threads for current user."""
    supabase = get_supabase()
    uid = g.user_id
    include_archived = request.args.get("archived", "false") == "true"
    try:
        convs = supabase.table("conversations").select("id, participant_1, participant_2, title, status, last_message_at, created_at, completed_at, completed_by") \
            .or_("participant_1.eq." + uid + ",participant_2.eq." + uid) \
            .order("last_message_at", desc=True).limit(50).execute().data or []

        # Get hidden message IDs for this user (migration 021)
        hidden_ids = set()
        try:
            hides = supabase.table("message_hides").select("notification_id").eq("user_id", uid).execute().data or []
            hidden_ids = set(h["notification_id"] for h in hides)
        except Exception:
            pass

        role = g.get("user_role")
        result = []
        for c in convs:
            # Skip soft-deleted (migration 021) — try both columns, skip if present
            try:
                del_col = "deleted_at_p1" if uid == c["participant_1"] else "deleted_at_p2"
                if c.get(del_col):
                    continue
            except Exception:
                pass
            other_id = c["participant_1"] if uid == c["participant_2"] else c["participant_2"]
            other_name = other_id[:8]
            other_role = None
            try:
                p = supabase.table("profiles").select("full_name, role").eq("id", other_id).single().execute().data
                if p:
                    other_name = p.get("full_name", other_id[:8])
                    other_role = p.get("role")
            except Exception:
                pass
            # Students can only converse with teachers/admins, not other students
            if role == "murid" and other_role == "murid":
                continue
            # Get messages in this conversation
            msgs = []
            try:
                msgs_data = supabase.table("notifications").select("id, sender_id, sender_role, title, message, created_at") \
                    .eq("conversation_id", c["id"]).order("created_at", asc=True).limit(100).execute().data or []
                for m in msgs_data:
                    if m["id"] in hidden_ids:
                        continue
                    msg_text = m.get("message", "")
                    msgs.append({
                        "id": m["id"], "sender_id": m.get("sender_id"),
                        "sender_role": m.get("sender_role"),
                        "title": m.get("title"), "message": msg_text,
                        "is_mine": m.get("sender_id") == uid,
                        "created_at": str(m.get("created_at", ""))[:19].replace("T", " "),
                    })
            except Exception:
                pass
            # Determine has_unread from notification_recipients.read_at
            has_unread = False
            unread_count = 0
            try:
                nids_unread = []
                for m in msgs:
                    if not m.get("is_mine"):
                        nids_unread.append(m["id"])
                if nids_unread:
                    unread_check = supabase.table("notification_recipients").select("id", count="exact") \
                        .eq("recipient_id", uid).is_("read_at", "null").in_("notification_id", nids_unread).execute()
                    has_unread = (unread_check.count or 0) > 0
                    unread_count = unread_check.count or 0
            except Exception:
                pass
            # Last message preview (truncated)
            last_msg_preview = ""
            if msgs:
                last_msg = msgs[-1].get("message", "")
                if last_msg:
                    last_msg_preview = last_msg[:80] + "..." if len(last_msg) > 80 else last_msg
            is_archived = c.get("status") == "archived"
            if is_archived and not include_archived:
                continue
            result.append({
                "id": c["id"], "title": c["title"], "status": c["status"],
                "other_name": other_name, "other_id": other_id, "other_role": other_role,
                "last_message_at": str(c.get("last_message_at", ""))[:19].replace("T", " "),
                "created_at": str(c.get("created_at", ""))[:19].replace("T", " "),
                "completed_at": str(c.get("completed_at", ""))[:19].replace("T", " ") if c.get("completed_at") else None,
                "messages": msgs,
                "last_msg_preview": last_msg_preview,
                "unread": has_unread,
                "has_unread": has_unread,
                "unread_count": unread_count,
                "is_archived": is_archived,
            })
        return jsonify({"conversations": result})
    except Exception as e:
        return jsonify({"error": str(e)[:100], "conversations": []}), 500


@api_bp.route("/broadcast/contacts", methods=["GET"])
@login_required
def api_broadcast_contacts():
    """Get contacts the current user can message, merged with conversation info."""
    supabase = get_supabase()
    uid = g.user_id
    role = g.get("user_role")
    school_id = g.get("user_school_id")

    roles_map = {
        "murid": ["guru", "admin_sekolah"],
        "guru": ["murid"],
        "admin_sekolah": ["guru", "murid"],
        "super_admin": ["guru", "murid", "admin_sekolah", "super_admin"],
    }
    target_roles = roles_map.get(role, [])
    if not target_roles:
        return jsonify({"contacts": []})

    try:
        query = supabase.table("profiles").select("id, full_name, role").in_("role", target_roles)
        if school_id and role != "super_admin":
            query = query.eq("school_id", school_id)
        profiles = query.order("full_name").execute().data or []
    except Exception:
        try:
            query = supabase.table("profiles").select("id, full_name, role").in_("role", target_roles)
            profiles = query.order("full_name").execute().data or []
        except Exception:
            profiles = []

    profile_map = {p["id"]: p for p in profiles}

    convs = []
    try:
        convs = supabase.table("conversations").select("id, participant_1, participant_2, title, status, last_message_at, created_at") \
            .or_("participant_1.eq." + uid + ",participant_2.eq." + uid) \
            .order("last_message_at", desc=True).limit(50).execute().data or []
    except Exception:
        pass

    hidden_ids = set()
    try:
        hides = supabase.table("message_hides").select("notification_id").eq("user_id", uid).execute().data or []
        hidden_ids = set(h["notification_id"] for h in hides)
    except Exception:
        pass

    contacts = []
    for p_id, p in profile_map.items():
        conv_id = None
        last_msg_preview = None
        unread_count = 0
        last_message_at = None
        for c in convs:
            other_id = c["participant_1"] if uid == c["participant_2"] else c["participant_2"]
            if other_id == p_id:
                conv_id = c["id"]
                last_message_at = str(c.get("last_message_at", ""))[:19].replace("T", " ")
                try:
                    msgs = supabase.table("notifications").select("id, message, sender_id, created_at") \
                        .eq("conversation_id", c["id"]).order("created_at", desc=True).limit(1).execute().data or []
                    if msgs:
                        last = msgs[0]
                        if last["id"] not in hidden_ids:
                            msg_text = last.get("message", "")
                            if msg_text:
                                last_msg_preview = msg_text[:80] + "..." if len(msg_text) > 80 else msg_text
                    nids = []
                    for m in supabase.table("notifications").select("id, sender_id") \
                            .eq("conversation_id", c["id"]).neq("sender_id", uid).execute().data or []:
                        nids.append(m["id"])
                    if nids:
                        chk = supabase.table("notification_recipients").select("id", count="exact") \
                            .eq("recipient_id", uid).is_("read_at", "null").in_("notification_id", nids).execute()
                        unread_count = chk.count or 0
                except Exception:
                    pass
                break
        contacts.append({
            "id": p_id,
            "full_name": p.get("full_name", p_id[:8]),
            "role": p.get("role"),
            "conversation_id": conv_id,
            "last_msg_preview": last_msg_preview or "",
            "unread_count": unread_count,
            "last_message_at": last_message_at or "",
        })

    contacts.sort(key=lambda x: (x["last_message_at"] or ""), reverse=True)
    return jsonify({"contacts": contacts})


@api_bp.route("/broadcast/update", methods=["POST"])
@login_required
def api_update_broadcast():
    """Edit notification title/message."""
    data = request.get_json() or {}
    notif_id = data.get("id")
    title = str(data.get("title", "")).strip()
    message = str(data.get("message", "")).strip()
    if not notif_id or not title or not message:
        return jsonify({"error": "Data tidak lengkap"}), 400
    supabase = get_supabase()
    role = g.get("user_role")
    try:
        existing = supabase.table("notifications").select("id,sender_id").eq("id", notif_id).single().execute().data
        if not existing:
            return jsonify({"error": "Notifikasi tidak ditemukan"}), 404
        if role != "super_admin" and existing.get("sender_id") != g.user_id:
            return jsonify({"error": "Tidak berhak mengedit"}), 403
        supabase.table("notifications").update({"title": title, "message": message}).eq("id", notif_id).execute()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": f"Gagal: {str(e)[:100]}"}), 500


@api_bp.route("/broadcast/delete", methods=["POST"])
@login_required
def api_delete_broadcast():
    """Delete a notification."""
    data = request.get_json() or {}
    notif_id = data.get("id")
    if not notif_id:
        return jsonify({"error": "ID notifikasi diperlukan"}), 400
    supabase = get_supabase()
    role = g.get("user_role")
    try:
        existing = supabase.table("notifications").select("id,sender_id").eq("id", notif_id).single().execute().data
        if not existing:
            return jsonify({"error": "Notifikasi tidak ditemukan"}), 404
        if role != "super_admin" and existing.get("sender_id") != g.user_id:
            return jsonify({"error": "Tidak berhak menghapus"}), 403
        supabase.table("notification_recipients").delete().eq("notification_id", notif_id).execute()
        supabase.table("notifications").delete().eq("id", notif_id).execute()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": f"Gagal: {str(e)[:100]}"}), 500


@api_bp.route("/broadcast/mark-read", methods=["POST"])
@login_required
def api_mark_read():
    """Mark a conversation's messages as read by the current user."""
    data = request.get_json() or {}
    conversation_id = data.get("conversation_id")
    if not conversation_id:
        return jsonify({"error": "conversation_id diperlukan"}), 400
    supabase = get_supabase()
    try:
        from datetime import datetime, timezone
        now_iso = datetime.now(timezone.utc).isoformat()
        notifs = supabase.table("notifications").select("id").eq("conversation_id", conversation_id).neq("sender_id", g.user_id).execute().data or []
        nids = [n["id"] for n in notifs]
        if nids:
            supabase.table("notification_recipients").update({"read_at": now_iso}).eq("recipient_id", g.user_id).in_("notification_id", nids).is_("read_at", "null").execute()
        return jsonify({"success": True, "marked": len(nids)})
    except Exception as e:
        return jsonify({"error": str(e)[:100]}), 500


@api_bp.route("/broadcast/mark-read-all", methods=["POST"])
@login_required
def api_mark_read_all():
    """Mark all user's unread notifications as read."""
    supabase = get_supabase()
    try:
        from datetime import datetime, timezone
        now_iso = datetime.now(timezone.utc).isoformat()
        res = supabase.table("notification_recipients").update({"read_at": now_iso}) \
            .eq("recipient_id", g.user_id).is_("read_at", "null").execute()
        marked = res.data if res.data else []
        return jsonify({"success": True, "marked": len(marked)})
    except Exception as e:
        return jsonify({"error": str(e)[:100]}), 500


@api_bp.route("/broadcast/list", methods=["GET"])
@login_required
def api_broadcast_list():
    """Get notifications for current user. ?type=sent|received filters."""
    supabase = get_supabase()
    user_id = g.user_id
    role = g.get("user_role")
    school_id = g.get("user_school_id")
    filter_type = request.args.get("type", "")

    notifs = []
    uid = g.user_id
    urole = g.get("user_role")

    # Notifications sent TO this user (direct recipients)
    if filter_type != "sent":
        try:
            direct = supabase.table("notification_recipients") \
                .select("notification_id, read_at") \
                .eq("recipient_id", user_id) \
                .order("notification_id", desc=True) \
                .limit(50) \
                .execute().data or []
            if direct:
                nids = [d["notification_id"] for d in direct]
                notif_map = {}
                try:
                    ndata = supabase.table("notifications") \
                        .select("id, sender_id, sender_role, target_role, title, message, created_at") \
                        .in_("id", nids) \
                        .execute().data or []
                    for n in ndata:
                        notif_map[n["id"]] = n
                except Exception:
                    pass
                for d in direct:
                    nid = d["notification_id"]
                    n = notif_map.get(nid, {})
                    notifs.append({
                        "id": nid, "title": n.get("title"),
                        "message": n.get("message"), "sender_role": n.get("sender_role"),
                        "target_role": n.get("target_role"),
                        "sender_id": n.get("sender_id"),
                        "created_at": str(n.get("created_at", ""))[:19].replace("T", " "),
                        "read": d.get("read_at") is not None,
                        "can_edit": urole == "super_admin" or n.get("sender_id") == uid,
                    })
        except Exception:
            pass

    # Role-based notifications (broadcasts to this role) - only for received
    if filter_type != "sent":
        try:
            role_notifs_query = supabase.table("notifications") \
                .select("id, sender_id, sender_role, target_role, title, message, created_at") \
                .in_("target_role", [role, None])
            if school_id:
                role_notifs_query = role_notifs_query.eq("target_school_id", school_id)
            role_notifs = role_notifs_query.order("created_at", desc=True).limit(50).execute().data or []
            existing_ids = {n["id"] for n in notifs}
            for n in role_notifs:
                if n["id"] not in existing_ids:
                    n["read"] = False
                    n["can_edit"] = urole == "super_admin" or n.get("sender_id") == uid
                    n["created_at"] = str(n.get("created_at", ""))[:19].replace("T", " ")
                    notifs.append(n)
        except Exception:
            pass

    # Notifications SENT BY this user
    if filter_type != "received":
        try:
            sent = supabase.table("notifications") \
                .select("id, sender_id, sender_role, target_role, title, message, created_at") \
                .eq("sender_id", user_id) \
                .order("created_at", desc=True) \
                .limit(50) \
                .execute().data or []
            sent_ids = {n["id"] for n in notifs}
            for n in sent:
                if n["id"] not in sent_ids:
                    n["read"] = True
                    n["can_edit"] = True
                    n["created_at"] = str(n.get("created_at", ""))[:19].replace("T", " ")
                    notifs.append(n)
        except Exception:
            pass

    notifs.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    return jsonify({"notifications": notifs[:50]})


@api_bp.route("/broadcast/unread", methods=["GET"])
@login_required
def api_broadcast_unread():
    """Get unread notification count (separate broadcast vs conversation)."""
    supabase = get_supabase()
    user_id = g.user_id
    broadcast_unread = 0
    conversation_unread = 0
    try:
        # Get all unread notification IDs for this user
        unread = supabase.table("notification_recipients").select("notification_id") \
            .eq("recipient_id", user_id).is_("read_at", "null").execute().data or []
        nids = [r["notification_id"] for r in unread]
        if nids:
            # Check which of these belong to conversations vs broadcasts
            conv_notifs = supabase.table("notifications").select("id") \
                .in_("id", nids).is_("conversation_id", "not.null").execute().data or []
            conv_nids = set(n["id"] for n in conv_notifs)
            broadcast_unread = sum(1 for nid in nids if nid not in conv_nids)
            conversation_unread = len(nids) - broadcast_unread
    except Exception:
        pass
    return jsonify({"unread": broadcast_unread + conversation_unread,
                    "broadcast_unread": broadcast_unread,
                    "conversation_unread": conversation_unread})


@api_bp.route("/broadcast/classes", methods=["GET"])
@login_required
def api_broadcast_classes():
    """Get classes in teacher's school."""
    supabase = get_supabase()
    school_id = g.get("user_school_id")
    try:
        query = supabase.table("classes").select("id, name, grade_level")
        if school_id:
            query = query.eq("school_id", school_id)
        classes = query.order("name").execute().data or []
    except Exception:
        classes = []
    return jsonify({"classes": classes})


@api_bp.route("/broadcast/stats", methods=["POST"])
@login_required
def api_broadcast_stats():
    """Get read/unread stats for a notification."""
    data = request.get_json() or {}
    notif_id = data.get("notification_id")
    if not notif_id:
        return jsonify({"error": "notification_id diperlukan"}), 400
    supabase = get_supabase()
    try:
        total = supabase.table("notification_recipients").select("id", count="exact") \
            .eq("notification_id", notif_id).execute()
        total_count = total.count or 0
        read = supabase.table("notification_recipients").select("id", count="exact") \
            .eq("notification_id", notif_id).is_("read_at", "not.null").execute()
        read_count = read.count or 0
        return jsonify({"total": total_count, "read": read_count, "unread": total_count - read_count})
    except Exception as e:
        return jsonify({"error": str(e)[:100]}), 500


@api_bp.route("/broadcast/students", methods=["GET"])
@login_required
def api_broadcast_students():
    """Get list of users for selective broadcast by role."""
    supabase = get_supabase()
    role = g.get("user_role")
    school_id = g.get("user_school_id")
    target_role = request.args.get("role", "murid")
    class_id = request.args.get("class_id")

    try:
        query = supabase.table("profiles").select("id, full_name").eq("status", "active")
        if target_role != "all":
            query = query.eq("role", target_role)
        if school_id:
            try:
                query = query.eq("school_id", school_id)
            except Exception:
                pass
        if class_id:
            try:
                query = query.eq("class_id", class_id)
            except Exception:
                pass
        users = query.order("full_name").execute().data or []
    except Exception:
        users = []
    return jsonify({"users": users})


@api_bp.route("/broadcast/teachers", methods=["GET"])
@login_required
def api_broadcast_teachers():
    """Get list of teachers for student one-on-one messaging."""
    supabase = get_supabase()
    school_id = g.get("user_school_id")
    try:
        query = supabase.table("profiles").select("id, full_name").eq("role", "guru").eq("status", "active")
        if school_id:
            try:
                query = query.eq("school_id", school_id)
            except Exception:
                pass
        teachers = query.order("full_name").execute().data or []
    except Exception:
        teachers = []
    return jsonify({"teachers": teachers})


@api_bp.route("/account/export-data", methods=["GET"])
@login_required
def api_export_data():
    """Export all personal data (UU PDP right to data portability)."""
    from app.services.data_retention_service import export_user_data
    data = export_user_data(g.user_id)
    return jsonify(data)


@api_bp.route("/account/delete-request", methods=["POST"])
@login_required
def api_delete_request():
    """Submit account deletion request (UU PDP right to erasure)."""
    from app.services.data_retention_service import request_deletion
    data = request.get_json() or {}
    reason = str(data.get("reason", "")).strip()
    result, status = request_deletion(g.user_id, reason)
    return jsonify(result), status


@api_bp.route("/account/delete-request/cancel", methods=["POST"])
@login_required
def api_cancel_delete_request():
    """Cancel a pending deletion request."""
    from app.services.data_retention_service import cancel_deletion_request
    result, status = cancel_deletion_request(g.user_id)
    return jsonify(result), status


@api_bp.route("/account/deletion-status", methods=["GET"])
@login_required
def api_deletion_status():
    """Check if user has a pending deletion request."""
    supabase = get_supabase()
    try:
        res = supabase.table("deletion_requests").select("id, status, reason, requested_at").eq("user_id", g.user_id).order("requested_at", desc=True).limit(1).execute().data
        return jsonify({"request": res[0] if res else None})
    except Exception:
        return jsonify({"request": None})


@api_bp.route("/admin/deletion-requests", methods=["GET"])
@login_required
def api_admin_deletion_requests():
    """List all deletion requests (admin only)."""
    role = g.get("user_role")
    if role not in ("super_admin", "admin_sekolah"):
        return jsonify({"error": "Forbidden"}), 403
    supabase = get_supabase()
    try:
        res = supabase.table("deletion_requests").select("id, user_id, reason, status, requested_at, notes").order("requested_at", desc=True).limit(100).execute().data or []
        enriched = []
        for r in res:
            uname = r["user_id"][:8]
            try:
                p = supabase.table("profiles").select("full_name").eq("id", r["user_id"]).single().execute().data
                if p:
                    uname = p.get("full_name", r["user_id"][:8])
            except Exception:
                pass
            enriched.append({**r, "user_name": uname})
        return jsonify({"requests": enriched})
    except Exception as e:
        return jsonify({"error": str(e)[:100], "requests": []}), 500


@api_bp.route("/admin/deletion-requests/process", methods=["POST"])
@login_required
def api_admin_process_deletion():
    """Approve or reject a deletion request (admin only)."""
    role = g.get("user_role")
    if role not in ("super_admin", "admin_sekolah"):
        return jsonify({"error": "Forbidden"}), 403
    data = request.get_json() or {}
    request_id = data.get("id")
    action = data.get("action")
    notes = str(data.get("notes", "")).strip()
    if not request_id or action not in ("approve", "reject"):
        return jsonify({"error": "Data tidak lengkap"}), 400
    from app.services.data_retention_service import process_deletion_request
    result, status = process_deletion_request(request_id, g.user_id, action, notes)
    return jsonify(result), status


@api_bp.route("/public/privacy-info")
def api_public_privacy_info():
    """Public endpoint: returns DPO contact and PSE registration number."""
    from app.utils.auth import get_supabase
    supabase = get_supabase()
    result = {"dpo_contact": "", "pse_reg_number": "", "data_controller_name": "", "data_controller_email": ""}
    try:
        rows = supabase.table("system_settings").select("key, value").in_("key", ["dpo_contact", "pse_reg_number", "data_controller_name", "data_controller_email"]).execute().data or []
        for r in rows:
            result[r["key"]] = r["value"]
    except Exception:
        pass
    return jsonify(result)
