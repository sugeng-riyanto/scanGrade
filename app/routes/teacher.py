import json
import io
import os
from datetime import datetime, timedelta, timezone
from flask import Blueprint, jsonify, render_template, request, redirect, url_for, flash, g, send_file, current_app
from app.utils.auth import teacher_or_admin_required, get_supabase, login_required, subscription_write_required
from app.decorators.security import require_school_access
from app.decorators.subscription import require_subscription
from app.services.export_service import export_to_xlsx, export_to_pdf
from app.services.answer_sheet_generator import generate_answer_sheet
from app.services.pdf_service import upload_pdf
from app.services.audit_service import log_activity

teacher_bp = Blueprint("teacher", __name__)


def _extract_mcq_answer(student_ans):
    if isinstance(student_ans, dict):
        return student_ans.get('answer', '')
    return student_ans or ''

def _is_mcq_correct(student_ans, key_val):
    ans = _extract_mcq_answer(student_ans)
    if key_val == "bonus":
        return bool(ans and str(ans).strip())
    if isinstance(key_val, list):
        if not ans:
            return False
        return ans in key_val
    return ans == key_val


def _recalculate_scores(exam_id):
    supabase = get_supabase()
    exam = supabase.table("exams").select("*").eq("id", exam_id).single().execute().data
    if not exam:
        return
    # Parse JSON fields that may be strings from Supabase
    for _fld in ("answer_key", "question_types", "question_weights", "question_pages"):
        _v = exam.get(_fld)
        if isinstance(_v, str):
            try:
                exam[_fld] = json.loads(_v)
            except (json.JSONDecodeError, TypeError):
                exam[_fld] = {}
    answer_key = exam.get("answer_key") or {}
    question_types = exam.get("question_types") or {}
    question_weights = exam.get("question_weights") or {}
    total_q = exam.get("total_questions", 0)
    if not question_weights and total_q > 0:
        mcq_count = sum(1 for i in range(total_q) if question_types.get(str(i), "mcq") == "mcq")
        essay_count = total_q - mcq_count
        mcq_pct = 70
        essay_pct = 30
        if mcq_count > 0 and essay_count > 0:
            pass
        elif mcq_count == 0:
            mcq_pct, essay_pct = 0, 100
        else:
            mcq_pct, essay_pct = 100, 0
        if mcq_count > 0:
            each = round(mcq_pct / mcq_count, 2)
            for i in range(total_q):
                if question_types.get(str(i), "mcq") == "mcq":
                    question_weights[str(i)] = each
        if essay_count > 0:
            each = round(essay_pct / essay_count, 2)
            for i in range(total_q):
                if question_types.get(str(i), "mcq") != "mcq":
                    question_weights[str(i)] = each
    subs = supabase.table("submissions").select("id, answers, penalty, teacher_feedback").eq("exam_id", exam_id).in_("status", ["submitted", "graded", "published"]).execute().data or []
    for sub in subs:
        # Parse JSON fields in each submission
        for _sf in ("answers", "teacher_feedback"):
            _sv = sub.get(_sf)
            if isinstance(_sv, str):
                try: sub[_sf] = json.loads(_sv)
                except (json.JSONDecodeError, TypeError): sub[_sf] = {}
        answers = sub.get("answers") or {}
        earned = 0.0
        for i in range(total_q):
            qtype = question_types.get(str(i), "mcq")
            key_val = answer_key.get(str(i))
            w = float(question_weights.get(str(i), 0))
            if w <= 0:
                continue
            if qtype == "mcq":
                if _is_mcq_correct(answers.get(str(i)), key_val):
                    earned += w
        fb = sub.get("teacher_feedback") or {}
        fb_scores = fb.get("scores", {}) or {}
        for qi, sv in fb_scores.items():
            if sv is not None and sv != "":
                ew = float(question_weights.get(str(qi), 0))
                if ew > 0:
                    earned += float(sv) / 100.0 * ew
        final = round(min(earned, 100), 2)
        penalty = float(sub.get("penalty") or 0)
        final = max(0, round(final - penalty, 2))
        mcq_correct = 0
        mcq_count = sum(1 for i in range(total_q) if question_types.get(str(i), "mcq") == "mcq")
        for i in range(total_q):
            qtype = question_types.get(str(i), "mcq")
            key_val = answer_key.get(str(i))
            if qtype == "mcq" and key_val:
                if _is_mcq_correct(answers.get(str(i)), key_val):
                    mcq_correct += 1
        mcq_score = round((mcq_correct / max(mcq_count, 1)) * 100, 2) if mcq_count > 0 else 0
        supabase.table("submissions").update({
            "score": mcq_score,
            "final_score": final,
        }).eq("id", sub["id"]).execute()


@teacher_bp.route("/dashboard")
@teacher_or_admin_required
def dashboard():
    supabase = get_supabase()
    res = supabase.table("exams").select("*").eq("teacher_id", g.user_id).order("created_at", desc=True).execute()
    exams = res.data or []

    exam_ids = [e["id"] for e in exams]
    total_students = 0
    all_scores = []
    pending_grading = 0
    upcoming_exams = []
    grading_progress = {}
    exams_no_key = []

    if exam_ids:
        subs = supabase.table("submissions").select("student_id,score,final_score,status,exam_id").in_("exam_id", exam_ids).execute().data or []
        unique_students = set(s["student_id"] for s in subs)
        total_students = len(unique_students)
        all_scores = [float(s.get("final_score") or s.get("score") or 0) for s in subs if s.get("final_score") or s.get("score")]

        # Count submissions that need manual grading (submitted/draft but exam has essay questions)
        from datetime import datetime, timezone
        now = datetime.now(timezone.utc)
        for e in exams:
            eid = e["id"]
            exam_subs = [s for s in subs if s.get("exam_id") == eid]
            graded = sum(1 for s in exam_subs if s.get("status") in ("graded", "published"))
            total = len(exam_subs)
            if total > 0:
                grading_progress[eid] = {"graded": graded, "total": total}

            # Check if exam has essay questions that need grading
            qt = e.get("question_types")
            if isinstance(qt, str):
                try: qt = json.loads(qt)
                except: qt = {}
            has_essay = any(v != "mcq" for v in (qt.values() if isinstance(qt, dict) else [])) if qt else False
            if has_essay:
                ungraded = [s for s in exam_subs if s.get("status") in ("submitted", "draft")]
                pending_grading += len(ungraded)

            # Upcoming exams (start in future but within 7 days)
            start_at = e.get("start_at")
            if start_at:
                try:
                    if isinstance(start_at, str):
                        start_dt = datetime.fromisoformat(start_at.replace("Z", "+00:00"))
                    else:
                        start_dt = start_at
                    days_until = (start_dt - now).days
                    if 0 <= days_until <= 7 and e.get("status") == "active":
                        upcoming_exams.append(e)
                except:
                    pass

        # Exams with missing answer keys and MCQ
        exams_no_key = []
        for e in exams:
            ak = e.get("answer_key")
            if not ak or ak == "{}" or ak == {}:
                qt = e.get("question_types")
                if isinstance(qt, str):
                    try: qt = json.loads(qt)
                    except: qt = {}
                if qt:
                    has_mcq = any(v == "mcq" for v in (qt.values() if isinstance(qt, dict) else []))
                    if has_mcq:
                        exams_no_key.append(e)
                    elif not qt:
                        exams_no_key.append(e)

    # Get teacher's school_id (column may not exist in older schema)
    school_id = None
    try:
        profile = supabase.table("profiles").select("*").eq("id", g.user_id).single().execute().data or {}
        school_id = profile.get("school_id")
    except Exception:
        pass

    # Get assignments and available classes/subjects (tables may not exist)
    assignments = []
    classes = []
    subjects = []
    if school_id:
        try:
            assignments = supabase.table("teacher_assignments") \
                .select("*, classes(id, name, grade_level), subjects(id, name, code)") \
                .eq("teacher_id", g.user_id) \
                .eq("school_id", school_id) \
                .execute().data or []
        except Exception:
            pass
        try:
            classes = supabase.table("classes").select("id, name, grade_level") \
                .eq("school_id", school_id) \
                .order("name").execute().data or []
        except Exception:
            pass
        try:
            subjects = supabase.table("subjects").select("id, name, code") \
                .eq("school_id", school_id) \
                .eq("is_active", True) \
                .order("name").execute().data or []
        except Exception:
            pass

    user_name = g.user_name or g.user_email or ""
    # School info for header
    school_info = {}
    if school_id:
        try:
            school_info = supabase.table("schools").select("name, npsn, logo_url").eq("id", school_id).single().execute().data or {}
        except Exception:
            pass
    avg_score = round(sum(all_scores) / len(all_scores), 1) if all_scores else "-"

    return render_template("teacher/dashboard.html", exams=exams, total_students=total_students,
                           avg_score=avg_score, all_scores=all_scores, user_name=user_name,
                           assignments=assignments, classes=classes, subjects=subjects,
                           school_info=school_info, exams_no_key=exams_no_key,
                           pending_grading=pending_grading, upcoming_exams=upcoming_exams,
                           grading_progress=grading_progress)


@teacher_bp.route("/templates")
@teacher_or_admin_required
def exam_templates():
    """Exam template marketplace — browse and copy templates."""
    supabase = get_supabase()
    school_id = g.get("user_school_id")
    templates = []
    if school_id:
        templates = supabase.table("exams") \
            .select("id,title,subject,total_questions,question_types,description,teacher_id,created_at") \
            .eq("is_template", True) \
            .order("created_at", desc=True) \
            .execute().data or []
    return render_template("teacher/templates.html", templates=templates, school_id=school_id)


@teacher_bp.route("/exams/parse-pdf", methods=["POST"])
@teacher_or_admin_required
def exam_parse_pdf():
    """Upload PDF → markdown → AI classification (MCQ vs Essay) → preview."""
    try:
        if "pdf" not in request.files:
            return jsonify({"error": "Tidak ada file PDF"}), 400
        pdf_file = request.files["pdf"]
        raw = pdf_file.read()
        if len(raw) > 50 * 1024 * 1024:
            return jsonify({"error": "PDF terlalu besar. Maksimal 50MB"}), 413

        ai_mode = request.form.get("ai_mode", "false") == "true"
        use_vision = request.form.get("use_vision", "false") == "true"

        # Step 1: PDF → Clean Markdown
        try:
            from app.services.pdf_parser import pdf_to_markdown, classify_with_ai, classify_heuristic, generate_preview_html, generate_answer_key, is_scanned_pdf
        except ImportError:
            return jsonify({"error": "Library tidak tersedia. Jalankan: pip install pymupdf"}), 500

        # Get API key for Gemini Vision if needed
        vision_api_key = ""
        if use_vision:
            try:
                from app.services.ai_service import _get_active_key
                vision_key = _get_active_key(g.user_id)
                if vision_key and vision_key.get("provider") == "gemini":
                    vision_api_key = vision_key.get("api_key", "")
                elif vision_key:
                    # Try to find a gemini key
                    supabase = get_supabase()
                    gemini_keys = supabase.table("teacher_ai_keys").select("*").eq("teacher_id", g.user_id).eq("provider", "gemini").limit(1).execute().data
                    if gemini_keys:
                        vision_api_key = gemini_keys[0].get("api_key", "")
            except:
                pass

        parsed = pdf_to_markdown(raw, use_vision=use_vision, vision_api_key=vision_api_key, lang=request.form.get("lang", "en"))
        if parsed.get("error"):
            return jsonify({"error": parsed["error"]}), 422

        questions = None
        ai_used = False
        answer_key_generated = False
        answer_key = {}
        answer_key_error = ""
        key = None

        # Step 2 & 3: AI processing (only in AI mode)
        if ai_mode:
            try:
                from app.services.ai_service import _get_active_key
                key = _get_active_key(g.user_id)
            except Exception as e:
                current_app.logger.warning("Failed to get AI key: %s", e)

            # Step 2: Classify questions
            try:
                if key and key.get("api_key"):
                    questions = classify_with_ai(
                        parsed["markdown"],
                        api_key=key["api_key"],
                        provider=key.get("provider", "groq"),
                    )
                    if questions and len(questions) > 0:
                        ai_used = True
            except Exception as e:
                current_app.logger.warning("AI classification failed: %s", e)

            if not questions:
                questions = classify_heuristic(parsed["markdown"])

            parsed["questions"] = questions
            parsed["mcq_count"] = sum(1 for q in questions if q.get("type") == "mcq")
            parsed["essay_count"] = sum(1 for q in questions if q.get("type") == "essay")

            # Step 3: Generate answer key
            try:
                if key and key.get("api_key"):
                    current_app.logger.info("Generating answer key with provider: %s", key.get("provider"))
                    ak = generate_answer_key(
                        parsed["markdown"], questions,
                        api_key=key["api_key"],
                        provider=key.get("provider", "groq"),
                        lang=request.form.get("lang", "en"))
                    if ak and len(ak) > 0:
                        if "_error" in ak:
                            answer_key_error = ak["_error"]
                        else:
                            answer_key = ak
                            answer_key_generated = True
                            current_app.logger.info("Answer key generated: %d answers", len(ak))
                    else:
                        current_app.logger.warning("Answer key returned empty")
                else:
                    answer_key_error = "Belum ada API key aktif. Atur di Pengaturan AI."
            except Exception as e:
                answer_key_error = f"Gagal: {str(e)[:100]}"
                current_app.logger.error("Answer key gen error: %s", e, exc_info=True)
        else:
            # Manual mode: no AI, just empty questions
            questions = parsed.get("questions", [])
            parsed["questions"] = questions
            parsed["mcq_count"] = 0
            parsed["essay_count"] = 0

        # Step 4: Save PDF for exam canvas
        pdf_url = ""
        try:
            import uuid
            upload_dir = os.path.join(current_app.root_path, "static", "uploads", "exams")
            os.makedirs(upload_dir, exist_ok=True)
            pdf_filename = f"temp_{str(uuid.uuid4())[:12]}.pdf"
            pdf_path = os.path.join(upload_dir, pdf_filename)
            with open(pdf_path, "wb") as f:
                f.write(raw)
            pdf_url = f"/static/uploads/exams/{pdf_filename}"
        except Exception as e:
            current_app.logger.warning("PDF save skipped: %s", e)

        # Step 4: Generate rubrics for essay questions
        lang = request.form.get("lang", "en")
        for q in questions or []:
            if q.get("type") == "essay":
                try:
                    from app.services.rubric_generator import generate_rubric
                    q["rubric"] = generate_rubric(q.get("text", ""), lang=lang)
                except Exception as e:
                    current_app.logger.warning("Rubric skipped: %s", e)

        preview = generate_preview_html(parsed)
        return jsonify({
            "success": True,
            "ai_classified": ai_used,
            "answer_key_generated": answer_key_generated,
            "answer_key": answer_key,
            "answer_key_error": answer_key_error,
            "markdown": parsed["markdown"][:500000],
            "page_count": parsed["page_count"],
            "mcq_count": parsed["mcq_count"],
            "essay_count": parsed["essay_count"],
            "questions": questions,
            "preview_html": preview,
            "pdf_url": pdf_url,
            "pdf_id": pdf_url.split("/")[-1].replace(".pdf", "").replace("temp_", "") if pdf_url else "",
        })
    except Exception as e:
        current_app.logger.error("parse-pdf error: %s", e, exc_info=True)
        return jsonify({"error": f"Gagal memproses PDF: {str(e)[:200]}"}), 500


@teacher_bp.route("/exams/generate-key", methods=["POST"])
@teacher_or_admin_required
def exam_generate_key():
    """Generate answer key for already-uploaded PDF (on-demand). Also classifies if needed."""
    from app.services.pdf_parser import generate_answer_key, classify_heuristic
    data = request.get_json() or {}
    markdown = data.get("markdown", "")
    questions = data.get("questions", [])
    lang = data.get("lang", "en")
    if not markdown:
        return jsonify({"error": "No markdown provided"}), 400

    # Classify questions if not provided
    if not questions:
        questions = classify_heuristic(markdown)
        if not questions:
            return jsonify({"error": "Tidak dapat mendeteksi soal dari PDF"}), 422

    from app.services.ai_service import _get_active_key
    key = _get_active_key(g.user_id)
    if not key or not key.get("api_key"):
        return jsonify({"error": "Belum ada API key aktif. Atur di Pengaturan AI."}), 400

    try:
        ak = generate_answer_key(markdown, questions, api_key=key["api_key"], provider=key.get("provider", "groq"), lang=lang)
        if ak and "_error" in ak:
            return jsonify({"error": ak["_error"]}), 429
        if ak and len(ak) > 0:
            return jsonify({"success": True, "answer_key": ak, "answer_key_generated": True, "questions": questions})
        current_app.logger.error("generate_answer_key returned empty for user %s. questions=%d, markdown=%d chars, provider=%s",
                                 g.user_id, len(questions), len(markdown), key.get("provider", "groq"))
    except Exception as e:
        current_app.logger.error("generate_answer_key exception: %s", str(e)[:500])
        return jsonify({"error": str(e)[:200]}), 500
    return jsonify({"error": "Gagal generate key"}), 500


@teacher_bp.route("/exams/parse-pdf/markdown", methods=["POST"])
@teacher_or_admin_required
def exam_pdf_markdown():
    """Return PDF as clean markdown text for download."""
    if "pdf" not in request.files:
        return jsonify({"error": "No PDF"}), 400
    raw = request.files["pdf"].read()
    from app.services.pdf_parser import pdf_to_markdown
    result = pdf_to_markdown(raw)
    if result.get("error"):
        return jsonify({"error": result["error"]}), 422
    return result["markdown"], 200, {"Content-Type": "text/markdown; charset=utf-8"}


@teacher_bp.route("/exams/export-scan", methods=["POST"])
@teacher_or_admin_required
def export_scan_results():
    """Export AI scan results as XLSX."""
    data = request.get_json() or {}
    questions = data.get("questions", [])
    fmt = data.get("format", "xlsx")

    if not questions:
        return jsonify({"error": "Tidak ada data"}), 400

    if fmt == "xlsx":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Scan AI"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
        thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        headers = ["No", "Tipe", "Teks Soal", "Rubrik"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal="center"); c.border = thin

        for i, q in enumerate(questions, 1):
            rubric_text = ""
            if q.get("rubric"):
                rubric_text = "; ".join(f"{r.get('kriteria','')} ({r.get('bobot',0)}%)" for r in q["rubric"])
            row = [i, "MCQ" if q.get("type") == "mcq" else "Essay", q.get("text", ""), rubric_text]
            for col, val in enumerate(row, 1):
                c = ws.cell(row=i + 1, column=col, value=val)
                c.border = thin
                c.alignment = Alignment(wrap_text=True, vertical="top")

        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 60
        ws.column_dimensions["D"].width = 50

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        as_attachment=True, download_name="scan_ai.xlsx")

    return jsonify({"error": "Format tidak didukung"}), 400


@teacher_bp.route("/exams/new", methods=["GET", "POST"])
@subscription_write_required
@teacher_or_admin_required
@require_subscription("create_exam")
def exam_form():
    supabase = get_supabase()
    if request.method == "GET":
        supabase = get_supabase()
        sid = g.get("user_school_id")
        subjects = []
        classes = []
        if sid:
            # Get teacher's assigned subjects and classes
            try:
                teacher_assignments = supabase.table("teacher_assignments") \
                    .select("*, subjects(id, name, code), classes(id, name, grade_level)") \
                    .eq("teacher_id", g.user_id) \
                    .execute().data or []
                for a in teacher_assignments:
                    if a.get("subjects"):
                        if a["subjects"] not in subjects:
                            subjects.append(a["subjects"])
                    if a.get("classes"):
                        if a["classes"] not in classes:
                            classes.append(a["classes"])
            except Exception:
                current_app.logger.warning("Failed to fetch teacher assignments, falling back to all")
            # Fallback: if no assignments, show all school subjects/classes
            if not subjects:
                subjects = supabase.table("subjects").select("*").eq("school_id", sid).order("name").execute().data or []
            if not classes:
                classes = supabase.table("classes").select("*").eq("school_id", sid).order("name").execute().data or []
        return render_template("teacher/exam_form.html", exam=None, subjects=subjects, classes=classes)

    title = request.form.get("title")
    subject = request.form.get("subject")
    subject_id = request.form.get("subject_id") or None
    class_ids = request.form.getlist("class_ids")
    is_template = request.form.get("is_template", "false") == "true"
    source_exam_id = request.form.get("source_exam_id") or None
    max_attempts = int(request.form.get("max_attempts", 1))
    publish_mode = request.form.get("publish_mode", "manual")
    total_questions = int(request.form.get("total_questions", 10))
    if total_questions < 1:
        flash("Minimal 1 soal", "error")
        return redirect(request.referrer or "/teacher/exams")
    duration_minutes = int(request.form.get("duration_minutes", 60))
    passing_score = int(request.form.get("passing_score", 70))
    description = request.form.get("description", "")
    action = request.form.get("action", "save_draft")
    start_at_str = request.form.get("start_at", "").strip()
    if action == "publish":
        start_at = None
    elif start_at_str:
        try:
            tz_off = g.get("tz_offset", 7)
            local_dt = datetime.fromisoformat(start_at_str)
            start_at = (local_dt - timedelta(hours=tz_off)).isoformat()
        except Exception:
            start_at = start_at_str
    else:
        start_at = None

    question_types = json.loads(request.form.get("question_types", "{}"))
    answer_key = json.loads(request.form.get("answer_key", "{}"))
    question_weights = json.loads(request.form.get("question_weights", "{}"))
    question_audio = {}
    question_canvas = {}
    anti_cheat_enabled = True
    penalty_per_violation = int(request.form.get("penalty_per_violation", 5))
    max_violations = int(request.form.get("max_violations", 5))
    auto_submit_on_max = request.form.get("auto_submit_on_max") == "true"
    fullscreen_required = request.form.get("fullscreen_required") == "true"
    randomize_questions = request.form.get("randomize_questions", "false") == "true"
    randomize_options = request.form.get("randomize_options", "false") == "true"
    watermark_name = request.form.get("watermark_name") == "true"
    block_copy_paste = request.form.get("block_copy_paste") == "true"
    block_right_click = request.form.get("block_right_click") == "true"
    block_screenshot = request.form.get("block_screenshot", "false") == "true"
    allow_calculator = request.form.get("allow_calculator", "false") == "true"
    for i in range(total_questions):
        qtype = question_types.get(str(i), "mcq")
        if qtype != "mcq":
            question_canvas[str(i)] = True
        audio_url = request.form.get(f"audio_{i}", "").strip()
        youtube_url = request.form.get(f"youtube_{i}", "").strip()
        media = {}
        if audio_url:
            media["audio"] = audio_url
        if youtube_url:
            media["youtube"] = youtube_url
        if media:
            question_audio[str(i)] = media

    data = {
        "teacher_id": g.user_id,
        "school_id": g.get("user_school_id"),
        "title": title,
        "subject": subject,
        "subject_id": subject_id,
        "class_ids": class_ids,
        "start_at": start_at,
        "is_template": is_template,
        "source_exam_id": source_exam_id,
        "max_attempts": max_attempts,
        "publish_mode": publish_mode,
        "question_pages": request.form.get("question_pages", "{}"),
        "total_questions": total_questions,
        "duration_minutes": duration_minutes,
        "passing_score": passing_score,
        "description": description,
        "status": "active" if action in ("save_active", "publish") else "draft",
        "is_published": action == "publish",
        "publish_mode": "auto" if action == "publish" else publish_mode,
        "answer_key": answer_key,
        "question_types": question_types,
        "question_weights": question_weights,
        "question_audio": question_audio,
        "question_canvas": question_canvas,
        "anti_cheat_enabled": anti_cheat_enabled,
        "penalty_per_violation": penalty_per_violation,
        "max_violations": max_violations,
        "auto_submit_on_max": auto_submit_on_max,
        "fullscreen_required": fullscreen_required,
        "randomize_questions": randomize_questions,
        "randomize_options": randomize_options,
        "watermark_name": watermark_name,
        "block_copy_paste": block_copy_paste,
        "block_right_click": block_right_click,
        "block_screenshot": block_screenshot,
        "allow_calculator": allow_calculator,
    }
    try:
        res = supabase.table("exams").insert(data).execute()
    except Exception:
        for key in ["question_weights", "question_texts", "anti_cheat_enabled", "penalty_per_violation", "max_violations", "auto_submit_on_max", "fullscreen_required", "randomize_questions", "randomize_options", "watermark_name", "block_copy_paste", "block_right_click", "block_screenshot", "allow_calculator", "subject_id", "class_ids", "start_at", "is_template", "source_exam_id", "max_attempts", "publish_mode", "question_pages"]:
            data.pop(key, None)
        res = supabase.table("exams").insert(data).execute()
    exam_id = res.data[0]["id"]
    log_activity("create", "exam", exam_id, new_data={"title": title, "subject": subject, "total_questions": total_questions}, user_id=g.user_id)
    # Handle PDF upload inline
    pdf_file = request.files.get("pdf")
    if pdf_file and pdf_file.filename:
        try:
            from app.services.pdf_service import upload_pdf
            result = upload_pdf(pdf_file, exam_id)
            supabase.table("exams").update({
                "pdf_url": result["pdf_path"],
                "pdf_page_urls": result["page_urls"],
            }).eq("id", exam_id).execute()
        except Exception as e:
            current_app.logger.error(f"PDF upload failed: {e}")
    # Handle AJAX-uploaded PDF via pdf_preview_url
    pdf_preview = request.form.get("pdf_preview_url", "")
    if pdf_preview and not pdf_preview.startswith("http") and not (pdf_file and pdf_file.filename):
        try:
            local_path = os.path.join(current_app.root_path, "static", "uploads", "exams", os.path.basename(pdf_preview))
            pdf_bytes = None
            if os.path.exists(local_path):
                with open(local_path, "rb") as f:
                    pdf_bytes = f.read()
            else:
                # Try alternative: maybe root_path is project root, not app/
                alt_path = os.path.join(os.path.dirname(current_app.root_path), "static", "uploads", "exams", os.path.basename(pdf_preview))
                if os.path.exists(alt_path):
                    with open(alt_path, "rb") as f:
                        pdf_bytes = f.read()
            if pdf_bytes:
                from app.services.pdf_service import upload_pdf
                class _MF:
                    def __init__(self, d, n): self._d = d; self.filename = n
                    def read(self): return self._d
                result = upload_pdf(_MF(pdf_bytes, "exam.pdf"), exam_id)
                supabase.table("exams").update({
                    "pdf_url": result["pdf_path"],
                    "pdf_page_urls": result["page_urls"],
                }).eq("id", exam_id).execute()
                current_app.logger.info("PDF processed for exam %s: %d pages", exam_id, result["total_pages"])
            else:
                current_app.logger.warning("PDF temp file not found for exam %s: %s", exam_id, pdf_preview)
        except Exception as e:
            current_app.logger.warning(f"PDF preview processing failed: {e}")
            flash("PDF gagal diproses untuk canvas siswa. Upload ulang PDF setelah menyimpan.", "warning")
    # If action is publish, also publish scores automatically
    if action == "publish":
        try:
            _recalculate_scores(exam_id)
            supabase.table("submissions").update({"is_published": True, "status": "published"}).eq("exam_id", exam_id).execute()
        except Exception:
            pass
        flash("✅ Ujian berhasil dipublikasikan! Siswa sekarang bisa mengerjakan.", "success")
    else:
        flash("✅ Ujian berhasil disimpan.", "success")
    return redirect("/teacher/exams" if action == "publish" else f"/teacher/exams/{exam_id}")


@teacher_bp.route("/exams/<exam_id>", methods=["GET", "POST", "DELETE"])
@subscription_write_required
@teacher_or_admin_required
@require_school_access("exams", "exam_id")
def exam_detail(exam_id):
    supabase = get_supabase()
    if request.method == "DELETE":
        supabase.table("exams").delete().eq("id", exam_id).execute()
        return jsonify({"success": True})
    if request.method == "GET":
        try:
            res = supabase.table("exams").select("*").eq("id", exam_id).single().execute()
        except Exception:
            res = supabase.table("exams").select("id,title,subject,total_questions,duration_minutes,passing_score,description,status,answer_key,question_types,question_audio,question_canvas,teacher_id,created_at").eq("id", exam_id).single().execute()
        exam_data = res.data
        exam_data.setdefault("question_weights", {})
        sid = g.get("user_school_id")
        subjects = []
        classes = []
        if sid:
            subjects = supabase.table("subjects").select("*").eq("school_id", sid).order("name").execute().data or []
            classes = supabase.table("classes").select("*").eq("school_id", sid).order("name").execute().data or []
        return render_template("teacher/exam_form.html", exam=exam_data, subjects=subjects, classes=classes)

    title = request.form.get("title")
    subject = request.form.get("subject")
    subject_id = request.form.get("subject_id") or None
    class_ids = request.form.getlist("class_ids")
    is_template = request.form.get("is_template", "false") == "true"
    source_exam_id = request.form.get("source_exam_id") or None
    max_attempts = int(request.form.get("max_attempts", 1))
    publish_mode = request.form.get("publish_mode", "manual")
    total_questions = int(request.form.get("total_questions", 10))
    if total_questions < 1:
        flash("Minimal 1 soal", "error")
        return redirect(request.referrer or "/teacher/exams")
    duration_minutes = int(request.form.get("duration_minutes", 60))
    passing_score = int(request.form.get("passing_score", 70))
    description = request.form.get("description", "")
    action = request.form.get("action", "save_draft")
    start_at_str = request.form.get("start_at", "").strip()
    if action == "publish":
        start_at = None
    elif start_at_str:
        try:
            tz_off = g.get("tz_offset", 7)
            local_dt = datetime.fromisoformat(start_at_str)
            start_at = (local_dt - timedelta(hours=tz_off)).isoformat()
        except Exception:
            start_at = start_at_str
    else:
        start_at = None

    question_types = json.loads(request.form.get("question_types", "{}"))
    answer_key = json.loads(request.form.get("answer_key", "{}"))
    question_weights = json.loads(request.form.get("question_weights", "{}"))
    question_audio = {}
    question_canvas = {}
    anti_cheat_enabled = True
    penalty_per_violation = int(request.form.get("penalty_per_violation", 5))
    max_violations = int(request.form.get("max_violations", 5))
    auto_submit_on_max = request.form.get("auto_submit_on_max") == "true"
    fullscreen_required = request.form.get("fullscreen_required") == "true"
    randomize_questions = request.form.get("randomize_questions", "false") == "true"
    randomize_options = request.form.get("randomize_options", "false") == "true"
    watermark_name = request.form.get("watermark_name") == "true"
    block_copy_paste = request.form.get("block_copy_paste") == "true"
    block_right_click = request.form.get("block_right_click") == "true"
    block_screenshot = request.form.get("block_screenshot", "false") == "true"
    allow_calculator = request.form.get("allow_calculator", "false") == "true"
    for i in range(total_questions):
        qtype = question_types.get(str(i), "mcq")
        if qtype != "mcq":
            question_canvas[str(i)] = True
        audio_url = request.form.get(f"audio_{i}", "").strip()
        youtube_url = request.form.get(f"youtube_{i}", "").strip()
        media = {}
        if audio_url:
            media["audio"] = audio_url
        if youtube_url:
            media["youtube"] = youtube_url
        if media:
            question_audio[str(i)] = media

    data = {
        "teacher_id": g.user_id,
        "school_id": g.get("user_school_id"),
        "title": title,
        "subject": subject,
        "subject_id": subject_id,
        "class_ids": class_ids,
        "start_at": start_at,
        "is_template": is_template,
        "source_exam_id": source_exam_id,
        "max_attempts": max_attempts,
        "publish_mode": publish_mode,
        "question_pages": request.form.get("question_pages", "{}"),
        "total_questions": total_questions,
        "duration_minutes": duration_minutes,
        "passing_score": passing_score,
        "description": description,
        "status": "active" if action in ("save_active", "publish") else "draft",
        "is_published": action == "publish",
        "publish_mode": "auto" if action == "publish" else publish_mode,
        "answer_key": answer_key,
        "question_types": question_types,
        "question_weights": question_weights,
        "question_audio": question_audio,
        "question_canvas": question_canvas,
        "anti_cheat_enabled": anti_cheat_enabled,
        "penalty_per_violation": penalty_per_violation,
        "max_violations": max_violations,
        "auto_submit_on_max": auto_submit_on_max,
        "fullscreen_required": fullscreen_required,
        "randomize_questions": randomize_questions,
        "randomize_options": randomize_options,
        "watermark_name": watermark_name,
        "block_copy_paste": block_copy_paste,
        "block_right_click": block_right_click,
        "block_screenshot": block_screenshot,
        "allow_calculator": allow_calculator,
    }
    try:
        supabase.table("exams").update(data).eq("id", exam_id).execute()
    except Exception:
        for key in ["question_weights", "question_texts", "anti_cheat_enabled", "penalty_per_violation", "max_violations", "auto_submit_on_max", "fullscreen_required", "randomize_questions", "randomize_options", "watermark_name", "block_copy_paste", "block_right_click", "block_screenshot", "allow_calculator", "subject_id", "class_ids", "start_at", "is_template", "source_exam_id", "max_attempts", "publish_mode", "question_pages"]:
            data.pop(key, None)
        supabase.table("exams").update(data).eq("id", exam_id).execute()

    # Process PDF: upload to Supabase, generate page images for student canvas
    pdf_preview = request.form.get("pdf_preview_url", "")
    if pdf_preview and not pdf_preview.startswith("http"):
        try:
            local_path = os.path.join(current_app.root_path, "static", "uploads", "exams", os.path.basename(pdf_preview))
            pdf_bytes = None
            if os.path.exists(local_path):
                with open(local_path, "rb") as f:
                    pdf_bytes = f.read()
            else:
                alt_path = os.path.join(os.path.dirname(current_app.root_path), "static", "uploads", "exams", os.path.basename(pdf_preview))
                if os.path.exists(alt_path):
                    with open(alt_path, "rb") as f:
                        pdf_bytes = f.read()
            if pdf_bytes:
                from app.services.pdf_service import upload_pdf
                class MockFile:
                    def __init__(self, data, name):
                        self._data = data
                        self.filename = name
                    def read(self): return self._data
                result = upload_pdf(MockFile(pdf_bytes, "exam.pdf"), exam_id)
                supabase.table("exams").update({
                    "pdf_url": result["pdf_path"],
                    "pdf_page_urls": result["page_urls"],
                }).eq("id", exam_id).execute()
                current_app.logger.info("PDF processed for exam %s: %d pages", exam_id, result["total_pages"])
            else:
                current_app.logger.warning("PDF temp file not found for exam %s: %s", exam_id, pdf_preview)
        except Exception as e:
            current_app.logger.warning("PDF processing failed for exam %s: %s", exam_id, e)

    _recalculate_scores(exam_id)
    log_activity("update", "exam", exam_id, new_data={"title": title, "status": data.get("status")}, user_id=g.user_id)
    return redirect("/teacher/exams" if action in ("publish", "save_active") else f"/teacher/exams/{exam_id}")


@teacher_bp.route("/exams/<exam_id>/preprocess-essays", methods=["POST"])
@teacher_or_admin_required
@require_school_access("exams", "exam_id")
def preprocess_exam_essays(exam_id):
    """Generate embeddings + rubric for all essay questions in an exam."""
    supabase = get_supabase()
    exam = supabase.table("exams").select("question_types,question_texts,question_rubrics,total_questions").eq("id", exam_id).single().execute().data
    if not exam:
        return jsonify({"error": "Exam not found"}), 404

    qtypes = exam.get("question_types") or {}
    total_q = exam.get("total_questions", 0)
    from app.services.ai_embedding import preprocess_exam_questions
    from app.services.rubric_generator import generate_rubric

    # Build questions list
    questions = []
    question_texts = exam.get("question_texts") or {}
    for i in range(total_q):
        qi = str(i)
        if qtypes.get(qi, "mcq") != "mcq":
            text = question_texts.get(qi, "")
            rubric = generate_rubric(text)
            questions.append({"number": i + 1, "type": "essay", "text": text, "rubric": rubric})

    result = preprocess_exam_questions(exam_id, questions, supabase)
    return jsonify({"success": True, **result})


@teacher_bp.route("/preview/<exam_id>")
@teacher_or_admin_required
@require_school_access("exams", "exam_id")
def preview_exam(exam_id):
    supabase = get_supabase()
    res = supabase.table("exams").select("*").eq("id", exam_id).maybe_single().execute()
    if not res.data:
        return redirect("/teacher/exams")
    return render_template("teacher/preview_exam.html", exam=res.data)


@teacher_bp.route("/exams/<exam_id>/publish-exam", methods=["POST"])
@subscription_write_required
@teacher_or_admin_required
@require_school_access("exams", "exam_id")
def publish_exam(exam_id):
    supabase = get_supabase()
    supabase.table("exams").update({
        "is_published": True,
        "status": "active",
    }).eq("id", exam_id).execute()
    log_activity("publish", "exam", exam_id, user_id=g.user_id)
    return redirect(f"/teacher/preview/{exam_id}")


@teacher_bp.route("/exams/<exam_id>/upload-pdf", methods=["GET", "POST"])
@subscription_write_required
@teacher_or_admin_required
@require_school_access("exams", "exam_id")
def upload_exam_pdf(exam_id):
    supabase = get_supabase()
    if request.method == "GET":
        exam = supabase.table("exams").select("*").eq("id", exam_id).single().execute().data
        return render_template("teacher/upload_pdf.html", exam=exam)

    if "pdf" not in request.files:
        exam = supabase.table("exams").select("*").eq("id", exam_id).single().execute().data
        return render_template("teacher/upload_pdf.html", exam=exam, error="Pilih file PDF")

    file = request.files["pdf"]
    try:
        result = upload_pdf(file, exam_id)
    except ValueError as e:
        exam = supabase.table("exams").select("*").eq("id", exam_id).single().execute().data
        return render_template("teacher/upload_pdf.html", exam=exam, error=str(e))
    supabase.table("exams").update({
        "pdf_url": result["pdf_path"],
        "pdf_page_urls": result["page_urls"],
    }).eq("id", exam_id).execute()
    log_activity("upload", "exam", exam_id, new_data={"pages": len(result.get("page_urls", []))}, user_id=g.user_id)
    return redirect(f"/teacher/preview/{exam_id}")


@teacher_bp.route("/exams")
@teacher_or_admin_required
def my_exams():
    supabase = get_supabase()
    res = supabase.table("exams").select("*").eq("teacher_id", g.user_id).order("created_at", desc=True).execute()
    return render_template("teacher/exams.html", exams=res.data or [])


@teacher_bp.route("/exams/<exam_id>/toggle-status", methods=["POST"])
@subscription_write_required
@teacher_or_admin_required
@require_school_access("exams", "exam_id")
def toggle_exam_status(exam_id):
    supabase = get_supabase()
    exam = supabase.table("exams").select("status").eq("id", exam_id).single().execute().data
    new_status = "draft" if exam["status"] == "active" else "active"
    supabase.table("exams").update({"status": new_status}).eq("id", exam_id).execute()
    if request.headers.get("Accept", "") == "application/json" or request.is_json:
        return jsonify({"success": True, "status": new_status})
    return redirect(request.referrer or "/teacher/exams")


@teacher_bp.route("/exams/<exam_id>/toggle-visibility", methods=["POST"])
@subscription_write_required
@teacher_or_admin_required
@require_school_access("exams", "exam_id")
def toggle_exam_visibility(exam_id):
    supabase = get_supabase()
    exam = supabase.table("exams").select("is_published").eq("id", exam_id).single().execute().data
    new_val = not exam["is_published"]
    supabase.table("exams").update({"is_published": new_val}).eq("id", exam_id).execute()
    if request.headers.get("Accept", "") == "application/json" or request.is_json:
        return jsonify({"success": True, "is_published": new_val})
    return redirect(request.referrer or "/teacher/exams")


@teacher_bp.route("/exams/<exam_id>/delete", methods=["POST"])
@subscription_write_required
@teacher_or_admin_required
@require_school_access("exams", "exam_id")
def delete_exam(exam_id):
    supabase = get_supabase()
    supabase.table("violation_logs").delete().eq("exam_id", exam_id).execute()
    supabase.table("exam_access_codes").delete().eq("exam_id", exam_id).execute()
    supabase.table("analytics_cache").delete().eq("exam_id", exam_id).execute()
    supabase.table("submissions").delete().eq("exam_id", exam_id).execute()
    supabase.table("exams").delete().eq("id", exam_id).execute()
    log_activity("delete", "exam", exam_id, user_id=g.user_id)
    if request.headers.get("Accept", "") == "application/json" or request.is_json:
        return jsonify({"success": True})
    return redirect("/teacher/exams")


@teacher_bp.route("/exams/<exam_id>/duplicate", methods=["POST"])
@subscription_write_required
@teacher_or_admin_required
@require_school_access("exams", "exam_id")
def duplicate_exam(exam_id):
    supabase = get_supabase()
    try:
        res = supabase.table("exams").select("*").eq("id", exam_id).single().execute()
        exam = res.data
        if not exam:
            flash("Ujian tidak ditemukan", "error")
            return redirect("/teacher/exams")
        import copy, uuid
        new_data = {k: v for k, v in exam.items() if k not in ("id", "created_at", "updated_at")}
        new_data["title"] = exam["title"] + " (salinan)"
        new_data["status"] = "draft"
        new_data["is_published"] = False
        new_data["is_template"] = False
        new_data["source_exam_id"] = exam_id
        new_exam = supabase.table("exams").insert(new_data).execute()
        new_id = new_exam.data[0]["id"]
        log_activity("duplicate", "exam", new_id, new_data={"source": exam_id, "title": new_data["title"]}, user_id=g.user_id)
        flash("Ujian berhasil digandakan. Silakan edit sesuai kebutuhan.", "success")
        return redirect(f"/teacher/exams/{new_id}")
    except Exception as e:
        flash(f"Gagal menggandakan: {str(e)[:60]}", "error")
        return redirect("/teacher/exams")


@teacher_bp.route("/exams/<exam_id>/answer-keys", methods=["GET", "POST"])
@teacher_or_admin_required
@require_school_access("exams", "exam_id")
def answer_keys(exam_id):
    supabase = get_supabase()
    exam = supabase.table("exams").select("*").eq("id", exam_id).single().execute().data
    if not exam:
        flash("Ujian tidak ditemukan", "error")
        return redirect("/teacher/exams")

    # Parse JSON fields
    for fld in ("answer_key", "question_types"):
        v = exam.get(fld)
        if isinstance(v, str):
            try: exam[fld] = json.loads(v)
            except: exam[fld] = {}

    if request.method == "POST":
        answer_key = request.form.get("answer_key", "{}")
        try:
            answer_key = json.loads(answer_key)
        except json.JSONDecodeError:
            answer_key = {}
        supabase.table("exams").update({"answer_key": json.dumps(answer_key)}).eq("id", exam_id).execute()
        # Recalculate scores
        try:
            from app.routes.teacher import _recalculate_scores
            _recalculate_scores(exam_id)
        except Exception:
            pass
        flash("Kunci jawaban berhasil disimpan & nilai diperbarui!", "success")
        return redirect(f"/teacher/exams/{exam_id}/answer-keys")

    # GET: build question list from question_types
    qtypes = exam.get("question_types", {})
    akey = exam.get("answer_key", {})
    questions = []
    for i in sorted(qtypes.keys(), key=int):
        idx = str(i)
        qtype = qtypes[idx]
        k = akey.get(idx)
        questions.append({
            "index": int(idx),
            "type": qtype,
            "key": k,
            "is_bonus": k == "bonus",
            "is_multi": isinstance(k, list),
        })

    return render_template("teacher/answer_keys.html", exam=exam, questions=questions)


@teacher_bp.route("/scan")
@teacher_or_admin_required
def scan_page():
    supabase = get_supabase()
    res = supabase.table("exams").select("*").eq("teacher_id", g.user_id).execute()
    students = supabase.table("profiles").select("id,full_name,phone").eq("role", "murid").execute()
    return render_template("teacher/scan.html", exams=res.data, students=students.data)


@teacher_bp.route("/retractions")
@teacher_or_admin_required
def retraction_requests():
    supabase = get_supabase()
    exam_ids = [e["id"] for e in supabase.table("exams").select("id").eq("teacher_id", g.user_id).execute().data or []]
    requests = []
    if exam_ids:
        subs = supabase.table("submissions").select("id,student_id,exam_id,answers,submitted_at,exams(title),profiles(full_name)").in_("exam_id", exam_ids).execute().data or []
        for s in subs:
            answers = s.get("answers")
            if isinstance(answers, str):
                try:
                    answers = json.loads(answers)
                except (json.JSONDecodeError, TypeError):
                    answers = {}
            if isinstance(answers, dict) and answers.get("_retract_request", {}).get("status") == "pending":
                s["exam_title"] = (s.get("exams") or {}).get("title", "-")
                s["student_name"] = (s.get("profiles") or {}).get("full_name", "-")
                s["requested_at"] = answers["_retract_request"].get("requested_at", "")
                requests.append(s)
    return render_template("teacher/retractions.html", requests=requests)


@teacher_bp.route("/retractions/<submission_id>/approve", methods=["POST"])
@teacher_or_admin_required
@require_school_access("submissions", "submission_id", ("exam_id", "exams"))
def approve_retraction(submission_id):
    supabase = get_supabase()
    sub = supabase.table("submissions").select("answers").eq("id", submission_id).single().execute().data
    if not sub:
        flash("Submission tidak ditemukan", "error")
        return redirect(url_for("teacher.retraction_requests"))
    answers = sub.get("answers")
    if isinstance(answers, str):
        try:
            answers = json.loads(answers)
        except (json.JSONDecodeError, TypeError):
            answers = {}
    if not isinstance(answers, dict) or "_retract_request" not in answers or not isinstance(answers["_retract_request"], dict):
        flash("Tidak ada permintaan retraction", "error")
        return redirect(url_for("teacher.retraction_requests"))
    answers["_retract_request"]["status"] = "approved"
    try:
        supabase.table("submissions").update({"answers": json.dumps(answers), "status": "retracted"}).eq("id", submission_id).execute()
    except Exception:
        supabase.table("submissions").update({"answers": json.dumps(answers)}).eq("id", submission_id).execute()
    log_activity("retract_approve", "submission", submission_id, user_id=g.user_id)
    flash("Retraction berhasil disetujui", "success")
    return redirect(url_for("teacher.retraction_requests"))


@teacher_bp.route("/retractions/<submission_id>/reject", methods=["POST"])
@teacher_or_admin_required
@require_school_access("submissions", "submission_id", ("exam_id", "exams"))
def reject_retraction(submission_id):
    supabase = get_supabase()
    sub = supabase.table("submissions").select("answers").eq("id", submission_id).single().execute().data
    if not sub:
        flash("Submission tidak ditemukan", "error")
        return redirect(url_for("teacher.retraction_requests"))
    answers = sub.get("answers")
    if isinstance(answers, str):
        try:
            answers = json.loads(answers)
        except (json.JSONDecodeError, TypeError):
            answers = {}
    if not isinstance(answers, dict) or "_retract_request" not in answers or not isinstance(answers["_retract_request"], dict):
        flash("Tidak ada permintaan retraction", "error")
        return redirect(url_for("teacher.retraction_requests"))
    answers["_retract_request"]["status"] = "rejected"
    supabase.table("submissions").update({"answers": json.dumps(answers)}).eq("id", submission_id).execute()
    log_activity("retract_reject", "submission", submission_id, user_id=g.user_id)
    flash("Retraction ditolak", "success")
    return redirect(url_for("teacher.retraction_requests"))


@teacher_bp.route("/results")
@teacher_or_admin_required
def results():
    exam_id = request.args.get("exam_id")
    supabase = get_supabase()
    user_role = g.get("user_role")
    school_id = g.get("user_school_id")

    if not exam_id:
        query = supabase.table("exams").select("id,title").eq("teacher_id", g.user_id)
        if user_role == "admin_sekolah" and school_id:
            query = supabase.table("exams").select("id,title").eq("school_id", school_id)
        exams = query.execute().data or []
        return render_template("teacher/results.html", submissions=[], stats={}, exam_id="", exams=exams, exam={}, scan_subs=[], online_subs=[])

    subs = supabase.table("submissions").select("*, profiles(full_name)").eq("exam_id", exam_id).execute().data or []
    query = supabase.table("exams").select("id,title,subject").eq("teacher_id", g.user_id)
    if user_role == "admin_sekolah" and school_id:
        query = supabase.table("exams").select("id,title,subject").eq("school_id", school_id)
    exams = query.execute().data or []
    exam = next((e for e in exams if e["id"] == exam_id), {})

    for s in subs:
        if s.get("profiles"):
            s["student_name"] = s.pop("profiles").get("full_name", "")

    # Deduplicate: keep latest per (student_id, source)
    seen = {}
    for s in sorted(subs, key=lambda x: x.get("submitted_at", "") or "", reverse=True):
        sid = s.get("student_id", "")
        ans = s.get("answers") or {}
        if isinstance(ans, str):
            try:
                ans = json.loads(ans)
            except Exception:
                ans = {}
        source = "scan" if isinstance(ans, dict) and ans.get("_nisn") else "online"
        key = (sid, source)
        if key not in seen:
            s["_source"] = source
            seen[key] = s
    subs = list(seen.values())
    scan_subs = [s for s in subs if s.get("_source") == "scan"]
    online_subs = [s for s in subs if s.get("_source") != "scan"]

    if subs:
        scores = [float(s.get("final_score") or s.get("score") or 0) for s in subs]
        stats = {
            "avg": round(sum(scores) / len(scores), 1),
            "max": max(scores),
            "min": min(scores),
            "count": len(scores),
        "question_texts": request.form.get("question_texts", "{}"),
        "pdf_url": request.form.get("pdf_preview_url", ""),
    }
    else:
        stats = {"avg": 0, "max": 0, "min": 0, "count": 0}

    return render_template("teacher/results.html", submissions=subs, stats=stats, exam_id=exam_id, exams=exams, exam=exam, scan_subs=scan_subs, online_subs=online_subs)


@teacher_bp.route("/grade-question/<exam_id>/<int:question_index>")
@teacher_or_admin_required
def grade_question(exam_id, question_index):
    """Grade a single question across all students."""
    supabase = get_supabase()
    exam = supabase.table("exams").select("title,total_questions,question_types,answer_key").eq("id", exam_id).single().execute().data or {}
    for f in ("question_types", "answer_key"):
        v = exam.get(f)
        if isinstance(v, str):
            try: exam[f] = json.loads(v)
            except: exam[f] = {}
    return render_template("teacher/grade_question.html", exam=exam, exam_id=exam_id)


@teacher_bp.route("/api/grade-question/<exam_id>/<int:question_index>")
@teacher_or_admin_required
def grade_question_api(exam_id, question_index):
    """API: return all students' answers for a specific question."""
    supabase = get_supabase()
    subs = supabase.table("submissions").select("id,student_id,answers,score,final_score,status,submitted_at,profiles(full_name)").eq("exam_id", exam_id).execute().data or []

    students = []
    for s in subs:
        answers = s.get("answers") or {}
        if isinstance(answers, str):
            try: answers = json.loads(answers)
            except: answers = {}

        student_name = (s.get("profiles") or {}).get("full_name", s.get("student_id", "")[:12])
        qi = str(question_index)
        ans_data = answers.get(qi)

        # Extract answer text/option
        answer_val = ""
        essay_text = ""
        has_canvas = False
        feedback = {"score": None, "feedback": None}

        if isinstance(ans_data, dict):
            answer_val = ans_data.get("answer", "")
            if ans_data.get("pages"):
                has_canvas = True
            if ans_data.get("text"):
                essay_text = ans_data["text"]
            feedback["score"] = ans_data.get("ai_score") or ans_data.get("score")
            feedback["feedback"] = ans_data.get("feedback") or ans_data.get("ai_feedback")
        elif isinstance(ans_data, str):
            answer_val = ans_data

        students.append({
            "submission_id": s["id"],
            "name": student_name,
            "answer": answer_val,
            "essayText": essay_text or answer_val if essay_text else "",
            "hasCanvas": has_canvas,
            "status": s.get("status", ""),
            "score": s.get("score"),
            "final_score": s.get("final_score"),
            "submitted_at": str(s.get("submitted_at", ""))[:19],
            "feedback": feedback,
        })

    return jsonify({"students": students})


@teacher_bp.route("/api/grade-question/<exam_id>/<int:question_index>/save", methods=["POST"])
@teacher_or_admin_required
def grade_question_save(exam_id, question_index):
    """Save a grade update for a specific question on a submission."""
    data = request.get_json()
    submission_id = data.get("submission_id")
    if not submission_id:
        return jsonify({"error": "No submission_id"}), 400

    supabase = get_supabase()
    sub = supabase.table("submissions").select("answers").eq("id", submission_id).single().execute().data
    if not sub:
        return jsonify({"error": "Not found"}), 404

    answers = sub.get("answers") or {}
    if isinstance(answers, str):
        try: answers = json.loads(answers)
        except: answers = {}

    qi = str(question_index)
    current = answers.get(qi, {})
    if not isinstance(current, dict):
        current = {"answer": current}

    if "answer" in data:
        current["answer"] = data["answer"]
    if "essay_score" in data:
        current["ai_score"] = float(data["essay_score"]) if data.get("essay_score") else None
    if "essay_feedback" in data:
        current["feedback"] = data["essay_feedback"]

    answers[qi] = current
    supabase.table("submissions").update({"answers": answers}).eq("id", submission_id).execute()
    return jsonify({"success": True})


@teacher_bp.route("/grade/<submission_id>")
@teacher_or_admin_required
@require_school_access("submissions", "submission_id", ("exam_id", "exams"))
def grade_detail(submission_id):
    supabase = get_supabase()
    sub = supabase.table("submissions").select("*").eq("id", submission_id).maybe_single().execute()
    if not sub.data:
        return redirect("/teacher/grading")
    sub = sub.data
    exam = supabase.table("exams").select("*").eq("id", sub["exam_id"]).single().execute().data
    exam.setdefault("question_weights", {})
    if not exam.get("question_weights") and exam.get("total_questions", 0) > 0:
        qtypes = exam.get("question_types", {})
        tq = exam["total_questions"]
        mcq_n = sum(1 for i in range(tq) if qtypes.get(str(i), "mcq") == "mcq")
        essay_n = tq - mcq_n
        mp, ep = (70, 30)
        if mcq_n == 0: mp, ep = 0, 100
        elif essay_n == 0: mp, ep = 100, 0
        if mcq_n:
            e = round(mp / mcq_n, 2)
            for i in range(tq):
                if qtypes.get(str(i), "mcq") == "mcq":
                    exam["question_weights"][str(i)] = e
        if essay_n:
            e = round(ep / essay_n, 2)
            for i in range(tq):
                if qtypes.get(str(i), "mcq") != "mcq":
                    exam["question_weights"][str(i)] = e
    student = supabase.table("profiles").select("id,full_name,phone").eq("id", sub["student_id"]).single().execute().data or {}
    # Parse JSON string fields
    for field in ("teacher_feedback", "answers"):
        val = sub.get(field)
        if isinstance(val, str):
            try:
                sub[field] = json.loads(val)
            except (json.JSONDecodeError, TypeError):
                sub[field] = {} if field != "answers" else {}
        if not isinstance(sub.get(field), dict):
            sub[field] = {} if field != "answers" else {}
    # Parse exam JSON fields
    for _field in ("question_types", "answer_key", "question_weights", "question_pages", "pdf_page_urls"):
        _val = exam.get(_field)
        if isinstance(_val, str):
            try:
                exam[_field] = json.loads(_val)
            except (json.JSONDecodeError, TypeError):
                exam[_field] = {}
    return render_template("teacher/grade_detail.html", submission=sub, exam=exam, exam_id=sub["exam_id"], student=student)


@teacher_bp.route("/grade/<submission_id>/override", methods=["POST"])
@subscription_write_required
@teacher_or_admin_required
@require_school_access("submissions", "submission_id", ("exam_id", "exams"))
def override_score(submission_id):
    if request.is_json:
        data = request.get_json()
        new_score = data.get("final_score")
        feedback = data.get("teacher_feedback", {})
    else:
        new_score = request.form.get("final_score")
        feedback_raw = request.form.get("teacher_feedback", "{}")
        try:
            feedback = json.loads(feedback_raw)
        except json.JSONDecodeError:
            feedback = {}
    supabase = get_supabase()
    final_score_val = float(new_score) if new_score is not None and new_score != '' else None
    supabase.table("submissions").update({
        "final_score": final_score_val,
        "status": "graded",
        "teacher_feedback": feedback,
    }).eq("id", submission_id).execute()
    # Recalculate after grading
    try:
        sub_updated = supabase.table("submissions").select("exam_id").eq("id", submission_id).single().execute().data
        if sub_updated:
            _recalculate_scores(sub_updated["exam_id"])
    except Exception:
        pass
    if request.is_json:
        return jsonify({"success": True, "final_score": final_score_val})
    return redirect(request.referrer or "/teacher/results")


@teacher_bp.route("/publish/<exam_id>", methods=["GET", "POST"])
@subscription_write_required
@teacher_or_admin_required
def publish_scores(exam_id):
    supabase = get_supabase()
    if request.method == "GET":
        exam = supabase.table("exams").select("id,title,passing_score").eq("id", exam_id).single().execute().data
        subs = supabase.table("submissions").select("id,student_id,final_score,status,profiles(full_name)").eq("exam_id", exam_id).order("profiles.full_name").execute().data or []
        return render_template("teacher/publish_preview.html", exam=exam, submissions=subs)
    _recalculate_scores(exam_id)
    supabase.table("submissions") \
        .update({"is_published": True, "status": "published"}) \
        .eq("exam_id", exam_id) \
        .execute()
    return redirect("/teacher/results?exam_id=" + exam_id)


@teacher_bp.route("/publish/<exam_id>/unpublish", methods=["POST"])
@subscription_write_required
@teacher_or_admin_required
def unpublish_scores(exam_id):
    supabase = get_supabase()
    supabase.table("submissions") \
        .update({"is_published": False, "status": "graded"}) \
        .eq("exam_id", exam_id) \
        .execute()
    return redirect("/teacher/results?exam_id=" + exam_id)


@teacher_bp.route("/publish/submission/<submission_id>", methods=["POST"])
@subscription_write_required
@teacher_or_admin_required
def publish_single(submission_id):
    supabase = get_supabase()
    sub = supabase.table("submissions").select("exam_id").eq("id", submission_id).single().execute().data
    supabase.table("submissions") \
        .update({"is_published": True, "status": "published"}) \
        .eq("id", submission_id) \
        .execute()
    return redirect("/teacher/results?exam_id=" + sub["exam_id"])


# --- Export ---
@teacher_bp.route("/export/xlsx")
@teacher_or_admin_required
def export_xlsx():
    exam_id = request.args.get("exam_id")
    supabase = get_supabase()
    exam = supabase.table("exams").select("*").eq("id", exam_id).single().execute().data or {}
    subs = supabase.table("submissions").select("*, profiles(full_name)").eq("exam_id", exam_id).execute().data or []
    for s in subs:
        s["student_name"] = (s.pop("profiles", None) or {}).get("full_name", s.get("student_id", "")[:12])
    buf = export_to_xlsx(subs, exam)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"nilai_{exam_id[:8]}.xlsx")


@teacher_bp.route("/export/pdf")
@teacher_or_admin_required
def export_pdf():
    exam_id = request.args.get("exam_id")
    supabase = get_supabase()
    exam = supabase.table("exams").select("*").eq("id", exam_id).single().execute().data or {}
    subs = supabase.table("submissions").select("*, profiles(full_name)").eq("exam_id", exam_id).execute().data or []
    for s in subs:
        s["student_name"] = (s.pop("profiles", None) or {}).get("full_name", s.get("student_id", "")[:12])
    buf = export_to_pdf(subs, exam.get("title", "Hasil"), exam)
    return send_file(buf, mimetype="application/pdf", as_attachment=True,
                     download_name=f"nilai_{exam_id[:8]}.pdf")


@teacher_bp.route("/export/bubble-sheet/<exam_id>")
@teacher_or_admin_required
def bubble_sheet(exam_id):
    supabase = get_supabase()
    exam = supabase.table("exams").select("*").eq("id", exam_id).single().execute().data
    qtypes = exam.get("question_types") or {}
    mcq_count = sum(1 for i in range(exam["total_questions"]) if qtypes.get(str(i), "mcq") == "mcq")
    if mcq_count == 0:
        mcq_count = exam["total_questions"]

    buf = generate_answer_sheet(
        total_questions=mcq_count,
        subject=exam.get("subject", ""),
        school_name="",
    )
    return send_file(buf, mimetype="application/pdf", as_attachment=True,
                     download_name=f"LJK_{exam_id[:8]}.pdf")


@teacher_bp.route("/grading")
@teacher_or_admin_required
def grading_center():
    supabase = get_supabase()
    user_role = g.get("user_role")
    school_id = g.get("user_school_id")
    exam_ids = [e["id"] for e in supabase.table("exams").select("id").eq("teacher_id", g.user_id).execute().data or []]
    # For admin_sekolah, also get exams from their school
    if not exam_ids and user_role == "admin_sekolah" and school_id:
        exam_ids = [e["id"] for e in supabase.table("exams").select("id").eq("school_id", school_id).execute().data or []]
    pending_subs = []
    graded_subs = []
    if exam_ids:
        subs = supabase.table("submissions").select("id,student_id,exam_id,score,final_score,status,submitted_at,exams(title),profiles(full_name)").in_("exam_id", exam_ids).order("submitted_at", desc=True).execute().data or []
        for s in subs:
            s["student_name"] = (s.get("profiles") or {}).get("full_name", "-") if s.get("profiles") else "-"
            s["exam_title"] = (s.get("exams") or {}).get("title", "-") if s.get("exams") else "-"
            if s["status"] == "submitted":
                pending_subs.append(s)
            elif s["status"] in ("graded", "published"):
                graded_subs.append(s)
    return render_template("teacher/grading.html", pending_subs=pending_subs, graded_subs=graded_subs)


@teacher_bp.route("/analytics")
@teacher_or_admin_required
def analytics():
    import statistics
    supabase = get_supabase()
    exams = supabase.table("exams").select("id,title,passing_score").eq("teacher_id", g.user_id).execute().data or []
    exam_ids = [e["id"] for e in exams]
    all_scores = []
    exam_breakdown = []
    dist_bins = [0, 0, 0, 0, 0]
    exam_labels = []
    exam_avgs = []
    exam_medians = []
    total_submissions = 0
    pass_count = 0
    for e in exams:
        subs = supabase.table("submissions").select("score,final_score").eq("exam_id", e["id"]).execute().data or []
        scores = [float(s.get("final_score") or s.get("score") or 0) for s in subs if s.get("final_score") or s.get("score")]
        all_scores.extend(scores)
        total_submissions += len(subs)
        passing = e.get("passing_score") or 70
        pc = sum(1 for sc in scores if sc >= passing)
        pass_count += pc
        if scores:
            sorted_s = sorted(scores)
            n = len(sorted_s)
            median = sorted_s[n // 2] if n % 2 == 1 else (sorted_s[n // 2 - 1] + sorted_s[n // 2]) / 2
            exam_breakdown.append({
                "title": e["title"],
                "count": len(scores),
                "avg": round(sum(scores) / len(scores), 1),
                "median": round(median, 1),
                "max": round(max(scores), 1),
                "min": round(min(scores), 1),
                "pass_pct": round(pc / len(scores) * 100),
            })
            exam_labels.append(e["title"][:20])
            exam_avgs.append(round(sum(scores) / len(scores), 1))
            exam_medians.append(round(median, 1))
    for sc in all_scores:
        if sc < 20: dist_bins[0] += 1
        elif sc < 40: dist_bins[1] += 1
        elif sc < 60: dist_bins[2] += 1
        elif sc < 80: dist_bins[3] += 1
        else: dist_bins[4] += 1
    avg_score = round(sum(all_scores) / len(all_scores), 1) if all_scores else 0
    pass_rate = round(pass_count / len(all_scores) * 100) if all_scores else 0
    std_dev = round(statistics.stdev(all_scores), 1) if len(all_scores) > 1 else 0
    stats = {
        "total_exams": len(exams),
        "total_submissions": total_submissions,
        "avg_score": avg_score,
        "pass_rate": pass_rate,
        "std_dev": std_dev,
    }
    return render_template("teacher/analytics.html", stats=stats, exam_breakdown=exam_breakdown, dist_bins=dist_bins, exam_labels=exam_labels, exam_avgs=exam_avgs, exam_medians=exam_medians)


@teacher_bp.route("/reset-password", methods=["POST"])
@login_required
def teacher_reset_password():
    if g.get("user_role") not in ("guru",):
        return jsonify({"error": "Forbidden"}), 403
    supabase = get_supabase()
    pw = request.form.get("password", "").strip()
    if len(pw) < 6:
        return jsonify({"error": "Password minimal 6 karakter"}), 400
    try:
        supabase.auth.admin.update_user_by_id(g.user_id, {"password": pw})
        log_activity("reset_password", "user", g.user_id, user_id=g.user_id)
        return jsonify({"success": True, "message": "Password berhasil diubah"})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@teacher_bp.route("/profile/update", methods=["POST"])
@login_required
def teacher_update_profile():
    if g.get("user_role") not in ("guru",):
        return jsonify({"error": "Forbidden"}), 403
    supabase = get_supabase()
    data = {}
    for key in ("phone",):
        val = request.form.get(key)
        if val is not None:
            data[key] = val.strip()
    if not data:
        return jsonify({"error": "Tidak ada data yang diubah"}), 400
    try:
        supabase.table("profiles").update(data).eq("id", g.user_id).execute()
        log_activity("update", "profile", g.user_id, new_data=data, user_id=g.user_id)
        return jsonify({"success": True, "message": "Profil berhasil diperbarui"})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@teacher_bp.route("/assignments", methods=["GET", "POST"])
@teacher_or_admin_required
def assignments():
    supabase = get_supabase()
    school_id = None
    try:
        profile = supabase.table("profiles").select("*").eq("id", g.user_id).single().execute().data or {}
        school_id = profile.get("school_id")
    except Exception:
        pass
    if not school_id:
        if request.is_json or request.headers.get("HX-Request"):
            return jsonify({"error": "School not found"}), 400
        return redirect("/teacher/dashboard")

    if request.method == "POST":
        class_id = request.form.get("class_id")
        subject_id = request.form.get("subject_id")
        if not class_id or not subject_id:
            if request.is_json or request.headers.get("HX-Request"):
                return jsonify({"error": "Class and subject required"}), 400
            return redirect("/teacher/dashboard")
        try:
            res = supabase.table("teacher_assignments").upsert({
                "teacher_id": g.user_id,
                "class_id": class_id,
                "subject_id": subject_id,
                "school_id": school_id,
            }).execute()
            aid = res.data[0]["id"] if res.data else None
            log_activity("create", "teacher_assignment", aid, new_data={"class_id": class_id, "subject_id": subject_id}, user_id=g.user_id)
            if request.is_json or request.headers.get("HX-Request"):
                return jsonify({"success": True})
            return redirect("/teacher/dashboard")
        except Exception as e:
            if request.is_json or request.headers.get("HX-Request"):
                return jsonify({"error": str(e)}), 400
            return redirect("/teacher/dashboard")

    try:
        assignments = supabase.table("teacher_assignments") \
            .select("*, classes(id, name, grade_level), subjects(id, name, code)") \
            .eq("teacher_id", g.user_id) \
            .eq("school_id", school_id) \
            .execute().data or []
    except Exception:
        assignments = []
    return jsonify(assignments)


@teacher_bp.route("/assignments/<assignment_id>", methods=["DELETE"])
@teacher_or_admin_required
def delete_assignment(assignment_id):
    supabase = get_supabase()
    try:
        supabase.table("teacher_assignments").delete().eq("id", assignment_id).eq("teacher_id", g.user_id).execute()
        log_activity("delete", "teacher_assignment", assignment_id, user_id=g.user_id)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@teacher_bp.route("/api/ai/wizard-status")
@teacher_or_admin_required
def ai_wizard_status():
    from app.services.ai_service import get_teacher_ai_status
    return jsonify(get_teacher_ai_status(g.user_id))


@teacher_bp.route("/api/ai/test-demo", methods=["POST"])
@teacher_or_admin_required
def ai_test_demo():
    from app.services.ai_service import _get_demo_key, _call_ai
    key = _get_demo_key()
    if not key:
        return jsonify({"error": "Demo key tidak tersedia. Hubungi admin."}), 400
    try:
        raw = _call_ai(key, 'Jawab dalam satu kata: Berapa 2+2? Format JSON: {"answer": <number>}')
        import json
        data = json.loads(raw.strip().replace("```json", "").replace("```", "").strip())
        if data.get("answer") == 4:
            return jsonify({"success": True, "message": "✅ Demo AI aktif! Koneksi berhasil."})
        return jsonify({"success": True, "message": f"✅ Demo AI aktif. Response: {raw[:80]}"})
    except Exception as e:
        return jsonify({"error": f"❌ Gagal: {str(e)[:120]}"}), 400


@teacher_bp.route("/ai-settings")
@teacher_or_admin_required
def ai_settings():
    supabase = get_supabase()
    keys = supabase.table("teacher_ai_keys").select("*").eq("teacher_id", g.user_id).order("created_at", desc=True).execute().data or []
    settings_res = supabase.table("teacher_ai_settings").select("*").eq("teacher_id", g.user_id).limit(1).execute()
    if not settings_res.data:
        default_prompts = [
            {"id": "default", "label": "Default (Semua Mapel)", "template": "Kamu adalah asisten koreksi ujian. Koreksi jawaban esai berikut berdasarkan soal dan bobot maksimal.\n\nSoal: {question}\nPedoman Penskoran: {rubric}\nBobot Maksimal: {max_score} poin\nJawaban Siswa: \"{answer}\"\n\nBerikan skor (0-{max_score}) dan feedback singkat dalam bahasa Indonesia.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
            {"id": "ipa", "label": "IPA / Sains", "template": "Kamu adalah asisten koreksi mata pelajaran IPA (Fisika, Kimia, Biologi, Earth Science).\n\nSoal: {question}\nPedoman Penskoran: {rubric}\nBobot Maksimal: {max_score} poin\nJawaban Siswa: \"{answer}\"\n\nBerikan skor (0-{max_score}). Nilai berdasarkan: keakuratan konsep sains, penggunaan istilah ilmiah yang tepat, logika ilmiah, dan kelengkapan jawaban. Feedback dalam bahasa Indonesia.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
            {"id": "matematika", "label": "Matematika", "template": "Kamu adalah asisten koreksi mata pelajaran Matematika.\n\nSoal: {question}\nPedoman Penskoran: {rubric}\nBobot Maksimal: {max_score} poin\nJawaban Siswa: \"{answer}\"\n\nBerikan skor (0-{max_score}). Nilai berdasarkan: kebenaran rumus, ketepatan langkah-langkah penyelesaian, keakuratan perhitungan, dan kesimpulan akhir. Feedback dalam bahasa Indonesia.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
            {"id": "bahasa", "label": "Bahasa (Inggris/Indonesia/Arab/Mandarin)", "template": "Kamu adalah asisten koreksi mata pelajaran Bahasa (Indonesia, Inggris, Arab, Mandarin).\n\nSoal: {question}\nPedoman Penskoran: {rubric}\nBobot Maksimal: {max_score} poin\nJawaban Siswa: \"{answer}\"\n\nBerikan skor (0-{max_score}). Nilai berdasarkan: tata bahasa (grammar/tata bahasa), kosa kata (vocabulary/kosakata), struktur tulisan, kesesuaian konteks, dan kreativitas. Feedback dalam bahasa Indonesia.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
            {"id": "ips", "label": "IPS / Sosial", "template": "Kamu adalah asisten koreksi mata pelajaran IPS (Geografi, Sosiologi, Ekonomi).\n\nSoal: {question}\nPedoman Penskoran: {rubric}\nBobot Maksimal: {max_score} poin\nJawaban Siswa: \"{answer}\"\n\nBerikan skor (0-{max_score}). Nilai berdasarkan: kedalaman analisis, penggunaan data/contoh konkret, argumen logis, dan keterkaitan antar konsep sosial. Feedback dalam bahasa Indonesia.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
            {"id": "ict", "label": "ICT / Coding", "template": "Kamu adalah asisten koreksi mata pelajaran ICT, Coding, dan Computer Science.\n\nSoal: {question}\nPedoman Penskoran: {rubric}\nBobot Maksimal: {max_score} poin\nJawaban Siswa: \"{answer}\"\n\nBerikan skor (0-{max_score}). Nilai berdasarkan: kebenaran logika algoritma, sintaks kode, efisiensi solusi, dan dokumentasi/penjelasan. Feedback dalam bahasa Indonesia.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
            {"id": "agama", "label": "Agama", "template": "Kamu adalah asisten koreksi mata pelajaran Pendidikan Agama.\n\nSoal: {question}\nPedoman Penskoran: {rubric}\nBobot Maksimal: {max_score} poin\nJawaban Siswa: \"{answer}\"\n\nBerikan skor (0-{max_score}). Nilai berdasarkan: pemahaman konsep keagamaan, ketepatan dalil/sumber, implementasi dalam kehidupan, dan sikap toleransi. Feedback dalam bahasa Indonesia.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
            {"id": "penjas", "label": "PJOK / Olahraga", "template": "Kamu adalah asisten koreksi mata pelajaran Pendidikan Jasmani, Olahraga, dan Kesehatan.\n\nSoal: {question}\nPedoman Penskoran: {rubric}\nBobot Maksimal: {max_score} poin\nJawaban Siswa: \"{answer}\"\n\nBerikan skor (0-{max_score}). Nilai berdasarkan: pemahaman teori olahraga, teknik gerakan, keselamatan, dan kebugaran jasmani. Feedback dalam bahasa Indonesia.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
            {"id": "iot", "label": "IoT / Teknologi", "template": "Kamu adalah asisten koreksi mata pelajaran Internet of Things dan Teknologi Embedded.\n\nSoal: {question}\nPedoman Penskoran: {rubric}\nBobot Maksimal: {max_score} poin\nJawaban Siswa: \"{answer}\"\n\nBerikan skor (0-{max_score}). Nilai berdasarkan: pemahaman sistem IoT, integrasi sensor, jaringan komunikasi, dan pemecahan masalah teknis. Feedback dalam bahasa Indonesia.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
            {"id": "ketat", "label": "Ketat (Semua Mapel)", "template": "Kamu adalah pemeriksa ujian yang sangat ketat. Koreksi jawaban esai berikut.\n\nSoal: {question}\nBobot Maksimal: {max_score} poin\nJawaban: \"{answer}\"\n\nBerikan skor (0-{max_score}). Jangan mudah memberi nilai tinggi. Feedback harus menyebutkan kekurangan secara spesifik.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
            {"id": "ringan", "label": "Ringan (Semua Mapel)", "template": "Kamu adalah guru yang baik hati dan memotivasi. Koreksi jawaban esai berikut.\n\nSoal: {question}\nBobot Maksimal: {max_score} poin\nJawaban: \"{answer}\"\n\nBerikan skor (0-{max_score}). Beri nilai maksimal jika jawaban mendekati benar. Feedback yang membangun dan memotivasi.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
        ]
        supabase.table("teacher_ai_settings").insert({
            "teacher_id": g.user_id, "prompts": default_prompts, "active_prompt_id": "default"
        }).execute()
        settings = {"teacher_id": g.user_id, "prompts": default_prompts, "active_prompt_id": "default"}
    else:
        settings = settings_res.data[0]
        pr = settings.get("prompts")
        if isinstance(pr, str):
            try:
                settings["prompts"] = json.loads(pr)
            except (json.JSONDecodeError, TypeError):
                settings["prompts"] = []
        elif pr is None:
            settings["prompts"] = []
        # Auto-upgrade if less than 3 prompts (old version)
        try:
            if len(settings.get("prompts", [])) < 3:
                default_prompts = [
                    {"id": "default", "label": "Default (Semua Mapel)", "template": "Kamu adalah asisten koreksi ujian. Koreksi jawaban esai berikut berdasarkan soal dan bobot maksimal.\n\nSoal: {question}\nPedoman Penskoran: {rubric}\nBobot Maksimal: {max_score} poin\nJawaban Siswa: \"{answer}\"\n\nBerikan skor (0-{max_score}) dan feedback singkat dalam bahasa Indonesia.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
                    {"id": "ipa", "label": "IPA / Sains", "template": "Kamu adalah asisten koreksi mata pelajaran IPA (Fisika, Kimia, Biologi, Earth Science).\n\nSoal: {question}\nPedoman Penskoran: {rubric}\nBobot Maksimal: {max_score} poin\nJawaban Siswa: \"{answer}\"\n\nBerikan skor (0-{max_score}). Nilai berdasarkan: keakuratan konsep sains, penggunaan istilah ilmiah yang tepat, logika ilmiah, dan kelengkapan jawaban. Feedback dalam bahasa Indonesia.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
                    {"id": "matematika", "label": "Matematika", "template": "Kamu adalah asisten koreksi mata pelajaran Matematika.\n\nSoal: {question}\nPedoman Penskoran: {rubric}\nBobot Maksimal: {max_score} poin\nJawaban Siswa: \"{answer}\"\n\nBerikan skor (0-{max_score}). Nilai berdasarkan: kebenaran rumus, ketepatan langkah-langkah penyelesaian, keakuratan perhitungan, dan kesimpulan akhir. Feedback dalam bahasa Indonesia.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
                    {"id": "bahasa", "label": "Bahasa (Inggris/Indonesia/Arab/Mandarin)", "template": "Kamu adalah asisten koreksi mata pelajaran Bahasa (Indonesia, Inggris, Arab, Mandarin).\n\nSoal: {question}\nPedoman Penskoran: {rubric}\nBobot Maksimal: {max_score} poin\nJawaban Siswa: \"{answer}\"\n\nBerikan skor (0-{max_score}). Nilai berdasarkan: tata bahasa (grammar/tata bahasa), kosa kata (vocabulary/kosakata), struktur tulisan, kesesuaian konteks, dan kreativitas. Feedback dalam bahasa Indonesia.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
                    {"id": "ips", "label": "IPS / Sosial", "template": "Kamu adalah asisten koreksi mata pelajaran IPS (Geografi, Sosiologi, Ekonomi).\n\nSoal: {question}\nPedoman Penskoran: {rubric}\nBobot Maksimal: {max_score} poin\nJawaban Siswa: \"{answer}\"\n\nBerikan skor (0-{max_score}). Nilai berdasarkan: kedalaman analisis, penggunaan data/contoh konkret, argumen logis, dan keterkaitan antar konsep sosial. Feedback dalam bahasa Indonesia.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
                    {"id": "ict", "label": "ICT / Coding", "template": "Kamu adalah asisten koreksi mata pelajaran ICT, Coding, dan Computer Science.\n\nSoal: {question}\nPedoman Penskoran: {rubric}\nBobot Maksimal: {max_score} poin\nJawaban Siswa: \"{answer}\"\n\nBerikan skor (0-{max_score}). Nilai berdasarkan: kebenaran logika algoritma, sintaks kode, efisiensi solusi, dan dokumentasi/penjelasan. Feedback dalam bahasa Indonesia.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
                    {"id": "agama", "label": "Agama", "template": "Kamu adalah asisten koreksi mata pelajaran Pendidikan Agama.\n\nSoal: {question}\nPedoman Penskoran: {rubric}\nBobot Maksimal: {max_score} poin\nJawaban Siswa: \"{answer}\"\n\nBerikan skor (0-{max_score}). Nilai berdasarkan: pemahaman konsep keagamaan, ketepatan dalil/sumber, implementasi dalam kehidupan, dan sikap toleransi. Feedback dalam bahasa Indonesia.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
                    {"id": "penjas", "label": "PJOK / Olahraga", "template": "Kamu adalah asisten koreksi mata pelajaran Pendidikan Jasmani, Olahraga, dan Kesehatan.\n\nSoal: {question}\nPedoman Penskoran: {rubric}\nBobot Maksimal: {max_score} poin\nJawaban Siswa: \"{answer}\"\n\nBerikan skor (0-{max_score}). Nilai berdasarkan: pemahaman teori olahraga, teknik gerakan, keselamatan, dan kebugaran jasmani. Feedback dalam bahasa Indonesia.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
                    {"id": "iot", "label": "IoT / Teknologi", "template": "Kamu adalah asisten koreksi mata pelajaran Internet of Things dan Teknologi Embedded.\n\nSoal: {question}\nPedoman Penskoran: {rubric}\nBobot Maksimal: {max_score} poin\nJawaban Siswa: \"{answer}\"\n\nBerikan skor (0-{max_score}). Nilai berdasarkan: pemahaman sistem IoT, integrasi sensor, jaringan komunikasi, dan pemecahan masalah teknis. Feedback dalam bahasa Indonesia.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
                    {"id": "ketat", "label": "Ketat (Semua Mapel)", "template": "Kamu adalah pemeriksa ujian yang sangat ketat. Koreksi jawaban esai berikut.\n\nSoal: {question}\nBobot Maksimal: {max_score} poin\nJawaban: \"{answer}\"\n\nBerikan skor (0-{max_score}). Jangan mudah memberi nilai tinggi. Feedback harus menyebutkan kekurangan secara spesifik.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
                    {"id": "ringan", "label": "Ringan (Semua Mapel)", "template": "Kamu adalah guru yang baik hati dan memotivasi. Koreksi jawaban esai berikut.\n\nSoal: {question}\nBobot Maksimal: {max_score} poin\nJawaban: \"{answer}\"\n\nBerikan skor (0-{max_score}). Beri nilai maksimal jika jawaban mendekati benar. Feedback yang membangun dan memotivasi.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
                ]
                supabase.table("teacher_ai_settings").update({
                    "prompts": default_prompts, "active_prompt_id": "default"
                }).eq("teacher_id", g.user_id).execute()
                settings["prompts"] = default_prompts
                settings["active_prompt_id"] = "default"
        except Exception as e:
            current_app.logger.error(f"Auto-upgrade prompts error: {e}")
    return render_template("teacher/ai_settings.html", keys=keys, settings=settings)


@teacher_bp.route("/ai-settings/add-key", methods=["POST"])
@subscription_write_required
@teacher_or_admin_required
def ai_add_key():
    supabase = get_supabase()
    provider = request.form.get("provider", "gemini")
    api_key = request.form.get("api_key", "").strip()
    label = request.form.get("label", "").strip()
    base_url = request.form.get("base_url", "").strip()
    model_name = request.form.get("model_name", "").strip()
    if not api_key:
        flash("API Key wajib diisi", "error")
        return redirect("/teacher/ai-settings")
    supabase.table("teacher_ai_keys").update({"is_active": False}).eq("teacher_id", g.user_id).execute()
    data = {
        "teacher_id": g.user_id, "provider": provider,
        "api_key": api_key, "label": label or provider,
        "is_active": True,
    }
    if provider == "custom":
        data["base_url"] = base_url
        data["model_name"] = model_name or "gpt-4o-mini"
    supabase.table("teacher_ai_keys").insert(data).execute()
    flash("API Key berhasil ditambahkan dan diaktifkan", "success")
    return redirect("/teacher/ai-settings")


@teacher_bp.route("/ai-settings/<key_id>/toggle", methods=["POST"])
@subscription_write_required
@teacher_or_admin_required
def ai_toggle_key(key_id):
    supabase = get_supabase()
    key = supabase.table("teacher_ai_keys").select("is_active").eq("id", key_id).eq("teacher_id", g.user_id).single().execute()
    if key.data:
        if key.data.get("is_active"):
            supabase.table("teacher_ai_keys").update({"is_active": False}).eq("id", key_id).execute()
        else:
            supabase.table("teacher_ai_keys").update({"is_active": False}).eq("teacher_id", g.user_id).execute()
            supabase.table("teacher_ai_keys").update({"is_active": True}).eq("id", key_id).execute()
    return redirect("/teacher/ai-settings")


@teacher_bp.route("/ai-settings/<key_id>/delete", methods=["POST"])
@subscription_write_required
@teacher_or_admin_required
def ai_delete_key(key_id):
    supabase = get_supabase()
    supabase.table("teacher_ai_keys").delete().eq("id", key_id).eq("teacher_id", g.user_id).execute()
    flash("API Key berhasil dihapus", "success")
    return redirect("/teacher/ai-settings")


@teacher_bp.route("/ai-settings/save-prompt", methods=["POST"])
@teacher_or_admin_required
def ai_save_prompt():
    supabase = get_supabase()
    prompts_raw = request.form.get("prompts", "[]").strip()
    active_id = request.form.get("active_prompt_id", "default").strip()
    try:
        prompts = json.loads(prompts_raw)
    except json.JSONDecodeError:
        flash("Data prompts tidak valid", "error")
        return redirect("/teacher/ai-settings")
    supabase.table("teacher_ai_settings").upsert({
        "teacher_id": g.user_id, "prompts": prompts, "active_prompt_id": active_id
    }).execute()
    flash("Prompt berhasil disimpan", "success")
    return redirect("/teacher/ai-settings")


@teacher_bp.route("/ai-settings/reset-prompt", methods=["POST"])
@teacher_or_admin_required
def ai_reset_prompt():
    default_prompts = [
        {"id": "default", "label": "Default (Semua Mapel)", "template": "Kamu adalah asisten koreksi ujian. Koreksi jawaban esai berikut berdasarkan soal dan bobot maksimal.\n\nSoal: {question}\nPedoman Penskoran: {rubric}\nBobot Maksimal: {max_score} poin\nJawaban Siswa: \"{answer}\"\n\nBerikan skor (0-{max_score}) dan feedback singkat dalam bahasa Indonesia.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
        {"id": "ipa", "label": "IPA / Sains", "template": "Kamu adalah asisten koreksi mata pelajaran IPA (Fisika, Kimia, Biologi, Earth Science).\n\nSoal: {question}\nPedoman Penskoran: {rubric}\nBobot Maksimal: {max_score} poin\nJawaban Siswa: \"{answer}\"\n\nBerikan skor (0-{max_score}). Nilai berdasarkan: keakuratan konsep sains, penggunaan istilah ilmiah yang tepat, logika ilmiah, dan kelengkapan jawaban. Feedback dalam bahasa Indonesia.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
        {"id": "matematika", "label": "Matematika", "template": "Kamu adalah asisten koreksi mata pelajaran Matematika.\n\nSoal: {question}\nPedoman Penskoran: {rubric}\nBobot Maksimal: {max_score} poin\nJawaban Siswa: \"{answer}\"\n\nBerikan skor (0-{max_score}). Nilai berdasarkan: kebenaran rumus, ketepatan langkah-langkah penyelesaian, keakuratan perhitungan, dan kesimpulan akhir. Feedback dalam bahasa Indonesia.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
        {"id": "bahasa", "label": "Bahasa (Inggris/Indonesia/Arab/Mandarin)", "template": "Kamu adalah asisten koreksi mata pelajaran Bahasa (Indonesia, Inggris, Arab, Mandarin).\n\nSoal: {question}\nPedoman Penskoran: {rubric}\nBobot Maksimal: {max_score} poin\nJawaban Siswa: \"{answer}\"\n\nBerikan skor (0-{max_score}). Nilai berdasarkan: tata bahasa (grammar/tata bahasa), kosa kata (vocabulary/kosakata), struktur tulisan, kesesuaian konteks, dan kreativitas. Feedback dalam bahasa Indonesia.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
        {"id": "ips", "label": "IPS / Sosial", "template": "Kamu adalah asisten koreksi mata pelajaran IPS (Geografi, Sosiologi, Ekonomi).\n\nSoal: {question}\nPedoman Penskoran: {rubric}\nBobot Maksimal: {max_score} poin\nJawaban Siswa: \"{answer}\"\n\nBerikan skor (0-{max_score}). Nilai berdasarkan: kedalaman analisis, penggunaan data/contoh konkret, argumen logis, dan keterkaitan antar konsep sosial. Feedback dalam bahasa Indonesia.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
        {"id": "ict", "label": "ICT / Coding", "template": "Kamu adalah asisten koreksi mata pelajaran ICT, Coding, dan Computer Science.\n\nSoal: {question}\nPedoman Penskoran: {rubric}\nBobot Maksimal: {max_score} poin\nJawaban Siswa: \"{answer}\"\n\nBerikan skor (0-{max_score}). Nilai berdasarkan: kebenaran logika algoritma, sintaks kode, efisiensi solusi, dan dokumentasi/penjelasan. Feedback dalam bahasa Indonesia.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
        {"id": "agama", "label": "Agama", "template": "Kamu adalah asisten koreksi mata pelajaran Pendidikan Agama.\n\nSoal: {question}\nPedoman Penskoran: {rubric}\nBobot Maksimal: {max_score} poin\nJawaban Siswa: \"{answer}\"\n\nBerikan skor (0-{max_score}). Nilai berdasarkan: pemahaman konsep keagamaan, ketepatan dalil/sumber, implementasi dalam kehidupan, dan sikap toleransi. Feedback dalam bahasa Indonesia.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
        {"id": "penjas", "label": "PJOK / Olahraga", "template": "Kamu adalah asisten koreksi mata pelajaran Pendidikan Jasmani, Olahraga, dan Kesehatan.\n\nSoal: {question}\nPedoman Penskoran: {rubric}\nBobot Maksimal: {max_score} poin\nJawaban Siswa: \"{answer}\"\n\nBerikan skor (0-{max_score}). Nilai berdasarkan: pemahaman teori olahraga, teknik gerakan, keselamatan, dan kebugaran jasmani. Feedback dalam bahasa Indonesia.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
        {"id": "iot", "label": "IoT / Teknologi", "template": "Kamu adalah asisten koreksi mata pelajaran Internet of Things dan Teknologi Embedded.\n\nSoal: {question}\nPedoman Penskoran: {rubric}\nBobot Maksimal: {max_score} poin\nJawaban Siswa: \"{answer}\"\n\nBerikan skor (0-{max_score}). Nilai berdasarkan: pemahaman sistem IoT, integrasi sensor, jaringan komunikasi, dan pemecahan masalah teknis. Feedback dalam bahasa Indonesia.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
        {"id": "ketat", "label": "Ketat (Semua Mapel)", "template": "Kamu adalah pemeriksa ujian yang sangat ketat. Koreksi jawaban esai berikut.\n\nSoal: {question}\nBobot Maksimal: {max_score} poin\nJawaban: \"{answer}\"\n\nBerikan skor (0-{max_score}). Jangan mudah memberi nilai tinggi. Feedback harus menyebutkan kekurangan secara spesifik.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
        {"id": "ringan", "label": "Ringan (Semua Mapel)", "template": "Kamu adalah guru yang baik hati dan memotivasi. Koreksi jawaban esai berikut.\n\nSoal: {question}\nBobot Maksimal: {max_score} poin\nJawaban: \"{answer}\"\n\nBerikan skor (0-{max_score}). Beri nilai maksimal jika jawaban mendekati benar. Feedback yang membangun dan memotivasi.\nFormat JSON: {\"score\": <number>, \"feedback\": \"<string>\"}"},
    ]
    supabase = get_supabase()
    supabase.table("teacher_ai_settings").upsert({
        "teacher_id": g.user_id, "prompts": default_prompts, "active_prompt_id": "default"
    }).execute()
    flash("Prompt dikembalikan ke default", "success")
    return redirect("/teacher/ai-settings")


@teacher_bp.route("/students")
@teacher_or_admin_required
def students():
    supabase = get_supabase()
    school_id = g.get("user_school_id")
    query = supabase.table("profiles").select("id,full_name,phone,role").eq("role", "murid")
    if school_id:
        query = query.eq("school_id", school_id)
    students = query.execute().data or []
    exam_ids = [e["id"] for e in supabase.table("exams").select("id").eq("teacher_id", g.user_id).execute().data or []]
    if exam_ids:
        subs = supabase.table("submissions").select("student_id,score,final_score").in_("exam_id", exam_ids).execute().data or []
        sub_map = {}
        for s in subs:
            sid = s["student_id"]
            if sid not in sub_map:
                sub_map[sid] = []
            sc = float(s.get("final_score") or s.get("score") or 0)
            sub_map[sid].append(sc)
        for st in students:
            st_scores = sub_map.get(st["id"], [])
            st["sub_count"] = len(st_scores)
            st["avg_score"] = round(sum(st_scores) / len(st_scores), 1) if st_scores else None
    return render_template("teacher/students.html", students=students)


@teacher_bp.route("/classes")
@teacher_or_admin_required
def teacher_classes():
    supabase = get_supabase()
    sid = g.get("user_school_id")
    assignments = []
    school_info = {}
    active_year = None
    if sid:
        try:
            assignments = supabase.table("teacher_assignments") \
                .select("*, classes!inner(id, name), subjects!inner(id, name)") \
                .eq("teacher_id", g.user_id) \
                .eq("school_id", sid) \
                .execute().data or []
        except Exception:
            assignments = []
        try:
            sch = supabase.table("schools").select("name, npsn").eq("id", sid).single().execute()
            if sch.data: school_info = sch.data
        except: pass
        try:
            years = supabase.table("school_years").select("*").eq("school_id", sid).eq("is_active", True).limit(1).execute()
            if years.data: active_year = years.data[0]
        except: pass
    return render_template("teacher/classes.html", assignments=assignments,
                           school_info=school_info, active_year=active_year)


@teacher_bp.route("/tools/exam/<exam_id>/check-pdf", methods=["GET"])
@teacher_or_admin_required
@require_school_access("exams", "exam_id")
def exam_check_pdf(exam_id):
    """Diagnostic endpoint: returns exam's PDF fields and file status."""
    supabase = get_supabase()
    exam = supabase.table("exams").select("id,pdf_url,pdf_page_urls,title,status,answer_key").eq("id", exam_id).single().execute().data
    if not exam:
        return jsonify({"error": "not found"}), 404
    page_urls = exam.get("pdf_page_urls") or []
    pdf_url = exam.get("pdf_url") or ""
    exam_dir = os.path.join(current_app.root_path, "static", "uploads", "exams", exam_id)
    files_on_disk = []
    if os.path.isdir(exam_dir):
        files_on_disk = os.listdir(exam_dir)
    return jsonify({
        "exam_id": exam_id,
        "title": exam.get("title"),
        "status": exam.get("status"),
        "pdf_url": pdf_url,
        "pdf_page_urls_count": len(page_urls),
        "pdf_page_urls": page_urls[:3],
        "has_valid_pages": len(page_urls) > 0,
        "files_in_exam_dir": files_on_disk,
        "exam_dir": exam_dir,
        "root_path": current_app.root_path,
    })


@teacher_bp.route("/tools/exam/<exam_id>/reprocess-pdf", methods=["POST"])
@teacher_or_admin_required
@require_school_access("exams", "exam_id")
def exam_reprocess_pdf(exam_id):
    """Reprocess PDF for existing exam: regenerate local page images."""
    from app.services.pdf_service import upload_pdf

    supabase = get_supabase()
    exam = supabase.table("exams").select("pdf_url,pdf_page_urls,title").eq("id", exam_id).single().execute().data
    if not exam:
        return jsonify({"error": "not found"}), 404

    # Strategy 1: local exam.pdf already exists → regenerate from that
    local_pdf = os.path.join(current_app.root_path, "static", "uploads", "exams", exam_id, "exam.pdf")
    if os.path.exists(local_pdf):
        try:
            with open(local_pdf, "rb") as f:
                raw = f.read()
            if raw[:4] == b'%PDF':
                class _MF:
                    def __init__(self, d, n): self._d = d; self.filename = n
                    def read(self): return self._d
                result = upload_pdf(_MF(raw, "exam.pdf"), exam_id)
                supabase.table("exams").update({
                    "pdf_url": result["pdf_path"],
                    "pdf_page_urls": result["page_urls"],
                }).eq("id", exam_id).execute()
                return jsonify({"success": True, "source": "local_pdf", "pages": result["total_pages"]})
        except Exception as e:
            current_app.logger.warning("Reprocess from local PDF failed: %s", e)

    # Strategy 2: look for temp files in uploads/exams
    upload_dir = os.path.join(current_app.root_path, "static", "uploads", "exams")
    if os.path.isdir(upload_dir):
        temp_files = sorted(
            [os.path.join(upload_dir, f) for f in os.listdir(upload_dir)
             if f.startswith("temp_") and f.endswith(".pdf")],
            key=os.path.getmtime, reverse=True
        )
        for fp in temp_files:
            try:
                with open(fp, "rb") as f:
                    raw = f.read()
                if raw[:4] != b'%PDF':
                    continue
                class _MF:
                    def __init__(self, d, n): self._d = d; self.filename = n
                    def read(self): return self._d
                result = upload_pdf(_MF(raw, "exam.pdf"), exam_id)
                supabase.table("exams").update({
                    "pdf_url": result["pdf_path"],
                    "pdf_page_urls": result["page_urls"],
                }).eq("id", exam_id).execute()
                return jsonify({"success": True, "source": "temp_file", "file": os.path.basename(fp), "pages": result["total_pages"]})
            except Exception as e:
                current_app.logger.warning("Reprocess from temp file failed: %s", e)
                continue

    return jsonify({"error": "No PDF source found (no local PDF, no temp files)"}), 404


@teacher_bp.route("/exams/<exam_id>/proctoring")
@teacher_or_admin_required
@require_school_access("exams", "exam_id")
def exam_proctoring(exam_id):
    """Proctoring dashboard — live view of student exam progress."""
    supabase = get_supabase()
    exam = supabase.table("exams").select("id,title,subject,total_questions,duration_minutes,start_at,status").eq("id", exam_id).single().execute().data or {}
    return render_template("teacher/proctoring.html", exam=exam, exam_id=exam_id)


@teacher_bp.route("/api/exams/<exam_id>/proctoring-data")
@teacher_or_admin_required
@require_school_access("exams", "exam_id")
def exam_proctoring_data(exam_id):
    """API: return live proctoring data (submissions + violations) for an exam."""
    supabase = get_supabase()

    # Get all students for this exam's class(es)
    exam = supabase.table("exams").select("class_ids,total_questions").eq("id", exam_id).single().execute().data or {}
    class_ids = exam.get("class_ids") or []
    total_q = exam.get("total_questions", 0)

    students = []
    if class_ids:
        profile_ids = supabase.table("student_classes") \
            .select("student_id") \
            .in_("class_id", class_ids) \
            .execute().data or []
        sids = [p["student_id"] for p in profile_ids]
        if sids:
            profiles = supabase.table("profiles") \
                .select("id,full_name") \
                .in_("id", sids) \
                .execute().data or []
            students = profiles

    # Get submissions for this exam
    subs = supabase.table("submissions") \
        .select("student_id,status,answers,submitted_at,updated_at") \
        .eq("exam_id", exam_id) \
        .execute().data or []

    sub_map = {s["student_id"]: s for s in subs}

    # Get violation counts
    try:
        viols = supabase.table("violation_logs") \
            .select("user_id") \
            .eq("exam_id", exam_id) \
            .execute().data or []
    except Exception:
        viols = []

    viol_count = {}
    for v in viols:
        uid = v.get("user_id", "")
        viol_count[uid] = viol_count.get(uid, 0) + 1

    now = datetime.now(timezone.utc)

    result = []
    for s in students:
        sid = s["id"]
        sub = sub_map.get(sid)
        answers = sub.get("answers") or {} if sub else {}
        if isinstance(answers, str):
            try:
                answers = json.loads(answers)
            except Exception:
                answers = {}

        # Count how many questions have answers
        ans_count = 0
        if isinstance(answers, dict):
            for v in answers.values():
                if isinstance(v, dict) and v.get("answer"):
                    ans_count += 1
                elif isinstance(v, str) and v:
                    ans_count += 1
                elif isinstance(v, dict):
                    ans_count += 1

        result.append({
            "id": sid,
            "name": s.get("full_name", sid[:12]),
            "status": sub.get("status", "not_started") if sub else "not_started",
            "answers_count": ans_count,
            "total_questions": total_q,
            "violations": viol_count.get(sid, 0),
            "updated_at": (sub.get("updated_at") or sub.get("submitted_at") or "").split(".")[0].replace("T", " ") if sub else "",
        })

    return jsonify({
        "students": result,
        "timestamp": now.isoformat(),
        "total_students": len(result),
        "started": sum(1 for r in result if r["status"] != "not_started"),
        "submitted": sum(1 for r in result if r["status"] in ("submitted", "graded", "published")),
    })


@teacher_bp.route("/exams/<exam_id>/generate-remedial", methods=["POST"])
@teacher_or_admin_required
@require_school_access("exams", "exam_id")
def generate_remedial(exam_id):
    """Analyze exam results and generate remedial questions via AI."""
    supabase = get_supabase()
    exam = supabase.table("exams").select("*").eq("id", exam_id).single().execute().data
    if not exam:
        return jsonify({"error": "Exam not found"}), 404

    # Get submissions with answers
    subs = supabase.table("submissions").select("id,student_id,answers,final_score").eq("exam_id", exam_id).execute().data or []

    # Analyze MCQ scores per question
    qtypes = exam.get("question_types") or {}
    if isinstance(qtypes, str):
        try: qtypes = json.loads(qtypes)
        except: qtypes = {}
    answer_key = exam.get("answer_key") or {}
    if isinstance(answer_key, str):
        try: answer_key = json.loads(answer_key)
        except: answer_key = {}
    question_texts = exam.get("question_texts") or {}
    if isinstance(question_texts, str):
        try: question_texts = json.loads(question_texts)
        except: question_texts = {}
    total_q = exam.get("total_questions", 0)

    # Count correct/wrong per MCQ question
    q_correct = {}
    q_total = {}
    for sub in subs:
        answers = sub.get("answers") or {}
        if isinstance(answers, str):
            try: answers = json.loads(answers)
            except: answers = {}
        for qi in range(total_q):
            qi_str = str(qi)
            if qtypes.get(qi_str) == "mcq":
                q_total[qi_str] = q_total.get(qi_str, 0) + 1
                if qi_str in answer_key and qi_str in answers:
                    stu_ans = answers[qi_str]
                    if isinstance(stu_ans, dict):
                        stu_ans = stu_ans.get("answer", "")
                    if stu_ans == answer_key[qi_str]:
                        q_correct[qi_str] = q_correct.get(qi_str, 0) + 1

    # Find top 3 most-failed MCQ questions
    fail_rate = []
    for qi in range(total_q):
        qi_str = str(qi)
        if q_total.get(qi_str, 0) > 0:
            correct = q_correct.get(qi_str, 0)
            total = q_total[qi_str]
            rate = (total - correct) / total
            fail_rate.append((qi, rate, q_total[qi_str]))

    fail_rate.sort(key=lambda x: x[1], reverse=True)
    worst_q = fail_rate[:3]

    # Build prompt for AI
    prompt_parts = ["Buat 5 soal remedial tipe MCQ berdasarkan analisis berikut:\n"]
    prompt_parts.append(f"Ujian: {exam.get('title', '')}\n")
    prompt_parts.append(f"Mata Pelajaran: {exam.get('subject', '')}\n\n")

    if worst_q:
        prompt_parts.append("Soal dengan tingkat kesalahan tertinggi:\n")
        for qi, rate, total in worst_q:
            q_text = question_texts.get(str(qi), f"Soal {qi+1}")
            q_key = answer_key.get(str(qi), "-")
            prompt_parts.append(f"Soal {qi+1} (salah {total - q_correct.get(str(qi), 0)}/{total} siswa): {q_text} (kunci: {q_key})")

    prompt_parts.append("""
    \nBuat 5 soal pilihan ganda dengan 5 opsi (A-E) yang mirip dengan soal-soal di atas.
    Setiap soal harus memiliki: nomor, pertanyaan, 5 opsi (A-E), dan kunci jawaban.

    Format output JSON:
    {"questions": [{"number": 1, "question": "teks soal", "options": {"A": "...", "B": "...", "C": "...", "D": "...", "E": "..."}, "answer": "A"}]}
    """)

    prompt = "\n".join(prompt_parts)

    # Call AI
    from app.services.ai_service import _get_active_key, _call_ai
    key = _get_active_key(g.user_id)
    if not key:
        return jsonify({"error": "Belum ada API key AI. Atur di Pengaturan AI."}), 400

    try:
        raw = _call_ai(key, prompt)
        import re
        cleaned = raw.strip()
        if cleaned.startswith("```"):
            cleaned = re.sub(r"^```(?:json)?\s*|```$", "", cleaned, flags=re.DOTALL).strip()
        data = json.loads(cleaned)
        questions = data.get("questions", [])
    except Exception as e:
        return jsonify({"error": f"AI gagal generate: {str(e)[:150]}"}), 500

    if not questions:
        return jsonify({"error": "AI tidak menghasilkan soal"}), 500

    # Create a new exam draft with remedial questions
    title = f"Remedial - {exam.get('title', 'Ujian')}"
    new_qtypes = {}
    new_answer_key = {}
    new_qtexts = {}
    for q in questions:
        qi = q["number"] - 1
        new_qtypes[str(qi)] = "mcq"
        new_answer_key[str(qi)] = q.get("answer", "A")
        opts = q.get("options", {})
        txt = q["question"]
        for k, v in opts.items():
            txt += f"\n{k}. {v}"
        new_qtexts[str(qi)] = txt

    total_new = len(questions)
    new_exam = {
        "teacher_id": g.user_id,
        "school_id": g.get("user_school_id"),
        "title": title,
        "subject": exam.get("subject", ""),
        "class_ids": exam.get("class_ids", []),
        "total_questions": total_new,
        "duration_minutes": min(exam.get("duration_minutes", 60) // 2, 30),
        "description": f"Soal remedial otomatis dari ujian {exam.get('title', '')}",
        "status": "draft",
        "is_published": False,
        "publish_mode": "manual",
        "question_types": new_qtypes,
        "answer_key": new_answer_key,
        "question_texts": new_qtexts,
    }

    try:
        res = supabase.table("exams").insert(new_exam).execute()
        new_id = res.data[0]["id"]
        return jsonify({"success": True, "redirect": f"/teacher/exams/{new_id}"})
    except Exception:
        # Fallback: try without newer fields
        for key in ["question_texts", "publish_mode"]:
            new_exam.pop(key, None)
        res = supabase.table("exams").insert(new_exam).execute()
        new_id = res.data[0]["id"]
        return jsonify({"success": True, "redirect": f"/teacher/exams/{new_id}"})


@teacher_bp.route("/exams/<exam_id>/cheat-analysis")
@teacher_or_admin_required
@require_school_access("exams", "exam_id")
def cheat_analysis(exam_id):
    """Cheat pattern detection dashboard."""
    return render_template("teacher/cheat_analysis.html", exam_id=exam_id)


@teacher_bp.route("/api/exams/<exam_id>/cheat-data")
@teacher_or_admin_required
@require_school_access("exams", "exam_id")
def cheat_analysis_data(exam_id):
    """API: analyze submissions for cheating patterns."""
    supabase = get_supabase()
    exam = supabase.table("exams").select("title,question_types,answer_key,total_questions").eq("id", exam_id).single().execute().data or {}

    answer_key = exam.get("answer_key") or {}
    if isinstance(answer_key, str):
        try: answer_key = json.loads(answer_key)
        except: answer_key = {}
    qtypes = exam.get("question_types") or {}
    if isinstance(qtypes, str):
        try: qtypes = json.loads(qtypes)
        except: qtypes = {}

    subs = supabase.table("submissions") \
        .select("id,student_id,answers,submitted_at,created_at,profiles(full_name)") \
        .eq("exam_id", exam_id) \
        .in_("status", ["submitted", "graded", "published"]) \
        .execute().data or []

    parsed = []
    for s in subs:
        answers = s.get("answers") or {}
        if isinstance(answers, str):
            try: answers = json.loads(answers)
            except: answers = {}
        profile = s.get("profiles") or {}
        parsed.append({
            "id": s["id"],
            "student_id": s["student_id"],
            "name": profile.get("full_name", s["student_id"][:12]),
            "answers": answers,
            "submitted_at": s.get("submitted_at") or s.get("created_at") or "",
        })

    # 1. Identical Wrong Answer Detection
    total_q = exam.get("total_questions", 0)
    wrong_answers = {}
    for p in parsed:
        ans = p["answers"]
        wrong_pattern = []
        for qi in range(total_q):
            qi_str = str(qi)
            stu_ans = ans.get(qi_str, "")
            if isinstance(stu_ans, dict):
                stu_ans = stu_ans.get("answer", "")
            key = answer_key.get(qi_str, "")
            if key and stu_ans and stu_ans != key:
                wrong_pattern.append(f"{qi}:{stu_ans}")
        if wrong_pattern:
            pattern = "|".join(wrong_pattern)
            if pattern not in wrong_answers:
                wrong_answers[pattern] = []
            wrong_answers[pattern].append(p["name"])

    identical_groups = [{"students": v, "count": len(v), "pattern": k[:100]}
                        for k, v in wrong_answers.items() if len(v) >= 2]
    identical_groups.sort(key=lambda x: x["count"], reverse=True)

    # 2. Submission Timing Cluster
    from collections import defaultdict
    time_clusters = []
    timestamps = [(p["name"], p["submitted_at"]) for p in parsed if p.get("submitted_at")]
    import datetime
    from datetime import timezone
    for i, (n1, t1) in enumerate(timestamps):
        cluster = [n1]
        for j, (n2, t2) in enumerate(timestamps):
            if i != j and t1 and t2:
                try:
                    dt1 = datetime.datetime.fromisoformat(t1.replace("Z", "+00:00").split(".")[0])
                    dt2 = datetime.datetime.fromisoformat(t2.replace("Z", "+00:00").split(".")[0])
                    diff = abs((dt1 - dt2).total_seconds())
                    if diff < 3:
                        cluster.append(n2)
                except: pass
        if len(cluster) >= 3:
            cluster.sort()
            key = ",".join(cluster)
            if not any(key == c.get("key") for c in time_clusters):
                time_clusters.append({"key": key, "students": list(set(cluster)), "count": len(set(cluster))})

    time_clusters.sort(key=lambda x: x["count"], reverse=True)

    return jsonify({
        "exam_title": exam.get("title", ""),
        "identical_groups": identical_groups[:10],
        "time_clusters": time_clusters[:10],
        "total_students": len(parsed),
    })


@teacher_bp.route("/exams/<exam_id>/accreditation-report")
@teacher_or_admin_required
@require_school_access("exams", "exam_id")
def accreditation_report(exam_id):
    """Generate school accreditation report as PDF."""
    supabase = get_supabase()
    exam = supabase.table("exams").select("title,subject,teacher_id,total_questions,passing_score").eq("id", exam_id).single().execute().data or {}
    subs = supabase.table("submissions").select("score,final_score,status,student_id,profiles(full_name)").eq("exam_id", exam_id).in_("status", ["graded", "published"]).execute().data or []

    from app.services.export_service import _wrap_text
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    import io

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    story = []

    # Title
    story.append(Paragraph(f"Laporan Hasil Ujian", styles["Title"]))
    story.append(Paragraph(f"{exam.get('title', '')} - {exam.get('subject', '')}", styles["Heading2"]))
    story.append(Spacer(1, 12))

    # Stats
    scores = [float(s.get("final_score") or s.get("score") or 0) for s in subs]
    avg = sum(scores) / len(scores) if scores else 0
    passed = sum(1 for s in scores if s >= (exam.get("passing_score", 70)))
    story.append(Paragraph(f"Jumlah Siswa: {len(subs)}", styles["Normal"]))
    story.append(Paragraph(f"Rata-rata: {avg:.1f}", styles["Normal"]))
    story.append(Paragraph(f"KKM: {exam.get('passing_score', 70)}", styles["Normal"]))
    story.append(Paragraph(f"Lulus: {passed}/{len(subs)} ({passed*100//len(subs) if subs else 0}%)", styles["Normal"]))
    story.append(Spacer(1, 20))

    # Score distribution table
    bins = {"0-39": 0, "40-59": 0, "60-79": 0, "80-100": 0}
    for s in scores:
        if s < 40: bins["0-39"] += 1
        elif s < 60: bins["40-59"] += 1
        elif s < 80: bins["60-79"] += 1
        else: bins["80-100"] += 1

    dist_data = [["Rentang Nilai", "Jumlah Siswa"]]
    for k, v in bins.items():
        dist_data.append([k, str(v)])
    dist_data.append(["Total", str(len(subs))])

    t = Table(dist_data, colWidths=[150, 100])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3b82f6")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    story.append(t)
    story.append(Spacer(1, 20))

    # Student list
    story.append(Paragraph("Daftar Nilai Siswa:", styles["Heading3"]))
    student_data = [["No", "Nama", "Nilai", "Status"]]
    for i, s in enumerate(subs, 1):
        profile = s.get("profiles") or {}
        name = profile.get("full_name", s["student_id"][:12])
        score = float(s.get("final_score") or s.get("score") or 0)
        passed_txt = "Lulus" if score >= (exam.get("passing_score", 70)) else "Remedial"
        student_data.append([str(i), name, f"{score:.0f}", passed_txt])

    t2 = Table(student_data, colWidths=[30, 200, 60, 80])
    t2.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3b82f6")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (2, 0), (3, -1), "CENTER"),
    ]))
    story.append(t2)
    story.append(Spacer(1, 12))

    story.append(Paragraph(f"Dicetak: {datetime.now(timezone.utc).strftime('%d %B %Y')}", styles["Normal"]))

    doc.build(story)
    buf.seek(0)
    return send_file(buf, mimetype="application/pdf", as_attachment=True,
                     download_name=f"Laporan_{exam.get('title', 'Ujian')[:30]}.pdf")
