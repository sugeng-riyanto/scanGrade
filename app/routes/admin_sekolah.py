import io
import json
import re
import secrets
import string
from datetime import datetime, timezone, date

from flask import Blueprint, render_template, g, request, jsonify, redirect, flash, send_file, current_app
from openpyxl import load_workbook, Workbook
from app.utils.auth import admin_sekolah_required, get_supabase, subscription_write_required
from app.services.audit_service import log_activity, log_create, log_update, log_delete

def _gen_password(length=12) -> str:
    chars = string.ascii_letters + string.digits + "!@#$%^&*"
    return "".join(secrets.choice(chars) for _ in range(length))

admin_sekolah_bp = Blueprint("admin_sekolah", __name__)


def _generate_email(full_name: str, domain: str) -> str:
    """Generate email from full name: Budi Santoso → budi.santoso@domain"""
    if not full_name or not full_name.strip():
        return f"user.{_gen_password(6)}@{domain or 'school.local'}"
    parts = full_name.strip().lower().split()
    if len(parts) == 1:
        return f"{parts[0]}@{domain or 'school.local'}"
    elif len(parts) == 2:
        return f"{parts[0]}.{parts[1]}@{domain or 'school.local'}"
    else:
        # first.middle_initial.last
        first = parts[0]
        middle = parts[1][0] if len(parts[1]) > 0 else ""
        last = parts[-1]
        return f"{first}.{middle}.{last}@{domain or 'school.local'}"


def _get_email_domain(sid) -> str:
    """Get custom email domain for a school."""
    try:
        from flask import current_app
        supabase = current_app.extensions["supabase"]
        school = supabase.table("schools").select("email_domain").eq("id", sid).single().execute().data or {}
        return (school.get("email_domain") or "").strip() or "school.local"
    except:
        return "school.local"


def _school_id() -> str | None:
    sid = g.get("user_school_id")
    if not sid or sid == "None":
        return None
    return sid





# ─── DASHBOARD ───────────────────────────────────────

@admin_sekolah_bp.route("/dashboard")
@admin_sekolah_required
def dashboard():
    sid = _school_id()
    if not sid:
        flash("Sekolah belum terdaftar. Hubungi Super Admin.", "error")
        return redirect("/auth/login")

    supabase = get_supabase()

    school = supabase.table("schools").select("*").eq("id", sid).single().execute().data or {}
    students = supabase.table("students").select("id", count="exact").eq("school_id", sid).eq("status", "active").execute()
    teachers = supabase.table("teachers").select("id", count="exact").eq("school_id", sid).execute()
    classes = supabase.table("classes").select("id", count="exact").eq("school_id", sid).execute()
    years = supabase.table("school_years").select("*").eq("school_id", sid).eq("is_active", True).execute().data or []
    active_year = years[0] if years else None

    ctx = {
        "school": school,
        "student_count": students.count or 0,
        "teacher_count": teachers.count or 0,
        "class_count": classes.count or 0,
        "active_year": active_year,
    }
    return render_template("admin_sekolah/dashboard.html", **ctx)


# ─── SCHOOL PROFILE ──────────────────────────────────

@admin_sekolah_bp.route("/profile", methods=["GET", "POST"])
@admin_sekolah_required
def profile():
    sid = _school_id()
    supabase = get_supabase()

    if request.method == "POST":
        # Handle logo upload
        logo_file = request.files.get("logo")
        logo_url = None
        if logo_file and logo_file.filename:
            import uuid, os
            ext = logo_file.filename.rsplit(".", 1)[-1].lower() if "." in logo_file.filename else "png"
            logo_name = f"logo_{sid[:8]}.{ext}"
            logo_dir = os.path.join(current_app.root_path, "static", "uploads", "logos")
            os.makedirs(logo_dir, exist_ok=True)
            logo_path = os.path.join(logo_dir, logo_name)
            logo_file.save(logo_path)
            logo_url = f"/static/uploads/logos/{logo_name}"

        data = {
            "name": request.form.get("name", "").strip(),
            "npsn": request.form.get("npsn", "").strip(),
            "address": request.form.get("address", "").strip(),
            "province": request.form.get("province", "").strip(),
            "city": request.form.get("city", "").strip(),
            "district": request.form.get("district", "").strip(),
            "postal_code": request.form.get("postal_code", "").strip(),
            "phone": request.form.get("phone", "").strip(),
            "email": request.form.get("email", "").strip(),
            "website": request.form.get("website", "").strip(),
            "principal_name": request.form.get("principal_name", "").strip(),
            "principal_nip": request.form.get("principal_nip", "").strip(),
            "tz_offset": int(request.form.get("tz_offset", 7)),
            "email_domain": request.form.get("email_domain", "").strip(),
        }
        if logo_url:
            data["logo_url"] = logo_url
        try:
            supabase.table("schools").update(data).eq("id", sid).execute()
            log_activity("update", "school", sid, new_data=data, user_id=g.user_id)
            flash("Profil sekolah berhasil diperbarui", "success")
        except Exception as e:
            flash(f"Gagal: {e}", "error")
        return redirect("/admin-sekolah/profile")

    school = supabase.table("schools").select("*").eq("id", sid).single().execute().data or {}
    return render_template("admin_sekolah/profile.html", school=school)


# ─── IMPORT EXCEL ────────────────────────────────────

@admin_sekolah_bp.route("/generate-email")
@admin_sekolah_required
def generate_email_preview():
    """Preview generated email + password for a given name."""
    name = request.args.get("name", "").strip()
    sid = _school_id()
    domain = _get_email_domain(sid)
    email = _generate_email(name, domain) if name else ""
    pw = _gen_password() if name else ""
    return jsonify({"email": email, "password": pw, "domain": domain})


@admin_sekolah_bp.route("/download-template/murid")
@admin_sekolah_required
def download_template_murid():
    """Download XLSX template for students with generated passwords."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    sid = _school_id()
    domain = _get_email_domain(sid)
    school = get_supabase().table("schools").select("name, npsn").eq("id", sid).single().execute().data or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Murid"
    # Headers
    headers = ["NPSN", "Tahun Ajaran", "NISN", "Email", "Nama Lengkap", "Kelas", "Password"]
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = Alignment(horizontal="center")
    # Example rows with generated emails + passwords
    for i, (nama, nisn) in enumerate([("Ahmad Budiman", "1234567801"), ("Citra Dewi", "1234567802")], 2):
        email = _generate_email(nama, domain)
        pw = _gen_password()
        ws.cell(row=i, column=1, value=school.get("npsn", ""))
        ws.cell(row=i, column=2, value="2025/2026")
        ws.cell(row=i, column=3, value=nisn)
        ws.cell(row=i, column=4, value=email)
        ws.cell(row=i, column=5, value=nama)
        ws.cell(row=i, column=6, value="VII-A")
        ws.cell(row=i, column=7, value=pw)
    for col in range(1, 8):
        ws.column_dimensions[chr(64+col)].width = 22
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="template_murid.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@admin_sekolah_bp.route("/download-template/guru")
@admin_sekolah_required
def download_template_guru():
    """Download XLSX template for teachers with generated passwords."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    sid = _school_id()
    domain = _get_email_domain(sid)
    supabase = get_supabase()
    school = supabase.table("schools").select("name, npsn").eq("id", sid).single().execute().data or {}
    subjects = supabase.table("subjects").select("name").eq("school_id", sid).execute().data or []
    subj_names = [s["name"] for s in subjects]
    wb = Workbook()
    ws = wb.active
    ws.title = "Guru"
    headers = ["NPSN", "Tahun Ajaran", "NIP", "Email", "Nama Lengkap", "Mapel 1", "Mapel 2", "Mapel 3", "Password"]
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = Alignment(horizontal="center")
    for i, (nama, nip, mapels) in enumerate([("Budi Santoso", "19870101", subj_names[:2]), ("Siti Rahma", "19900202", subj_names[:1])], 2):
        email = _generate_email(nama, domain)
        pw = _gen_password()
        ws.cell(row=i, column=1, value=school.get("npsn", ""))
        ws.cell(row=i, column=2, value="2025/2026")
        ws.cell(row=i, column=3, value=nip)
        ws.cell(row=i, column=4, value=email)
        ws.cell(row=i, column=5, value=nama)
        for j, mn in enumerate(mapels):
            ws.cell(row=i, column=6+j, value=mn)
        ws.cell(row=i, column=9, value=pw)
    for col in range(1, 10):
        ws.column_dimensions[chr(64+col)].width = 20
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="template_guru.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@admin_sekolah_bp.route("/import", methods=["GET", "POST"])
@admin_sekolah_required
def import_excel():
    if request.method == "GET":
        return render_template("admin_sekolah/import.html")

    file = request.files.get("file")
    if not file:
        flash("File tidak ditemukan", "error")
        return redirect("/admin-sekolah/import")

    sid = _school_id()
    supabase = get_supabase()

    try:
        wb = load_workbook(filename=io.BytesIO(file.read()))
    except Exception as e:
        flash(f"Gagal membaca file: {e}", "error")
        return redirect("/admin-sekolah/import")

    results = {"students": 0, "teachers": 0, "subjects": 0, "errors": []}
    sheet_names_lower = {s.lower(): s for s in wb.sheetnames}

    # ── Sheet: Murid / Students ──
    for key in ("murid", "siswa", "students", "student"):
        if key in sheet_names_lower:
            ws = wb[sheet_names_lower[key]]
            _import_students(ws, sid, supabase, results)
            break

    # ── Sheet: Guru / Teachers ──
    for key in ("guru", "teachers", "teacher"):
        if key in sheet_names_lower:
            ws = wb[sheet_names_lower[key]]
            _import_teachers(ws, sid, supabase, results)
            break

    # ── Sheet: Mata Pelajaran / Subjects ──
    for key in ("mata pelajaran", "pelajaran", "subjects", "subject", "mapel"):
        if key in sheet_names_lower:
            ws = wb[sheet_names_lower[key]]
            _import_subjects(ws, sid, supabase, results)
            break

    log_activity("import", "school", sid, new_data={"students": results["students"], "teachers": results["teachers"], "subjects": results["subjects"], "errors": len(results["errors"])}, user_id=g.user_id)
    flash(
        f"Impor selesai: {results['students']} murid, {results['teachers']} guru, {results['subjects']} mapel. "
        f"{len(results['errors'])} error.",
        "success" if not results["errors"] else "warning",
    )
    return redirect("/admin-sekolah/import")


def _import_students(ws, sid, supabase, results):
    classes_cache = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or not row[0]:
            continue
        try:
            cols = [str(c or "").strip() for c in row]
            # Support both old format (NISN, Nama, Kelas, Level, Email)
            # and new template format (NPSN, Thn Ajaran, NISN, Email, Nama, Kelas, Password)
            if len(cols) >= 7 and cols[2].isdigit() and len(cols[2]) >= 8:
                # New template: NPSN(0), ThnAjaran(1), NISN(2), Email(3), Nama(4), Kelas(5), Pw(6)
                nisn = cols[2]; nama = cols[4]; kelas = cols[5]; email = cols[3]
            else:
                # Old format: NISN(0), Nama(1), Kelas(2), Level(3), Email(4)
                nisn = cols[0]; nama = cols[1]; kelas = cols[2] if len(cols) > 2 else ""
                email = cols[4] if len(cols) > 4 else ""

            if not nisn or not nama:
                continue

            # Resolve class_id
            class_id = None
            if kelas:
                if kelas not in classes_cache:
                    c = supabase.table("classes").select("id").eq("school_id", sid).eq("name", kelas).maybe_single().execute()
                    classes_cache[kelas] = c.data["id"] if c.data else None
                class_id = classes_cache.get(kelas)

            user_email = email or _generate_email(nama, _get_email_domain(sid))
            user_pw = _gen_password()

            res = supabase.auth.admin.create_user({
                "email": user_email,
                "password": user_pw,
                "user_metadata": {"role": "murid", "full_name": nama},
                "email_confirm": True,
            })
            uid = res.user.id
            supabase.table("profiles").upsert({
                "id": uid, "full_name": nama, "phone": "", "role": "murid",
                "nisn": nisn, "class_id": class_id, "status": "active", "school_id": sid,
            }).execute()
            supabase.table("students").upsert({
                "id": uid, "school_id": sid, "class_id": class_id, "nisn": nisn,
                "status": "active",
            }).execute()
            results["students"] += 1
        except Exception as e:
            results["errors"].append(f"Baris {row_idx}: {e}")


def _import_teachers(ws, sid, supabase, results):
    subjects_cache = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or not row[0]:
            continue
        try:
            cols = [str(c or "").strip() for c in row]
            # Support both old format and new template format
            if len(cols) >= 9 and cols[2].isdigit() and len(cols[2]) >= 5:
                # New template: NPSN(0), ThnAjaran(1), NIP(2), Email(3), Nama(4), Mapel1(5), Mapel2(6), Mapel3(7), Pw(8)
                nip = cols[2]; nama = cols[4]; email = cols[3]
                mapels = [cols[i] for i in range(5, min(8, len(cols))) if cols[i]]
                hp = ""
            else:
                # Old format: NIP(0), Nama(1), Mapel(2), Email(3), HP(4)
                nip = cols[0]; nama = cols[1]; mapels = [cols[2]] if len(cols) > 2 and cols[2] else []
                email = cols[3] if len(cols) > 3 else ""; hp = cols[4] if len(cols) > 4 else ""

            if not nip or not nama:
                continue

            # Resolve subject_id (use first mapel from the list)
            subject_id = None
            for mn in mapels:
                if mn and mn not in subjects_cache:
                    s = supabase.table("subjects").select("id").eq("school_id", sid).eq("name", mn).maybe_single().execute()
                    subjects_cache[mn] = s.data["id"] if s.data else None
                if mn and subjects_cache.get(mn):
                    subject_id = subjects_cache[mn]
                    break

            user_email = email or _generate_email(nama, _get_email_domain(sid))
            user_pw = _gen_password()

            res = supabase.auth.admin.create_user({
                "email": user_email,
                "password": user_pw,
                "user_metadata": {"role": "guru", "full_name": nama},
                "email_confirm": True,
            })
            uid = res.user.id
            supabase.table("profiles").upsert({
                "id": uid, "full_name": nama, "phone": hp, "role": "guru",
                "status": "active", "school_id": sid,
            }).execute()
            supabase.table("teachers").upsert({
                "id": uid, "school_id": sid, "employee_id": nip,
                "subject_id": subject_id,
            }).execute()
            results["teachers"] += 1
        except Exception as e:
            results["errors"].append(f"Baris {row_idx}: {e}")


def _import_subjects(ws, sid, supabase, results):
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or not row[0]:
            continue
        try:
            name = str(row[0] or "").strip()
            if not name:
                continue
            code = str(row[1] or "").strip() if len(row) > 1 else ""
            supabase.table("subjects").upsert({
                "school_id": sid, "name": name, "code": code,
            }).execute()
            results["subjects"] += 1
        except Exception as e:
            results["errors"].append(f"Baris {row_idx}: {e}")


# ─── EXPORT EXCEL ────────────────────────────────────

@admin_sekolah_bp.route("/export")
@admin_sekolah_required
def export_excel():
    sid = _school_id()
    supabase = get_supabase()
    school = supabase.table("schools").select("npsn, name, email_domain").eq("id", sid).single().execute().data or {}
    npsn = school.get("npsn", "")
    years = supabase.table("school_years").select("name").eq("school_id", sid).eq("is_active", True).execute().data or []
    academic_year = years[0]["name"] if years else "2025/2026"
    domain = _get_email_domain(sid)
    _email_map = {}
    try:
        for u in supabase.auth.admin.list_users():
            _email_map[u.id] = u.email
    except:
        pass

    wb = Workbook()
    # Students sheet (sama format dengan template download)
    ws1 = wb.active
    ws1.title = "Murid"
    ws1.append(["NPSN", "Tahun Ajaran", "NISN", "Email", "Nama Lengkap", "Kelas"])
    students = supabase.table("students").select("*, profiles!inner(id, full_name), classes(name)").eq("school_id", sid).execute().data or []
    for s in students:
        prof = s.get("profiles") or {}
        uid = prof.get("id", "")
        ws1.append([npsn, academic_year, s.get("nisn", ""), _email_map.get(uid, ""),
                     prof.get("full_name", ""), (s.get("classes") or {}).get("name", "")])

    # Teachers sheet (sama format dengan template download)
    ws2 = wb.create_sheet("Guru")
    ws2.append(["NPSN", "Tahun Ajaran", "NIP", "Email", "Nama Lengkap", "Mapel"])
    teachers = supabase.table("teachers").select("*, profiles!inner(id, full_name), subjects(name)").eq("school_id", sid).execute().data or []
    for t in teachers:
        prof = t.get("profiles") or {}
        uid = prof.get("id", "")
        ws2.append([npsn, academic_year, t.get("employee_id", ""), _email_map.get(uid, ""),
                     prof.get("full_name", ""), (t.get("subjects") or {}).get("name", "")])

    # Subjects sheet
    ws3 = wb.create_sheet("Mata Pelajaran")
    ws3.append(["Nama Mata Pelajaran", "Kode"])
    subjects = supabase.table("subjects").select("*").eq("school_id", sid).execute().data or []
    for s in subjects:
        ws3.append([s.get("name", ""), s.get("code", "")])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="data_sekolah.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ─── SCHOOL YEARS ────────────────────────────────────

@admin_sekolah_bp.route("/school-years", methods=["GET", "POST"])
@admin_sekolah_required
def school_years():
    sid = _school_id()
    supabase = get_supabase()

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        start_date = request.form.get("start_date", "").strip()
        end_date = request.form.get("end_date", "").strip()
        is_active = request.form.get("is_active") == "1"

        if not name or not start_date or not end_date:
            flash("Nama, tanggal mulai, dan tanggal berakhir wajib diisi", "error")
            return redirect("/admin-sekolah/school-years")

        try:
            if is_active:
                supabase.table("school_years").update({"is_active": False}).eq("school_id", sid).execute()
            res = supabase.table("school_years").insert({
                "school_id": sid, "name": name,
                "start_date": start_date, "end_date": end_date,
                "is_active": is_active,
            }).execute()
            new_id = res.data[0]["id"] if res.data else None
            log_activity("create", "school_year", new_id, new_data={"name": name, "start_date": start_date, "end_date": end_date, "is_active": is_active}, user_id=g.user_id)
            flash("Tahun ajaran berhasil ditambahkan", "success")
        except Exception as e:
            flash(f"Gagal: {e}", "error")
        return redirect("/admin-sekolah/school-years")

    years = supabase.table("school_years").select("*").eq("school_id", sid).order("name", desc=True).execute().data or []
    return render_template("admin_sekolah/school_years.html", years=years)


@admin_sekolah_bp.route("/school-years/<year_id>/toggle", methods=["POST"])
@admin_sekolah_required
def toggle_school_year(year_id):
    sid = _school_id()
    supabase = get_supabase()
    supabase.table("school_years").update({"is_active": False}).eq("school_id", sid).execute()
    supabase.table("school_years").update({"is_active": True}).eq("id", year_id).execute()
    return jsonify({"success": True})


@admin_sekolah_bp.route("/school-years/<year_id>/delete", methods=["POST"])
@admin_sekolah_required
def delete_school_year(year_id):
    supabase = get_supabase()
    try:
        supabase.table("school_years").delete().eq("id", year_id).execute()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


# ─── CLASSES ─────────────────────────────────────────

@admin_sekolah_bp.route("/classes")
@admin_sekolah_required
def classes():
    sid = _school_id()
    supabase = get_supabase()
    classes_list = supabase.table("classes").select("*, profiles!classes_teacher_id_fkey(full_name)").eq("school_id", sid).order("name").execute().data or []
    for c in classes_list:
        c["wali_kelas"] = (c.get("profiles") or {}).get("full_name")
    teachers = supabase.table("profiles").select("id, full_name").eq("role", "guru").eq("school_id", sid).execute().data or []
    years = supabase.table("school_years").select("*").eq("school_id", sid).order("name", desc=True).execute().data or []
    return render_template("admin_sekolah/classes.html", classes=classes_list, teachers=teachers, years=years)


@admin_sekolah_bp.route("/classes/create", methods=["POST"])
@admin_sekolah_required
def create_class():
    sid = _school_id()
    supabase = get_supabase()
    name = request.form.get("name", "").strip()
    grade_level = request.form.get("grade_level", "").strip()
    wali_id = request.form.get("wali_kelas_id") or None
    year_id = request.form.get("school_year_id") or None
    if not name:
        flash("Nama kelas wajib diisi", "error")
        return redirect("/admin-sekolah/classes")
    # Check duplicate across all roles
    dup = supabase.table("classes").select("id").eq("school_id", sid).eq("name", name).limit(1).execute()
    if dup.data:
        flash(f"Kelas '{name}' sudah ada", "error")
        return redirect("/admin-sekolah/classes")
    try:
        res = supabase.table("classes").insert({
            "name": name, "grade_level": grade_level, "school_id": sid,
            "teacher_id": wali_id, "school_year_id": year_id, "created_by": g.user_id,
        }).execute()
        cid = res.data[0]["id"] if res.data else None
        log_activity("create", "class", cid, new_data={"name": name, "grade_level": grade_level}, user_id=g.user_id)
        flash("Kelas berhasil ditambahkan", "success")
    except Exception as e:
        flash(f"Gagal: {e}", "error")
    return redirect("/admin-sekolah/classes")


@admin_sekolah_bp.route("/classes/<class_id>/edit", methods=["POST"])
@admin_sekolah_required
def edit_class(class_id):
    supabase = get_supabase()
    data = {}
    for key in ("name", "grade_level"):
        val = request.form.get(key)
        if val is not None:
            data[key] = val.strip()
    wali = request.form.get("wali_kelas_id")
    data["teacher_id"] = wali if wali else None
    year_id = request.form.get("school_year_id")
    data["school_year_id"] = year_id if year_id else None
    try:
        supabase.table("classes").update(data).eq("id", class_id).execute()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@admin_sekolah_bp.route("/classes/<class_id>/delete", methods=["POST"])
@admin_sekolah_required
def delete_class(class_id):
    supabase = get_supabase()
    try:
        supabase.table("students").update({"class_id": None}).eq("class_id", class_id).execute()
        supabase.table("classes").delete().eq("id", class_id).execute()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


# ─── SUBJECTS CRUD ────────────────────────────────────

@admin_sekolah_bp.route("/subjects")
@admin_sekolah_required
def admin_subjects():
    sid = _school_id()
    supabase = get_supabase()
    sort = request.args.get("sort", "asc")
    q = request.args.get("q", "")
    data = supabase.table("subjects").select("*").eq("school_id", sid).order("name", desc=(sort == "desc")).execute().data or []
    if q:
        data = [s for s in data if q.lower() in s.get("name", "").lower()]
    return render_template("admin_sekolah/subjects.html", subjects=data, sort=sort, q=q)


@admin_sekolah_bp.route("/subjects/create", methods=["POST"])
@admin_sekolah_required
def admin_subject_create():
    sid = _school_id()
    supabase = get_supabase()
    name = request.form.get("name", "").strip()
    if not name:
        flash("Nama mapel wajib diisi", "error")
        return redirect("/admin-sekolah/subjects")
    # Check duplicate across all roles
    dup = supabase.table("subjects").select("id").eq("school_id", sid).eq("name", name).limit(1).execute()
    if dup.data:
        flash(f"Mapel '{name}' sudah ada", "error")
        return redirect("/admin-sekolah/subjects")
    try:
        supabase.table("subjects").insert({"school_id": sid, "name": name, "is_active": True, "created_by": g.user_id}).execute()
        log_activity("create", "subject", name, new_data={"name": name}, user_id=g.user_id)
        flash("Mapel berhasil ditambahkan", "success")
    except Exception as e:
        flash(f"Gagal: {e}", "error")
    return redirect("/admin-sekolah/subjects")


@admin_sekolah_bp.route("/subjects/<subject_id>/delete", methods=["POST"])
@admin_sekolah_required
def admin_subject_delete(subject_id):
    supabase = get_supabase()
    try:
        supabase.table("teacher_assignments").delete().eq("subject_id", subject_id).execute()
        supabase.table("subjects").delete().eq("id", subject_id).execute()
        log_activity("delete", "subject", str(subject_id), user_id=g.user_id)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


# ─── PROMOTE (Naik Kelas) ────────────────────────────

@admin_sekolah_bp.route("/promote", methods=["GET", "POST"])
@admin_sekolah_required
def promote():
    sid = _school_id()
    supabase = get_supabase()

    if request.method == "POST":
        source_class_id = request.form.get("source_class_id")
        target_class_id = request.form.get("target_class_id")
        create_new = request.form.get("create_new") == "1"

        if not source_class_id:
            flash("Pilih kelas asal", "error")
            return redirect("/admin-sekolah/promote")

        # Get source class info for level inference
        src = supabase.table("classes").select("*").eq("id", source_class_id).single().execute().data

        if create_new:
            new_name = request.form.get("new_class_name", "").strip()
            new_level = request.form.get("new_grade_level", "").strip()
            year_id = request.form.get("school_year_id") or None
            if not new_name:
                flash("Nama kelas baru wajib diisi", "error")
                return redirect("/admin-sekolah/promote")
            res = supabase.table("classes").insert({
                "name": new_name, "grade_level": new_level or src.get("grade_level", ""),
                "school_id": sid, "school_year_id": year_id,
            }).execute()
            target_class_id = res.data[0]["id"] if res.data else None

        if not target_class_id:
            flash("Pilih atau buat kelas tujuan", "error")
            return redirect("/admin-sekolah/promote")

        # Move students
        students = supabase.table("students").select("id").eq("class_id", source_class_id).eq("status", "active").execute().data or []
        moved = 0
        for s in students:
            try:
                supabase.table("students").update({"class_id": target_class_id}).eq("id", s["id"]).execute()
                supabase.table("profiles").update({"class_id": target_class_id}).eq("id", s["id"]).execute()
                moved += 1
            except Exception:
                pass
        log_activity("promote", "class", source_class_id, new_data={"target_class_id": target_class_id, "moved": moved, "school_year_id": request.form.get("school_year_id")}, user_id=g.user_id)
        flash(f"{moved} murid berhasil dipindahkan ke kelas tujuan", "success")
        return redirect("/admin-sekolah/promote")

    classes_list = supabase.table("classes").select("*").eq("school_id", sid).order("name").execute().data or []
    teachers = supabase.table("profiles").select("id, full_name").eq("role", "guru").eq("school_id", sid).execute().data or []
    years = supabase.table("school_years").select("*").eq("school_id", sid).order("name", desc=True).execute().data or []
    return render_template("admin_sekolah/promote.html", classes=classes_list, teachers=teachers, years=years)


# ─── TEACHERS CRUD ───────────────────────────────────

@admin_sekolah_bp.route("/teachers")
@admin_sekolah_required
def teachers():
    sid = _school_id()
    supabase = get_supabase()
    q = request.args.get("q", "").strip()
    sort_by = request.args.get("sort", "name")
    sort_dir = request.args.get("dir", "asc")

    subjects = supabase.table("subjects").select("*").eq("school_id", sid).order("name").execute().data or []
    teachers_raw = supabase.table("teachers").select(
        "*, profiles!inner(id, full_name, phone), subjects(name)"
    ).eq("school_id", sid).execute().data or []

    # Fetch auth emails once for all users
    _email_map = {}
    try:
        for u in supabase.auth.admin.list_users():
            _email_map[u.id] = u.email
    except:
        pass

    teachers_list = []
    for t in teachers_raw:
        prof = t.get("profiles") or {}
        subj = t.get("subjects") or {}
        uid = t["id"]
        teachers_list.append({
            "id": uid,
            "name": prof.get("full_name", "-"),
            "employee_number": t.get("employee_id", ""),
            "email": _email_map.get(uid, ""),
            "subject_name": subj.get("name", "-"),
            "subject_id": t.get("subject_id"),
            "phone": prof.get("phone", ""),
            "employee_id": t.get("employee_id", ""),
        })
    if q:
        ql = q.lower()
        teachers_list = [t for t in teachers_list if ql in t["name"].lower() or ql in t["employee_number"].lower()]
    reverse = sort_dir == "desc"
    if sort_by == "name":
        teachers_list.sort(key=lambda t: t["name"].lower(), reverse=reverse)
    elif sort_by == "employee_number":
        teachers_list.sort(key=lambda t: t["employee_number"], reverse=reverse)
    return render_template("admin_sekolah/teachers.html", teachers=teachers_list, subjects=subjects, q=q, sort_by=sort_by, sort_dir=sort_dir)


@admin_sekolah_bp.route("/teachers/create", methods=["POST"])
@subscription_write_required
@admin_sekolah_required
def create_teacher():
    sid = _school_id()
    supabase = get_supabase()
    nip = request.form.get("employee_number", request.form.get("employee_id", "")).strip()
    nama = request.form.get("name", request.form.get("full_name", "")).strip()
    email = request.form.get("email", "").strip().lower()
    hp = request.form.get("phone", "").strip()
    subject_id = request.form.get("subject_id") or None
    password = request.form.get("password", "").strip() or _gen_password()

    if not nama:
        flash("Nama guru wajib diisi", "error")
        return redirect("/admin-sekolah/teachers")

    try:
        user_email = email or _generate_email(nama, _get_email_domain(sid))
        res = supabase.auth.admin.create_user({
            "email": user_email, "password": password,
            "user_metadata": {"role": "guru", "full_name": nama},
            "email_confirm": True,
        })
        uid = res.user.id
        supabase.table("profiles").upsert({
            "id": uid, "full_name": nama, "phone": hp, "role": "guru",
            "status": "active", "school_id": sid,
        }).execute()
        supabase.table("teachers").upsert({
            "id": uid, "school_id": sid, "employee_id": nip, "subject_id": subject_id,
        }).execute()
        log_activity("create", "teacher", uid, new_data={"full_name": nama, "employee_id": nip}, user_id=g.user_id)
        flash(f"Guru berhasil ditambahkan. Email: {user_email}, Password: {password}", "success")
    except Exception as e:
        flash(f"Gagal: {e}", "error")
    return redirect("/admin-sekolah/teachers")


@admin_sekolah_bp.route("/teachers/<teacher_id>/edit", methods=["POST"])
@subscription_write_required
@admin_sekolah_required
def edit_teacher(teacher_id):
    supabase = get_supabase()
    data = {}
    emp_id = request.form.get("employee_number", request.form.get("employee_id", ""))
    if emp_id:
        data["employee_id"] = emp_id.strip()
    subj = request.form.get("subject_id")
    data["subject_id"] = subj if subj else None

    profile_data = {}
    nama = request.form.get("name", request.form.get("full_name", ""))
    if nama:
        profile_data["full_name"] = nama.strip()
    hp = request.form.get("phone", "")
    if hp:
        profile_data["phone"] = hp.strip()
    # Email is stored in auth.users, not profiles table — cannot update via this API

    try:
        if data:
            supabase.table("teachers").update(data).eq("id", teacher_id).execute()
        if profile_data:
            supabase.table("profiles").update(profile_data).eq("id", teacher_id).execute()
        log_activity("update", "teacher", teacher_id, new_data={**data, **profile_data}, user_id=g.user_id)
        flash("Guru berhasil diperbarui", "success")
        return redirect("/admin-sekolah/teachers")
    except Exception as e:
        flash(f"Gagal: {e}", "error")
        return redirect("/admin-sekolah/teachers")


@admin_sekolah_bp.route("/teachers/<teacher_id>/delete", methods=["POST"])
@subscription_write_required
@admin_sekolah_required
def delete_teacher(teacher_id):
    supabase = get_supabase()
    try:
        supabase.table("teachers").delete().eq("id", teacher_id).execute()
        supabase.table("profiles").delete().eq("id", teacher_id).execute()
        supabase.auth.admin.delete_user(teacher_id)
        log_activity("delete", "teacher", teacher_id, user_id=g.user_id)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@admin_sekolah_bp.route("/teachers/<teacher_id>/reset-password", methods=["POST"])
@admin_sekolah_required
def reset_teacher_password(teacher_id):
    supabase = get_supabase()
    password = request.form.get("password", "").strip() or _gen_password()
    try:
        supabase.auth.admin.update_user_by_id(teacher_id, {"password": password})
        log_activity("reset_password", "teacher", teacher_id, user_id=g.user_id)
        return jsonify({"success": True, "password": password})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


# ─── STUDENTS CRUD ───────────────────────────────────

@admin_sekolah_bp.route("/students")
@admin_sekolah_required
def students():
    sid = _school_id()
    supabase = get_supabase()
    q = request.args.get("q", "").strip()
    sort_by = request.args.get("sort", "name")
    sort_dir = request.args.get("dir", "asc")

    classes_list = supabase.table("classes").select("*").eq("school_id", sid).order("name").execute().data or []
    students_raw = supabase.table("students").select(
        "*, profiles!inner(id, full_name, phone, nisn), classes(name)"
    ).eq("school_id", sid).order("nisn").execute().data or []

    _email_map = {}
    try:
        for u in supabase.auth.admin.list_users():
            _email_map[u.id] = u.email
    except:
        pass

    students_list = []
    for s in students_raw:
        prof = s.get("profiles") or {}
        cls = s.get("classes") or {}
        uid = s["id"]
        students_list.append({
            "id": uid,
            "nisn": s.get("nisn", "") or prof.get("nisn", ""),
            "name": prof.get("full_name", "-"),
            "class_name": cls.get("name", "-"),
            "class_id": s.get("class_id"),
            "phone": prof.get("phone", ""),
            "email": _email_map.get(uid, ""),
        })
    if q:
        ql = q.lower()
        students_list = [s for s in students_list if ql in s["name"].lower() or ql in s["nisn"]]
    reverse = sort_dir == "desc"
    if sort_by == "name":
        students_list.sort(key=lambda s: s["name"].lower(), reverse=reverse)
    elif sort_by == "nisn":
        students_list.sort(key=lambda s: s["nisn"], reverse=reverse)
    elif sort_by == "class_name":
        students_list.sort(key=lambda s: s["class_name"], reverse=reverse)
    return render_template("admin_sekolah/students.html", students=students_list, classes=classes_list, q=q, sort_by=sort_by, sort_dir=sort_dir)


@admin_sekolah_bp.route("/students/create", methods=["POST"])
@subscription_write_required
@admin_sekolah_required
def create_student():
    sid = _school_id()
    supabase = get_supabase()
    nisn = request.form.get("nisn", "").strip()
    nama = request.form.get("name", request.form.get("full_name", "")).strip()
    email = request.form.get("email", "").strip().lower()
    class_id = request.form.get("class_id") or None
    password = request.form.get("password", "").strip() or _gen_password()

    if not nama:
        flash("Nama murid wajib diisi", "error")
        return redirect("/admin-sekolah/students")

    try:
        user_email = email or _generate_email(nama, _get_email_domain(sid))
        res = supabase.auth.admin.create_user({
            "email": user_email, "password": password,
            "user_metadata": {"role": "murid", "full_name": nama},
            "email_confirm": True,
        })
        uid = res.user.id
        supabase.table("profiles").upsert({
            "id": uid, "full_name": nama, "role": "murid", "nisn": nisn,
            "class_id": class_id, "status": "active", "school_id": sid,
        }).execute()
        supabase.table("students").upsert({
            "id": uid, "school_id": sid, "class_id": class_id, "nisn": nisn,
            "status": "active",
        }).execute()
        log_activity("create", "student", uid, new_data={"full_name": nama, "nisn": nisn, "class_id": class_id}, user_id=g.user_id)
        flash(f"Murid berhasil ditambahkan. Email: {user_email}, Password: {password}", "success")
    except Exception as e:
        flash(f"Gagal: {e}", "error")
    return redirect("/admin-sekolah/students")


@admin_sekolah_bp.route("/students/<student_id>/edit", methods=["POST"])
@subscription_write_required
@admin_sekolah_required
def edit_student(student_id):
    supabase = get_supabase()
    data = {}
    nisn = request.form.get("nisn", "")
    if nisn:
        data["nisn"] = nisn.strip()
    cls = request.form.get("class_id")
    data["class_id"] = cls if cls else None

    profile_data = {}
    nama = request.form.get("name", request.form.get("full_name", ""))
    if nama:
        profile_data["full_name"] = nama.strip()

    try:
        if data:
            supabase.table("students").update(data).eq("id", student_id).execute()
            supabase.table("profiles").update({"nisn": data.get("nisn"), "class_id": data.get("class_id")}).eq("id", student_id).execute()
        if profile_data:
            supabase.table("profiles").update(profile_data).eq("id", student_id).execute()
        log_activity("update", "student", student_id, new_data={**data, **profile_data}, user_id=g.user_id)
        flash("Murid berhasil diperbarui", "success")
        return redirect("/admin-sekolah/students")
    except Exception as e:
        flash(f"Gagal: {e}", "error")
        return redirect("/admin-sekolah/students")


@admin_sekolah_bp.route("/teachers/bulk-reset-password", methods=["POST"])
@subscription_write_required
@admin_sekolah_required
def bulk_reset_teachers_password():
    sid = _school_id()
    supabase = get_supabase()
    data = request.get_json() if request.is_json else request.form
    user_ids = data.get("user_ids", [])
    if isinstance(user_ids, str):
        user_ids = json.loads(user_ids)
    if not user_ids:
        return jsonify({"error": "Tidak ada user dipilih"}), 400
    results = []
    for uid in user_ids:
        try:
            prof = supabase.table("profiles").select("school_id").eq("id", uid).single().execute().data or {}
            if prof.get("school_id") != sid:
                results.append({"id": uid, "error": "Not in school"})
                continue
            new_pw = _gen_password()
            supabase.auth.admin.update_user_by_id(uid, {"password": new_pw})
            log_activity("reset_password", "user", uid, user_id=g.user_id)
            results.append({"id": uid, "password": new_pw, "success": True})
        except Exception as e:
            results.append({"id": uid, "error": str(e)})
    return jsonify({"results": results, "total": len(results)})


@admin_sekolah_bp.route("/teachers/bulk-delete", methods=["POST"])
@subscription_write_required
@admin_sekolah_required
def bulk_delete_teachers():
    sid = _school_id()
    supabase = get_supabase()
    data = request.get_json() if request.is_json else request.form
    user_ids = data.get("user_ids", [])
    if isinstance(user_ids, str):
        user_ids = json.loads(user_ids)
    if not user_ids:
        return jsonify({"error": "Tidak ada user dipilih"}), 400
    results = []
    for uid in user_ids:
        try:
            prof = supabase.table("profiles").select("school_id").eq("id", uid).single().execute().data or {}
            if prof.get("school_id") != sid:
                continue
            supabase.table("teachers").delete().eq("id", uid).execute()
            supabase.table("profiles").delete().eq("id", uid).execute()
            supabase.auth.admin.delete_user(uid)
            results.append({"id": uid, "success": True})
        except:
            pass
    return jsonify({"results": results, "total": len(results)})


@admin_sekolah_bp.route("/students/bulk-reset-password", methods=["POST"])
@subscription_write_required
@admin_sekolah_required
def bulk_reset_students_password():
    sid = _school_id()
    supabase = get_supabase()
    data = request.get_json() if request.is_json else request.form
    user_ids = data.get("user_ids", [])
    if isinstance(user_ids, str):
        user_ids = json.loads(user_ids)
    if not user_ids:
        return jsonify({"error": "Tidak ada user dipilih"}), 400
    results = []
    for uid in user_ids:
        try:
            prof = supabase.table("profiles").select("school_id").eq("id", uid).single().execute().data or {}
            if prof.get("school_id") != sid:
                results.append({"id": uid, "error": "Not in school"})
                continue
            new_pw = _gen_password()
            supabase.auth.admin.update_user_by_id(uid, {"password": new_pw})
            log_activity("reset_password", "user", uid, user_id=g.user_id)
            results.append({"id": uid, "password": new_pw, "success": True})
        except Exception as e:
            results.append({"id": uid, "error": str(e)})
    return jsonify({"results": results, "total": len(results)})


@admin_sekolah_bp.route("/users/<user_id>/reset-password", methods=["POST"])
@admin_sekolah_required
def admin_reset_user_password(user_id):
    """Admin sekolah reset password untuk guru/murid di sekolahnya."""
    sid = _school_id()
    supabase = get_supabase()
    try:
        # Verify user belongs to this school
        prof = supabase.table("profiles").select("school_id, role").eq("id", user_id).single().execute().data or {}
        if prof.get("school_id") != sid:
            return jsonify({"error": "User tidak berada di sekolah Anda"}), 403
        new_pw = _gen_password()
        supabase.auth.admin.update_user_by_id(user_id, {"password": new_pw})
        log_activity("reset_password", "user", user_id, user_id=g.user_id)
        return jsonify({"success": True, "password": new_pw})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@admin_sekolah_bp.route("/students/<student_id>/delete", methods=["POST"])
@subscription_write_required
@admin_sekolah_required
def delete_student(student_id):
    supabase = get_supabase()
    try:
        supabase.table("students").delete().eq("id", student_id).execute()
        supabase.table("profiles").delete().eq("id", student_id).execute()
        supabase.auth.admin.delete_user(student_id)
        log_activity("delete", "student", student_id, user_id=g.user_id)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@admin_sekolah_bp.route("/students/<student_id>/reset-password", methods=["POST"])
@subscription_write_required
@admin_sekolah_required
def reset_student_password(student_id):
    supabase = get_supabase()
    password = request.form.get("password", "").strip() or _gen_password()
    try:
        supabase.auth.admin.update_user_by_id(student_id, {"password": password})
        log_activity("reset_password", "student", student_id, user_id=g.user_id)
        return jsonify({"success": True, "password": password})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


# ─── Langganan / Subscription ─────────────────────────────────────────

@admin_sekolah_bp.route("/subscription")
@admin_sekolah_required
def subscription():
    supabase = get_supabase()
    from app.services.midtrans_service import get_school_subscription

    school_id = g.user_school_id
    sub = get_school_subscription(school_id) if school_id else None

    plans = []
    try:
        plans = supabase.table("subscription_plans").select("*").eq("is_active", True).order("sort_order").execute().data or []
    except Exception:
        pass

    transactions = []
    try:
        transactions = supabase.table("payment_transactions").select("*, subscription_plans!left(name, duration_label)").eq("school_id", school_id).order("created_at", desc=True).limit(10).execute().data or []
    except Exception:
        pass

    trial_days = 14
    try:
        tr = supabase.table("trial_settings").select("trial_days").limit(1).execute()
        if tr.data:
            trial_days = tr.data[0].get("trial_days", 14)
    except Exception:
        pass

    from app.services.midtrans_service import get_pricing_config, calculate_plan_price, get_student_count_for_school, get_payment_fee_config
    pricing_config = get_pricing_config()
    student_count = get_student_count_for_school(school_id) if school_id else 0
    scaled_active = pricing_config.get("model") == "scaled"
    scaled_tiers = pricing_config.get("tiers", [])
    fee_config = get_payment_fee_config()

    return render_template("admin_sekolah/subscription.html",
        sub=sub, plans=plans, transactions=transactions, trial_days=trial_days,
        pricing_config=pricing_config, student_count=student_count,
        scaled_active=scaled_active, scaled_tiers=scaled_tiers,
        fee_config=fee_config)


@admin_sekolah_bp.route("/subscription/subscribe", methods=["POST"])
@admin_sekolah_required
def subscribe():
    supabase = get_supabase()
    plan_id = request.form.get("plan_id", type=int)
    if not plan_id:
        flash("Pilih plan terlebih dahulu", "error")
        return redirect("/admin-sekolah/subscription")

    school_id = g.user_school_id
    if not school_id:
        flash("Sekolah tidak terdaftar", "error")
        return redirect("/admin-sekolah/subscription")

    # Get school info
    school_name = ""
    admin_email = ""
    try:
        sch = supabase.table("schools").select("name").eq("id", school_id).single().execute()
        if sch.data:
            school_name = sch.data.get("name", "")
        # Email is in Auth, not profiles table
        user_info = supabase.auth.admin.get_user_by_id(g.user_id)
        if user_info and user_info.user:
            admin_email = user_info.user.email or ""
    except Exception:
        pass
    # Ensure email is valid for Midtrans
    if not admin_email or "@" not in admin_email:
        admin_email = "srphysics04@gmail.com"

    from app.services.midtrans_service import create_snap_transaction
    result, error = create_snap_transaction(school_id, plan_id, school_name, admin_email)

    if error:
        flash(error, "error")
        return redirect("/admin-sekolah/subscription")

    settings = {}
    try:
        res = supabase.table("midtrans_settings").select("*").limit(1).execute()
        if res.data:
            settings = res.data[0]
    except Exception:
        pass

    from app.services.midtrans_service import get_payment_fee_config, calculate_total_with_fee
    _total, fee_info = calculate_total_with_fee(result.get("base_amount", result["gross_amount"]))
    base_price = result.get("base_amount", result["gross_amount"])

    return render_template("admin_sekolah/payment.html",
        token=result["token"],
        redirect_url=result["redirect_url"],
        order_id=result["order_id"],
        gross_amount=result["gross_amount"],
        base_price=base_price,
        fee_info=fee_info,
        settings=settings,
    )


@admin_sekolah_bp.route("/payment/success")
@admin_sekolah_required
def payment_success():
    order_id = request.args.get("order_id", "")
    supabase = get_supabase()
    tx = None
    try:
        res = supabase.table("payment_transactions").select("*, subscription_plans!left(name)").eq("order_id", order_id).single().execute()
        tx = res.data
    except Exception:
        pass
    if not tx:
        flash("Transaksi tidak ditemukan", "error")
        return redirect("/admin-sekolah/subscription")
    return render_template("admin_sekolah/payment_success.html", tx=tx)


@admin_sekolah_bp.route("/payment/failure")
@admin_sekolah_required
def payment_failure():
    order_id = request.args.get("order_id", "")
    flash("Pembayaran gagal atau dibatalkan", "error")
    return redirect("/admin-sekolah/subscription")
